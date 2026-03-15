"""
workbook_builder.py
────────────────────
Master orchestrator for Excel generation.
Instantiates sheet writers in the correct order,
then saves the workbook.
Phase 2: Assumption + Cost & Means only.
Remaining sheets added in Phases 3-5.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

from core.session_store import SessionStore
from core.layout_engine  import LayoutEngine
from excel.sheet_assumption    import AssumptionWriter
from excel.sheet_cost_means    import CostMeansWriter
from excel.sheet_revenue       import RevenueWriter
from excel.sheet_manpower      import ManPowerWriter
from excel.sheet_depreciation  import DepreciationWriter
from excel.sheet_expenses      import ExpensesWriter
from excel.sheet_term_loan     import TermLoanWriter
from excel.sheet_wcap          import WCapWriter
from excel.sheet_tax           import TaxWriter
from excel.sheet_pl            import PLWriter
from excel.sheet_bs            import BSWriter
from excel.sheet_cfs_ratio     import CFSWriter, RatioWriter


class WorkbookBuilder:

    def __init__(self, store: SessionStore):
        self.store  = store
        self.layout = LayoutEngine(store)
        self.wb     = Workbook()
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    def build(self, output_path: str) -> str:
        self._build_index()
        self._build_assumption()
        self._build_cost_means()
        self._build_revenue()
        self._build_manpower()
        self._build_depreciation()
        self._build_expenses()
        self._build_term_loan()
        self._build_wcap()
        self._build_tax()
        self._build_pl()
        self._build_bs()
        self._build_cfs()
        self._build_ratio()
        self._set_workbook_properties()
        self.wb.save(output_path)
        return output_path

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _build_index(self):
        ws = self.wb.create_sheet("Index")
        ws.sheet_properties.tabColor = "1F3864"
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 36

        # Title
        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value = self.store.project_profile.company_name
        c.font  = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.merge_cells("A2:C2")
        c = ws["A2"]
        c.value = "DETAILED PROJECT REPORT"
        c.font  = Font(name="Arial", size=10, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16

        ws["B3"].value = "S.No"
        ws["C3"].value = "Sheet Name"
        ws["B3"].font  = Font(name="Arial", size=10, bold=True, color="595959")
        ws["C3"].font  = Font(name="Arial", size=10, bold=True, color="595959")
        ws.row_dimensions[3].height = 14

        sheets = [
            "Assumption", "Cost & Means", "IDC", "Revenue",
            "ManPower", "Depreciation", "Expenses", "Term Loan",
            "W Cap", "Tax", "PL", "BS", "CFS", "Ratio"
        ]
        for i, name in enumerate(sheets, 1):
            r = 3 + i
            ws.cell(row=r, column=2, value=i).font = Font(name="Arial", size=10)
            ws.cell(row=r, column=3, value=name).font = Font(name="Arial", size=10)
            ws.row_dimensions[r].height = 14

    def _build_assumption(self):
        AssumptionWriter(self.wb, self.store, self.layout)

    def _build_cost_means(self):
        CostMeansWriter(self.wb, self.store, self.layout)

    def _build_revenue(self):
        RevenueWriter(self.wb, self.store, self.layout)

    def _build_manpower(self):
        ManPowerWriter(self.wb, self.store, self.layout)

    def _build_depreciation(self):
        DepreciationWriter(self.wb, self.store, self.layout)

    def _build_expenses(self):
        ExpensesWriter(self.wb, self.store, self.layout)

    def _build_term_loan(self):
        TermLoanWriter(self.wb, self.store, self.layout)

    def _build_wcap(self):
        WCapWriter(self.wb, self.store, self.layout)

    def _build_tax(self):
        TaxWriter(self.wb, self.store, self.layout)

    def _build_pl(self):
        PLWriter(self.wb, self.store, self.layout)

    def _build_bs(self):
        BSWriter(self.wb, self.store, self.layout)

    def _build_cfs(self):
        CFSWriter(self.wb, self.store, self.layout)

    def _build_ratio(self):
        RatioWriter(self.wb, self.store, self.layout)

    # ── Workbook properties ───────────────────────────────────────────────────

    def _set_workbook_properties(self):
        """Set workbook-level properties including iterative calculation."""
        self.wb.calculation.iterateCount = 100
        self.wb.calculation.iterateDelta = 0.001

        props = self.wb.properties
        props.creator  = "DPR Agent"
        props.title    = (f"DPR — {self.store.project_profile.company_name}")
        props.subject  = "Detailed Project Report"
        props.keywords = "DPR, MSME, Bank Appraisal, Financial Model"
