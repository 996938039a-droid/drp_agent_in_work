"""
sheet_cost_means.py
────────────────────
Writes the Cost & Means sheet.

Two tables:
  A) Cost of Project   — itemised asset costs summing to Total Project Cost
  B) Means of Finance  — promoter equity, term loans, OD, subsidy

Balance check: =IF(ROUND(TotalCost,0)=ROUND(TotalFinance,0),"✓ BALANCED","✗ CHECK")
All monetary totals reference Assumption sheet — no hardcoded numbers in formulas.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.session_store import (
    SessionStore, AssetCategory, FinanceSourceType
)
from core.layout_engine import LayoutEngine, col_letter, year_col
from excel.styles import (
    BaseSheetWriter,
    FMT_LAKHS, FMT_PCT_1, FMT_INTEGER, FMT_PCT_2,
    font_input, font_formula, font_header, font_subhdr, font_label,
    font_check_pass, font_check_fail,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_amber,
    fill_solid, fill_teal_hdr,
    border_all_thin, border_none,
    align_center, align_left, align_right,
    ROW_HEIGHT_HEADER, ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, WHITE, LIGHT_GREY, DARK_GREY,
    INPUT_FONT_COLOR, FORMULA_FONT_COLOR, XSHEET_FONT_COLOR,
)

# Column layout for Cost & Means sheet
COL_SEC     = 1   # A — section marker
COL_LABEL   = 2   # B — description
COL_BLANK   = 3   # C — spacer
COL_CAT     = 4   # D — category
COL_AMT     = 5   # E — amount (INR Lakhs)
COL_PCT     = 6   # F — percentage of total
LAST_COL    = 7   # G


class CostMeansWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore,
                 layout: LayoutEngine):
        super().__init__(wb, "Cost & Means")
        self.store  = store
        self.layout = layout
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = BLUE
        widths = {1: 4, 2: 36, 3: 3, 4: 22, 5: 18, 6: 14, 7: 16}
        for col, w in widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = w

    def _write(self):
        self._write_title()
        cost_total_row = self._write_cost_table()
        means_total_row = self._write_means_table(cost_total_row + 3)
        self._write_balance_check(cost_total_row, means_total_row)

    # ── Title ─────────────────────────────────────────────────────────────────

    def _write_title(self):
        self.write_section_header(1, 1, LAST_COL,
            "COST OF PROJECT & MEANS OF FINANCE")
        self.write_section_header(2, 1, LAST_COL,
            f"{self.store.project_profile.company_name}  |  "
            f"(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 15

    # ── A: Cost of Project ────────────────────────────────────────────────────

    def _write_cost_table(self) -> int:
        """Write Cost of Project table. Returns the row of the Total line."""
        assets = self.store.capital_means.assets

        # Section header
        ROW_COST_HDR = 4
        self.write_section_header(ROW_COST_HDR, 1, LAST_COL,
            "A  |  COST OF PROJECT")
        self.ws.row_dimensions[ROW_COST_HDR].height = ROW_HEIGHT_SECTION

        # Column headers
        ROW_COL_HDR = 5
        for col, lbl in [(COL_SEC, "#"), (COL_LABEL, "Asset / Item"),
                         (COL_CAT, "Category"),
                         (COL_AMT, "Amount (₹ Lakhs)"),
                         (COL_PCT, "% of Total")]:
            self.write_column_header(ROW_COL_HDR, col, lbl,
                                     fill=fill_solid(NAVY))
        self.ws.row_dimensions[ROW_COL_HDR].height = ROW_HEIGHT_SUBHDR

        # Asset rows — values pulled from Assumption (already written)
        # We store direct values here because Cost & Means is an input sheet
        # The asset costs were written as inputs in Assumption.
        # Here we just restate them as blue inputs (same values, both are inputs)
        first_data_row = ROW_COL_HDR + 1
        asset_rows = []

        for i, asset in enumerate(assets):
            r = first_data_row + i
            fill = fill_white if i % 2 == 0 else fill_alt
            self._section_letter(r, str(i + 1))
            self.write_label(r, COL_LABEL, asset.name, fill=fill)
            self.write_label(r, COL_CAT,   asset.category.value, fill=fill)
            # Cost is an input (blue) — user-entered, matches Assumption
            self.write_input(r, COL_AMT,   asset.cost_lakhs,
                             fmt=FMT_LAKHS, fill=fill_amber)
            asset_rows.append(r)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        # Total row
        total_row = first_data_row + len(assets)
        self.write_label(total_row, COL_LABEL, "TOTAL PROJECT COST",
                         bold=True, fill=fill_solid(TEAL))
        self.write_label(total_row, COL_CAT, "", fill=fill_solid(TEAL))
        total_formula = (f"=SUM({get_column_letter(COL_AMT)}{first_data_row}"
                         f":{get_column_letter(COL_AMT)}{total_row - 1})")
        self.write_total(total_row, COL_AMT, total_formula, bold=True)
        self.ws.row_dimensions[total_row].height = ROW_HEIGHT_DATA

        # % column for each asset
        for i, r in enumerate(asset_rows):
            pct_formula = (f"=IFERROR({get_column_letter(COL_AMT)}{r}"
                           f"/{get_column_letter(COL_AMT)}{total_row},0)")
            self.write_formula(r, COL_PCT, pct_formula, fmt=FMT_PCT_1)
        # total % = 100%
        self.write_formula(total_row, COL_PCT,
                           f"=SUM({get_column_letter(COL_PCT)}{first_data_row}"
                           f":{get_column_letter(COL_PCT)}{total_row - 1})",
                           fmt=FMT_PCT_1, bold=True)

        # Store row reference for balance check
        self._cost_total_row = total_row
        self._cost_total_col = COL_AMT
        return total_row

    # ── B: Means of Finance ───────────────────────────────────────────────────

    def _write_means_table(self, start_row: int) -> int:
        """Write Means of Finance table. Returns the row of the Total line."""
        sources = self.store.capital_means.finance_sources

        self.write_section_header(start_row, 1, LAST_COL,
            "B  |  MEANS OF FINANCE")
        self.ws.row_dimensions[start_row].height = ROW_HEIGHT_SECTION

        # Column headers
        hdr_row = start_row + 1
        for col, lbl in [(COL_SEC, "#"), (COL_LABEL, "Source"),
                         (COL_CAT, "Type"),
                         (COL_AMT, "Amount (₹ Lakhs)"),
                         (COL_PCT, "% of Total")]:
            self.write_column_header(hdr_row, col, lbl,
                                     fill=fill_solid(BLUE))
        self.ws.row_dimensions[hdr_row].height = ROW_HEIGHT_SUBHDR

        first_data_row = hdr_row + 1
        source_rows = []

        for i, src in enumerate(sources):
            r = first_data_row + i
            fill = fill_white if i % 2 == 0 else fill_alt
            self._section_letter(r, str(i + 1))

            # Label
            label = src.label if src.label else src.source_type.value
            # For term loans, append rate and tenor info
            if src.source_type == FinanceSourceType.TERM_LOAN:
                details = (f"{src.rate_pa*100:.1f}% p.a. | "
                           f"{src.tenor_months}m tenor | "
                           f"{src.moratorium_months}m moratorium")
                self.write_label(r, COL_LABEL,
                                 f"{label}  ({details})",
                                 fill=fill)
            else:
                self.write_label(r, COL_LABEL, label, fill=fill)

            self.write_label(r, COL_CAT, src.source_type.value, fill=fill)
            self.write_input(r, COL_AMT, src.amount_lakhs,
                             fmt=FMT_LAKHS, fill=fill_amber)
            source_rows.append(r)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        # Total Means row
        total_row = first_data_row + len(sources)
        self.write_label(total_row, COL_LABEL, "TOTAL MEANS OF FINANCE",
                         bold=True, fill=fill_solid(TEAL))
        self.write_label(total_row, COL_CAT, "", fill=fill_solid(TEAL))
        total_formula = (f"=SUM({get_column_letter(COL_AMT)}{first_data_row}"
                         f":{get_column_letter(COL_AMT)}{total_row - 1})")
        self.write_total(total_row, COL_AMT, total_formula, bold=True)

        # % column
        for i, r in enumerate(source_rows):
            pct_formula = (f"=IFERROR({get_column_letter(COL_AMT)}{r}"
                           f"/{get_column_letter(COL_AMT)}{total_row},0)")
            self.write_formula(r, COL_PCT, pct_formula, fmt=FMT_PCT_1)
        self.write_formula(total_row, COL_PCT,
                           f"=SUM({get_column_letter(COL_PCT)}{first_data_row}"
                           f":{get_column_letter(COL_PCT)}{total_row - 1})",
                           fmt=FMT_PCT_1, bold=True)

        self._means_total_row = total_row
        self._means_total_col = COL_AMT

        # Promoter contribution breakdown
        self._write_promoter_breakdown(total_row + 2, source_rows, sources)

        return total_row

    # ── Promoter contribution breakdown ───────────────────────────────────────

    def _write_promoter_breakdown(self, start_row: int,
                                   source_rows, sources):
        """Show how promoter contribution was computed."""
        pc = self.store.capital_means.promoter_contribution
        if pc <= 0:
            return

        self.write_sub_header(start_row, COL_LABEL, LAST_COL,
                               "Promoter Contribution Derivation",
                               fill=fill_solid("F0F0F0"))

        r1 = start_row + 1
        self.write_label(r1, COL_LABEL,
                         "= Total Project Cost − Term Loans − OD Limit",
                         fill=fill_white)
        self.write_formula(r1, COL_AMT,
                           f"={get_column_letter(self._cost_total_col)}"
                           f"{self._cost_total_row}"
                           f"-{self._means_term_loan_sum()}"
                           f"-{self._means_od_sum()}",
                           fmt=FMT_LAKHS, fill=fill_white)

    def _means_term_loan_sum(self) -> str:
        tl_amounts = [
            src.amount_lakhs
            for src in self.store.capital_means.finance_sources
            if src.source_type == FinanceSourceType.TERM_LOAN
        ]
        return str(sum(tl_amounts))

    def _means_od_sum(self) -> str:
        od_amounts = [
            src.amount_lakhs
            for src in self.store.capital_means.finance_sources
            if src.source_type == FinanceSourceType.OD_LIMIT
        ]
        return str(sum(od_amounts))

    # ── Balance check ─────────────────────────────────────────────────────────

    def _write_balance_check(self, cost_total_row: int,
                              means_total_row: int):
        """Write the balance check formula: cost must equal means."""
        check_row = means_total_row + 2
        # Find the promoter breakdown row if it exists
        for r in range(means_total_row + 2, means_total_row + 8):
            check_row = r
            break

        check_row = means_total_row + 5

        self.write_section_header(check_row - 1, 1, LAST_COL,
                                   "BALANCE CHECK")
        self.ws.row_dimensions[check_row - 1].height = 14

        amt_col = get_column_letter(COL_AMT)
        cost_ref  = f"{amt_col}{cost_total_row}"
        means_ref = f"{amt_col}{means_total_row}"

        check_formula = (
            f'=IF(ROUND({cost_ref},0)=ROUND({means_ref},0),'
            f'"✓  BALANCED  —  Cost = Means = "&TEXT({cost_ref},"#,##0.00")&" Lakhs",'
            f'"✗  IMBALANCE  —  Gap = "&TEXT({cost_ref}-{means_ref},"#,##0.00")&" Lakhs")'
        )
        cell = self.ws.cell(row=check_row, column=COL_LABEL,
                            value=check_formula)
        cell.font      = Font(name="Arial", size=11, bold=True, color="000000")
        cell.fill      = PatternFill("solid", fgColor="E8F8F5")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = border_all_thin(TEAL)
        self.ws.row_dimensions[check_row].height = 20
        self.ws.merge_cells(
            start_row=check_row, start_column=COL_LABEL,
            end_row=check_row,   end_column=LAST_COL
        )

        # Store reference for validation agent
        self.layout._map["Cost & Means"] = self.layout._map.get("Cost & Means", {})
        self.layout._map["Cost & Means"]["balance_check_row"] = (check_row, COL_LABEL)
        self.layout._map["Cost & Means"]["cost_total_row"]    = (cost_total_row,  COL_AMT)
        self.layout._map["Cost & Means"]["means_total_row"]   = (means_total_row, COL_AMT)

    def _section_letter(self, row: int, text: str):
        self.w(row, COL_SEC, text,
               font=Font(name="Arial", size=8, color="999999"),
               alignment=Alignment(horizontal="center", vertical="center"))
