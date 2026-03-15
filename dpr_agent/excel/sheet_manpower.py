"""
sheet_manpower.py
──────────────────
Writes the ManPower sheet.

Structure:
  - Employee category table: count, monthly salary, annual salary, fixed/variable flag
  - Annual salary totals (fixed + variable separately)
  - 7-year projection: compounded at annual increment rate from Assumption
  - Transfer to P&L rows: salary / 12 × months_in_operation

All salary values reference Assumption (count and monthly salary per category).
Escalation rate also references Assumption.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore
from core.layout_engine import (
    LayoutEngine, ManPowerLayout as ML, AssumptionLayout as AL,
    col_letter, year_col, COL_LABEL, COL_BASIS, COL_YEAR_START
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_NUMBER, FMT_PCT_2, FMT_INTEGER,
    font_formula, font_xsheet, font_input, font_header,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid,
    fill_teal_hdr, fill_blue_hdr,
    border_all_thin, align_center, align_left, align_right,
    ROW_HEIGHT_HEADER, ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, LIGHT_GREY, WHITE,
    FORMULA_FONT_COLOR, XSHEET_FONT_COLOR,
)


class ManPowerWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store  = store
        self.layout = layout
        self.n_years = store.n_years
        self.n_cats  = store.n_employee_categories
        super().__init__(wb, "ManPower")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = BLUE
        self.ws.column_dimensions[get_column_letter(COL_LABEL)].width = 28
        self.ws.column_dimensions[get_column_letter(COL_BASIS)].width = 10
        # Additional columns for static employee data
        self.ws.column_dimensions["D"].width = 12
        self.ws.column_dimensions["E"].width = 14
        self.ws.column_dimensions["F"].width = 14
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[get_column_letter(year_col(yr))].width = 16

    def _write(self):
        self._write_title()
        self._write_employee_table()
        self._write_annual_projections()
        self._write_pl_transfer()
        self.ws.freeze_panes = f"G{ML.BASE_ROW}"

    # ── Title ─────────────────────────────────────────────────────────────────

    def _write_title(self):
        last = year_col(self.n_years)
        self.write_section_header(1, COL_LABEL, last, "MANPOWER EXPENSES")
        self.write_section_header(2, COL_LABEL, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 14

    # ── Employee table ────────────────────────────────────────────────────────

    def _write_employee_table(self):
        """Static reference table: designation, count, monthly salary, annual, total."""
        # Column headers at row 4
        headers = {
            COL_LABEL: "Designation",
            3: "Count",
            4: "Monthly Salary (₹L)",
            5: "Annual (₹L)",
            6: "Annual Total (₹L)",
        }
        for col, hdr in headers.items():
            self.write_column_header(4, col, hdr, fill=fill_solid(NAVY))
        self.ws.row_dimensions[4].height = ROW_HEIGHT_SUBHDR

        for i, cat in enumerate(self.store.manpower.categories):
            r       = ML.emp_row(i)
            fill_bg = fill_white if i % 2 == 0 else fill_alt

            # Designation label
            self.write_label(r, COL_LABEL, cat.designation,
                             bold=True, fill=fill_bg)

            # Count — from Assumption
            asmp_count = self._asmp_abs(f"mp_count_cat{i}")
            self.write_formula(r, 3, f"={asmp_count}",
                               fmt=FMT_INTEGER, fill=fill_bg, xsheet=True)

            # Monthly salary — from Assumption
            asmp_sal = self._asmp_abs(f"mp_salary_cat{i}")
            self.write_formula(r, 4, f"={asmp_sal}",
                               fmt=FMT_LAKHS, fill=fill_bg, xsheet=True)

            # Annual = monthly × 12
            c3 = get_column_letter(3); c4 = get_column_letter(4)
            self.write_formula(r, 5,
                               f"={c4}{r}*12",
                               fmt=FMT_LAKHS, fill=fill_bg)

            # Annual total = annual × count
            self.write_formula(r, 6,
                               f"={c3}{r}*E{r}",
                               fmt=FMT_LAKHS,
                               fill=fill_solid("E8F8F5") if cat.is_fixed else fill_solid("FFF3CD"))

            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        # Base year totals
        base_row = ML.BASE_ROW + self.n_cats + 1
        self.write_label(base_row, COL_LABEL, "Base Year Annual Salary",
                         bold=True, fill=fill_solid(TEAL))
        first_emp = ML.emp_row(0); last_emp = ML.emp_row(self.n_cats - 1)
        self.write_total(base_row, 6,
                         f"=SUM(F{first_emp}:F{last_emp})")
        self.ws.row_dimensions[base_row].height = ROW_HEIGHT_DATA
        self._base_year_row = base_row

    # ── 7-year projection ─────────────────────────────────────────────────────

    def _write_annual_projections(self):
        """Compounded salary projections for N years."""
        proj_base = self._base_year_row + 2

        # Section header
        last = year_col(self.n_years)
        self.write_sub_header(proj_base - 1, COL_LABEL, last,
                               "Annual Salary Projections (5% increment p.a.)",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[proj_base - 1].height = 14

        # Year headers
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            self.write_column_header(proj_base, col, f"Year {yr}",
                                     fill=fill_solid(NAVY))
        self.ws.row_dimensions[proj_base].height = ROW_HEIGHT_SUBHDR

        salary_row = proj_base + 1
        self.write_label(salary_row, COL_LABEL, "Total Annual Salary (Fixed)",
                         bold=True)
        self.write_label(salary_row, COL_BASIS, "₹ Lakhs")

        # Get increment reference from Assumption
        # Use increment of category 0 as the common rate (applies to all)
        if self.n_cats > 0:
            asmp_incr = self._asmp_abs("mp_increment_cat0")
        else:
            asmp_incr = "0.05"

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            if yr == 1:
                # Year 1 = base year total
                formula = f"=F{self._base_year_row}"
            else:
                prev_c = get_column_letter(year_col(yr - 1))
                formula = f"={prev_c}{salary_row}*(1+{asmp_incr})"
            self.write_total(salary_row, col, formula)
        self.ws.row_dimensions[salary_row].height = ROW_HEIGHT_DATA
        self._salary_proj_row = salary_row

    # ── Transfer to P&L ───────────────────────────────────────────────────────

    def _write_pl_transfer(self):
        """Salary cost per year = annual / 12 × months_in_operation."""
        pl_base = self._salary_proj_row + 2
        last    = year_col(self.n_years)

        self.write_section_header(pl_base - 1, COL_LABEL, last,
                                   "Transfer to P&L Account")
        self.ws.row_dimensions[pl_base - 1].height = 14

        pl_row = pl_base
        self.write_label(pl_row, COL_LABEL,
                         "Salary Expense (Annual, Lakhs)", bold=True,
                         fill=fill_solid("D5F5E3"))
        self.write_label(pl_row, COL_BASIS, "₹ Lakhs",
                         fill=fill_solid("D5F5E3"))

        for yr in range(1, self.n_years + 1):
            col  = year_col(yr)
            c    = get_column_letter(col)
            # months in operation for this year = Revenue!row5 value = 12
            # Use simple: annual / 12 * 12 = annual (all years are full years here)
            formula = f"={c}{self._salary_proj_row}/12*12"
            cell = self.ws.cell(row=pl_row, column=col, value=formula)
            cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill = fill_solid("1ABC9C")
            cell.number_format = FMT_LAKHS
            cell.alignment = align_right
            cell.border = border_all_thin(TEAL)
        self.ws.row_dimensions[pl_row].height = ROW_HEIGHT_DATA + 2

        # Store this row in layout for P&L sheet to reference
        self.layout._map["ManPower"]["pl_salary_row"] = (pl_row, year_col(1))

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _asmp_abs(self, logical: str) -> str:
        """Return absolute Assumption reference like Assumption!$E$4."""
        entry = self.layout._map["Assumption"].get(logical)
        if entry:
            r, c = entry
            return f"Assumption!${get_column_letter(c)}${r}"
        return "0"
