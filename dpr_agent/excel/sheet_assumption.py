"""
sheet_assumption.py
────────────────────
Writes the Assumption sheet — the SINGLE SOURCE OF TRUTH for every
parameter in the DPR model.

Design rules:
  1. Every user-entered value → BLUE font (input cell)
  2. Zero derived formulas in this sheet — pure inputs only
  3. Organized into 8 labeled sections with section headers
  4. Assumption cells are referenced by EVERY other sheet via absolute refs
  5. No other sheet has hardcoded constants — everything comes from here
"""

from openpyxl import Workbook
from core.session_store import (
    SessionStore, AssetCategory, EntityType, FinanceSourceType
)
from core.layout_engine import LayoutEngine, col_letter
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_PCT_2, FMT_INTEGER, FMT_TEXT,
    FMT_NUMBER, FMT_PCT_1, FMT_DATE, FMT_ZERO_DASH,
    font_input, font_formula, font_header, font_subhdr, font_label,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_amber,
    fill_solid, fill_teal_hdr, fill_blue_hdr,
    border_all_thin, border_none,
    align_center, align_left, align_right, align_wrap_l,
    ROW_HEIGHT_HEADER, ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, WHITE, LIGHT_GREY, AMBER, DARK_GREY,
    INPUT_FONT_COLOR, FORMULA_FONT_COLOR,
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Column layout for Assumption sheet
COL_SECTION = 1   # A — section letter
COL_LABEL   = 2   # B — parameter label
COL_BLANK   = 3   # C — spacer
COL_UNIT    = 4   # D — unit description
COL_VALUE   = 5   # E — VALUE (the key input cell)
COL_NOTES   = 6   # F — notes / escalation / secondary param

LAST_COL    = 8   # H


class AssumptionWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore,
                 layout: LayoutEngine):
        super().__init__(wb, "Assumption")
        self.store  = store
        self.layout = layout
        self._write()

    # ── Sheet setup ───────────────────────────────────────────────────────────

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = NAVY
        # Column widths
        widths = {1: 4, 2: 38, 3: 3, 4: 18, 5: 16, 6: 16, 7: 16, 8: 22}
        for col, w in widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = w

    # ── Main write method ─────────────────────────────────────────────────────

    def _write(self):
        s = self.store
        L = self.layout

        # Title
        self._write_title()

        # Section A — Capacity
        self._write_capacity()

        # Section B — Revenue (one sub-block per product)
        self._write_revenue()

        # Section C — Raw Materials (one sub-block per material)
        self._write_raw_materials()

        # Section D — Operating Expenses
        self._write_expenses()

        # Section E — Manpower
        self._write_manpower()

        # Section F — Finance (one sub-block per loan)
        self._write_finance()

        # Section G — Working Capital
        self._write_working_capital()

        # Section H — Depreciation
        self._write_depreciation()

        # Section I — Implementation Schedule
        self._write_implementation()

        # Freeze panes: freeze rows 1-2 and column A-B
        self.ws.freeze_panes = "E4"

    # ── Title ────────────────────────────────────────────────────────────────

    def _write_title(self):
        self.write_section_header(1, 1, LAST_COL,
            f"ASSUMPTIONS & INPUTS  —  {self.store.project_profile.company_name}")
        self.ws.row_dimensions[1].height = 22

        self.write_section_header(2, 1, LAST_COL,
            "(All monetary values in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[2].height = 15

        # Legend row
        self.w(3, 2, "LEGEND:", font=font_label(bold=True),
               alignment=align_left, height=14)
        self.w(3, 3, "Blue = User Input",
               font=Font(name="Arial", size=9, color=INPUT_FONT_COLOR, bold=True),
               alignment=align_left)
        self.w(3, 5, "Yellow bg = Key Assumption",
               font=Font(name="Arial", size=9, color="806000", bold=False),
               fill=fill_amber, alignment=align_center)

    # ── Section A: Capacity ───────────────────────────────────────────────────

    def _write_capacity(self):
        s = self.store
        rm = s.revenue_model

        row = self.layout.row("Assumption", "cap_year1_util") - 1
        self.write_section_header(row, 1, LAST_COL, "A  |  CAPACITY PARAMETERS")

        self._param_row(
            self.layout.row("Assumption", "cap_year1_util"),
            "A1", "Year 1 Capacity Utilisation",
            rm.year1_utilization, "fraction (0–1)",
            fmt=FMT_PCT_2, is_key=True
        )
        self._param_row(
            self.layout.row("Assumption", "cap_annual_increment"),
            "A2", "Annual Utilisation Increment",
            rm.annual_utilization_increment, "fraction p.a.",
            fmt=FMT_PCT_2
        )
        self._param_row(
            self.layout.row("Assumption", "cap_max_util"),
            "A3", "Maximum Utilisation Ceiling",
            rm.max_utilization, "fraction (0–1)",
            fmt=FMT_PCT_2
        )
        self._param_row(
            self.layout.row("Assumption", "cap_working_days"),
            "A4", "Working Days per Month",
            rm.working_days_per_month, "days",
            fmt=FMT_INTEGER
        )
        self._param_row(
            self.layout.row("Assumption", "cap_months_year"),
            "A5", "Months in a Year",
            12, "months",
            fmt=FMT_INTEGER
        )

    # ── Section B: Revenue ────────────────────────────────────────────────────

    def _write_revenue(self):
        from core.layout_engine import AssumptionLayout as AL
        s  = self.store
        products = s.revenue_model.products

        # Section header row (just above first product)
        hdr_row = AL.HDR_REVENUE
        self.write_section_header(hdr_row, 1, LAST_COL,
            "B  |  REVENUE PARAMETERS  (one block per product / service)")
        self.ws.row_dimensions[hdr_row].height = ROW_HEIGHT_SECTION

        # Column sub-headers for this section
        shdr = hdr_row + 1
        self._col_headers(shdr,
            ["", "Product / Service Name", "", "Unit",
             "Base Price (Yr 1)", "Annual Price Escalation"])
        self.ws.row_dimensions[shdr].height = ROW_HEIGHT_SUBHDR

        for i, prod in enumerate(products):
            price_row = AL.rev_price_row(i)
            esc_row   = AL.rev_escalation_row(i)

            # Update layout map with correct column
            self.layout._map["Assumption"][f"rev_price_p{i}"]      = (price_row, COL_VALUE)
            self.layout._map["Assumption"][f"rev_escalation_p{i}"] = (esc_row,   COL_VALUE)

            self._section_letter(price_row, f"B{i+1}")
            self.write_label(price_row, COL_LABEL, prod.name,
                             bold=True, fill=fill_solid("EAF4FB"))
            self.write_label(price_row, COL_UNIT,  prod.unit,
                             fill=fill_solid("EAF4FB"))
            self.write_input(price_row, COL_VALUE, prod.price_per_unit,
                             fmt=FMT_NUMBER, fill=fill_amber)

            # Escalation row
            self._section_letter(esc_row, "")
            self.write_label(esc_row, COL_LABEL, "  Annual Price Escalation",
                             fill=fill_alt)
            self.write_label(esc_row, COL_UNIT,  "fraction p.a.", fill=fill_alt)
            self.write_input(esc_row, COL_VALUE, prod.price_escalation_pa,
                             fmt=FMT_PCT_2, fill=fill_alt)

            # Blank separator row
            if i < len(products) - 1:
                sep = esc_row + 1
                self.ws.row_dimensions[sep].height = 4

    # ── Section C: Raw Materials ──────────────────────────────────────────────

    def _write_raw_materials(self):
        from core.layout_engine import AssumptionLayout as AL
        s         = self.store
        materials = s.cost_structure.raw_materials
        n_prods   = s.n_products

        hdr_row = AL.HDR_RAWMATERIAL
        self.write_section_header(hdr_row, 1, LAST_COL,
            "C  |  RAW MATERIAL PARAMETERS  (one block per input material)")

        shdr = hdr_row + 1
        self._col_headers(shdr,
            ["", "Material Name", "", "Unit",
             "Base Price (Yr 1)", "Annual Cost Escalation"])

        for i, mat in enumerate(materials):
            price_row = AL.rm_price_row(i, n_prods)
            esc_row   = price_row + 1
            ipu_row   = esc_row + 1

            # Keep layout map in sync with actual written rows
            self.layout._map["Assumption"][f"rm_price_m{i}"]           = (price_row, COL_VALUE)
            self.layout._map["Assumption"][f"rm_escalation_m{i}"]      = (esc_row,   COL_VALUE)
            self.layout._map["Assumption"][f"rm_input_per_output_m{i}"]= (ipu_row,   COL_VALUE)

            self._section_letter(price_row, f"C{i+1}")
            self.write_label(price_row, COL_LABEL, mat.name,
                             bold=True, fill=fill_solid("EAF4FB"))
            self.write_label(price_row, COL_UNIT,  f"per {mat.unit}",
                             fill=fill_solid("EAF4FB"))
            self.write_input(price_row, COL_VALUE, mat.price_per_unit,
                             fmt=FMT_NUMBER, fill=fill_amber)

            self._section_letter(esc_row, "")
            self.write_label(esc_row, COL_LABEL, "  Annual Price Escalation",
                             fill=fill_alt)
            self.write_label(esc_row, COL_UNIT,  "fraction p.a.", fill=fill_alt)
            self.write_input(esc_row, COL_VALUE, mat.price_escalation_pa,
                             fmt=FMT_PCT_2, fill=fill_alt)

            # Input-per-output unit
            self._section_letter(ipu_row, "")
            self.write_label(ipu_row, COL_LABEL,
                             f"  Input per Output Unit ({mat.unit} per {self.store.revenue_model.products[0].unit if self.store.revenue_model.products else 'unit'})",
                             fill=fill_white)
            self.write_label(ipu_row, COL_UNIT,
                             f"{mat.unit} / output unit", fill=fill_white)
            self.write_input(ipu_row, COL_VALUE, mat.input_per_output_unit,
                             fmt=FMT_NUMBER, fill=fill_amber)

    # ── Section D: Operating Expenses ─────────────────────────────────────────

    def _write_expenses(self):
        from core.layout_engine import AssumptionLayout as AL
        cs  = self.store.cost_structure
        hdr = AL.HDR_EXPENSES

        self.write_section_header(hdr, 1, LAST_COL,
            "D  |  OPERATING EXPENSE PARAMETERS")

        shdr = hdr + 1
        self._col_headers(shdr,
            ["", "Expense Item", "", "Basis",
             "Rate / Base Amount", "Annual Escalation"])

        def exp_row(logical, label, value, unit, fmt, esc_logical, esc_value,
                    is_key=False):
            r = self.layout.row("Assumption", logical)
            # Ensure map has correct column (COL_VALUE)
            self.layout._map["Assumption"][logical] = (r, COL_VALUE)
            self._section_letter(r, "D")
            self.write_label(r, COL_LABEL, label, fill=fill_white if not is_key else fill_amber)
            self.write_label(r, COL_UNIT,  unit,  fill=fill_white)
            self.write_input(r, COL_VALUE, value, fmt=fmt,
                             fill=fill_amber if is_key else fill_white)
            if esc_logical:
                r2 = self.layout.row("Assumption", esc_logical)
                self.layout._map["Assumption"][esc_logical] = (r2, COL_VALUE)
                self.write_label(r2, COL_LABEL, f"  └ Escalation rate", fill=fill_alt)
                self.write_label(r2, COL_UNIT,  "fraction p.a.", fill=fill_alt)
                self.write_input(r2, COL_VALUE, esc_value, fmt=FMT_PCT_2, fill=fill_alt)

        exp_row("exp_rm_pct_fa",   "Repair & Maintenance",  cs.rm_pct_of_fa,
                "% of Net Fixed Assets", FMT_PCT_2,
                "exp_rm_escalation", cs.rm_escalation_pa, is_key=True)

        exp_row("exp_ins_pct_fa",  "Insurance",             cs.insurance_pct_of_fa,
                "% of Net Fixed Assets", FMT_PCT_2,
                "exp_ins_escalation", cs.insurance_escalation_pa)

        exp_row("exp_power_pct_rev","Power & Fuel",         cs.power_pct_revenue,
                "% of Revenue", FMT_PCT_2,
                "exp_power_escalation", cs.power_escalation_pa, is_key=True)

        exp_row("exp_mkt_pct_rev", "Marketing Expenses",   cs.marketing_pct_revenue,
                "% of Revenue", FMT_PCT_2,
                "exp_mkt_escalation", cs.marketing_escalation_pa)

        exp_row("exp_transport_base","Transportation Cost (Base)",
                cs.transport_base_lakhs, "INR Lakhs (Year 1)", FMT_LAKHS,
                "exp_transport_esc", cs.transport_escalation_pa)

        exp_row("exp_misc_base",   "Miscellaneous Expenses (Base)",
                cs.misc_base_lakhs, "INR Lakhs (Year 1)", FMT_LAKHS,
                "exp_misc_esc", cs.misc_escalation_pa)

        exp_row("exp_sga_base",    "Selling, General & Admin (Base)",
                cs.sga_base_lakhs, "INR Lakhs (Year 1)", FMT_LAKHS,
                "exp_sga_esc", cs.sga_escalation_pa)

    # ── Section E: Manpower ───────────────────────────────────────────────────

    def _write_manpower(self):
        from core.layout_engine import AssumptionLayout as AL
        cats = self.store.manpower.categories
        hdr  = AL.HDR_MANPOWER

        self.write_section_header(hdr, 1, LAST_COL,
            "E  |  MANPOWER PARAMETERS")

        shdr = hdr + 1
        self._col_headers(shdr,
            ["", "Designation", "", "Head Count",
             "Monthly Salary (Lakhs)", "Annual Increment"])
        self.ws.row_dimensions[shdr].height = ROW_HEIGHT_SUBHDR

        for i, cat in enumerate(cats):
            r = AL.mp_count_row(i)
            self._section_letter(r, f"E{i+1}")
            self.write_label(r, COL_LABEL, cat.designation, bold=True,
                             fill=fill_solid("EAF4FB") if i % 2 == 0 else fill_alt)
            self.write_input(r, COL_UNIT,  cat.count, fmt=FMT_INTEGER,
                             fill=fill_solid("EAF4FB") if i % 2 == 0 else fill_alt)
            self.write_input(r, COL_VALUE, cat.monthly_salary_lakhs,
                             fmt=FMT_LAKHS, fill=fill_amber)
            self.write_input(r, COL_NOTES, cat.annual_increment_pa,
                             fmt=FMT_PCT_2,
                             fill=fill_solid("EAF4FB") if i % 2 == 0 else fill_alt)
            # Keep layout map aligned
            self.layout._map["Assumption"][f"mp_count_cat{i}"] = (r, COL_UNIT)
            self.layout._map["Assumption"][f"mp_salary_cat{i}"] = (r, COL_VALUE)
            self.layout._map["Assumption"][f"mp_increment_cat{i}"] = (r, COL_NOTES)

    # ── Section F: Finance ────────────────────────────────────────────────────

    def _write_finance(self):
        from core.layout_engine import AssumptionLayout as AL
        s       = self.store
        loans   = s.capital_means.term_loans
        od_list = s.capital_means.od_sources
        hdr     = AL.HDR_FINANCE

        self.write_section_header(hdr, 1, LAST_COL,
            "F  |  FINANCE PARAMETERS")

        shdr = hdr + 1
        self._col_headers(shdr,
            ["", "Parameter", "", "Description",
             "Value", ""])

        for i, loan in enumerate(loans):
            ar = AL.fin_amount_row(i)
            rr = AL.fin_rate_row(i)
            tr = AL.fin_tenor_row(i)
            mr = AL.fin_moratorium_row(i)

            lbl = loan.label or f"Term Loan {i+1}"
            self.write_section_header(ar - 1, COL_LABEL, COL_NOTES,
                f"  {lbl}")
            self.ws.row_dimensions[ar - 1].height = 14

            self._fin_param(ar, f"F{i+1}a", "Loan Amount",
                            loan.amount_lakhs, "INR Lakhs")
            self._fin_param(rr, f"F{i+1}b", "Annual Interest Rate",
                            loan.rate_pa, "fraction p.a.", fmt=FMT_PCT_2,
                            is_key=True)
            self._fin_param(tr, f"F{i+1}c", "Total Tenor",
                            loan.tenor_months, "months", fmt=FMT_INTEGER)
            self._fin_param(mr, f"F{i+1}d", "Moratorium Period",
                            loan.moratorium_months, "months", fmt=FMT_INTEGER)

            # Derived: repayment months (formula, not input)
            rep_row = mr + 1
            self.write_label(rep_row, COL_LABEL, "  Repayment Months (derived)",
                             fill=fill_alt)
            self.write_label(rep_row, COL_UNIT,  "months", fill=fill_alt)
            self.write_formula(
                rep_row, COL_VALUE,
                f"={get_column_letter(COL_VALUE)}{tr}"
                f"-{get_column_letter(COL_VALUE)}{mr}",
                fmt=FMT_INTEGER, fill=fill_alt
            )
            self.layout._map["Assumption"][f"fin_repayment_months_l{i}"] = (
                rep_row, COL_VALUE
            )

        # OD / CC Limit
        od_hdr_row = AL.FIN_OD_LIMIT_ROW - 1
        self.write_section_header(od_hdr_row, COL_LABEL, COL_NOTES,
            "  Overdraft / Working Capital Limit")
        self.ws.row_dimensions[od_hdr_row].height = 14

        od_amount = od_list[0].amount_lakhs if od_list else 0.0
        od_rate   = od_list[0].rate_pa      if od_list else 0.0

        self._fin_param(AL.FIN_OD_LIMIT_ROW, "F_OD1", "OD / CC Limit",
                        od_amount, "INR Lakhs")
        self._fin_param(AL.FIN_OD_RATE_ROW,  "F_OD2", "OD Interest Rate",
                        od_rate, "fraction p.a.", fmt=FMT_PCT_2)

    # ── Section G: Working Capital ─────────────────────────────────────────────

    def _write_working_capital(self):
        from core.layout_engine import AssumptionLayout as AL
        wc  = self.store.finance_wc
        hdr = AL.HDR_WORKING_CAPITAL

        self.write_section_header(hdr, 1, LAST_COL,
            "G  |  WORKING CAPITAL PARAMETERS")

        params = [
            ("wc_debtor_days",    "G1", "Debtor Days (Receivables)",
             wc.debtor_days,       "days", FMT_INTEGER),
            ("wc_creditor_rm",    "G2", "Creditor Days — Raw Materials",
             wc.creditor_days_rm,  "days", FMT_INTEGER),
            ("wc_creditor_admin", "G3", "Creditor Days — Admin Expenses",
             wc.creditor_days_admin,"days", FMT_INTEGER),
            ("wc_stock_rm",       "G4", "Raw Material Stock Days",
             wc.stock_days_rm,    "days", FMT_INTEGER),
            ("wc_stock_fg",       "G5", "Finished Goods Stock Days",
             wc.stock_days_fg,    "days", FMT_INTEGER),
            ("wc_loan_amount",    "G6", "Working Capital Loan",
             wc.wc_loan_amount,   "INR Lakhs", FMT_LAKHS),
            ("wc_interest_rate",  "G7", "WC Loan Interest Rate",
             wc.wc_interest_rate, "fraction p.a.", FMT_PCT_2),
        ]
        for logical, sec, label, value, unit, fmt in params:
            r = self.layout.row("Assumption", logical)
            self._section_letter(r, sec)
            self.write_label(r, COL_LABEL, label, fill=fill_white)
            self.write_label(r, COL_UNIT,  unit,  fill=fill_white)
            self.write_input(r, COL_VALUE, value, fmt=fmt,
                             fill=fill_amber if "days" in unit else fill_white)

    # ── Section H: Depreciation ────────────────────────────────────────────────

    def _write_depreciation(self):
        from core.layout_engine import AssumptionLayout as AL
        dr  = self.store.depreciation_rates
        hdr = AL.HDR_DEPRECIATION

        self.write_section_header(hdr, 1, LAST_COL,
            "H  |  DEPRECIATION RATES  (Income Tax Act — WDV Method)")

        params = [
            ("dep_pm_rate",    "H1", "Plant & Machinery",    dr.plant_machinery),
            ("dep_civil_rate", "H2", "Civil Works / Building",dr.civil_works),
            ("dep_furn_rate",  "H3", "Furniture & Fixtures",  dr.furniture),
            ("dep_veh_rate",   "H4", "Vehicles",               dr.vehicle),
            ("dep_elec_rate",  "H5", "Electrical & Fittings",  dr.electrical),
            ("dep_preop_rate", "H6", "Pre-operative Assets",   dr.pre_operative),
        ]
        for logical, sec, label, value in params:
            r = self.layout.row("Assumption", logical)
            self._section_letter(r, sec)
            self.write_label(r, COL_LABEL, label, fill=fill_white)
            self.write_label(r, COL_UNIT,  "WDV % p.a.", fill=fill_white)
            self.write_input(r, COL_VALUE, value, fmt=FMT_PCT_2)

    # ── Section I: Implementation Schedule ────────────────────────────────────

    def _write_implementation(self):
        from core.layout_engine import AssumptionLayout as AL
        fw  = self.store.finance_wc
        hdr = AL.HDR_IMPL_SCHEDULE

        self.write_section_header(hdr, 1, LAST_COL,
            "I  |  IMPLEMENTATION SCHEDULE")

        r = self.layout.row("Assumption", "impl_months")
        self._section_letter(r, "I1")
        self.write_label(r, COL_LABEL,
                         "Implementation Period (months before COD)",
                         fill=fill_white)
        self.write_label(r, COL_UNIT,  "months", fill=fill_white)
        self.write_input(r, COL_VALUE, fw.implementation_months,
                         fmt=FMT_INTEGER)

        # Tax entity type (referenced by Tax sheet)
        te_row = r + 2
        self.write_section_header(te_row - 1, 1, LAST_COL,
            "J  |  TAX PARAMETERS")
        self._section_letter(te_row, "J1")
        self.write_label(te_row, COL_LABEL, "Entity Type", fill=fill_white)
        self.write_label(te_row, COL_UNIT,  "type", fill=fill_white)
        self.write_input(te_row, COL_VALUE,
                         self.store.project_profile.entity_type.value,
                         fmt=FMT_TEXT, fill=fill_amber)
        self.layout._map["Assumption"]["entity_type"] = (te_row, COL_VALUE)

    # ── Private helpers ───────────────────────────────────────────────────────

    def _param_row(self, row: int, sec_label: str, label: str,
                   value, unit: str, fmt=FMT_LAKHS,
                   is_key=False, notes: str = ""):
        """Write one standard parameter row in Assumption."""
        fill_v = fill_amber if is_key else fill_white
        fill_l = fill_white

        self._section_letter(row, sec_label)
        self.write_label(row, COL_LABEL, label,     fill=fill_l, bold=is_key)
        self.write_label(row, COL_UNIT,  unit,       fill=fill_l)
        self.write_input(row, COL_VALUE, value,
                         fmt=fmt, fill=fill_v)
        if notes:
            self.write_label(row, COL_NOTES, notes,
                             fill=fill_solid("FFFDE7"))

    def _fin_param(self, row: int, sec: str, label: str,
                   value, unit: str, fmt=FMT_LAKHS, is_key=False):
        fill_v = fill_amber if is_key else fill_white
        self._section_letter(row, sec)
        self.write_label(row, COL_LABEL, label, fill=fill_white)
        self.write_label(row, COL_UNIT,  unit,  fill=fill_white)
        self.write_input(row, COL_VALUE, value, fmt=fmt, fill=fill_v)

    def _section_letter(self, row: int, letter: str):
        """Write small section-letter identifier in column A."""
        self.w(row, COL_SECTION, letter,
               font=Font(name="Arial", size=8, color="999999"),
               alignment=Alignment(horizontal="center", vertical="center"),
               height=ROW_HEIGHT_DATA)

    def _col_headers(self, row: int, labels: list):
        """Write column sub-header labels."""
        for i, lbl in enumerate(labels):
            col = 1 + i
            if lbl:
                self.write_column_header(row, col, lbl,
                                         fill=fill_solid("2E75B6"))
        self.ws.row_dimensions[row].height = ROW_HEIGHT_SUBHDR
