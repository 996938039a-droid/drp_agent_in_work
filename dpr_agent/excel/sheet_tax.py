"""
sheet_tax.py
─────────────
Writes the Tax sheet.

Contains tax computation tables for three entity types:
  1. Individual / Sole Proprietor — progressive slab rates
  2. Company — 30% + surcharge + 4% HEC
  3. Partnership Firm — 30% flat + surcharge + 4% HEC

The active entity type is read from Assumption!entity_type cell.
P&L pulls tax using an IF formula that selects the right table row.

Taxable income is pulled from PL!PBT row, converted to absolute INR.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore, EntityType
from core.layout_engine import (
    LayoutEngine, TaxLayout as TxL, AssumptionLayout as AL,
    col_letter, year_col, COL_LABEL
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_NUMBER, FMT_PCT_2, FMT_INTEGER,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid, fill_amber,
    border_all_thin, align_center, align_left, align_right,
    ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, WHITE, LIGHT_GREY,
)

# Tax computation column layout
TC_DATA_COL_START = 3   # C = Year 1


class TaxWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years = store.n_years
        self._entity = store.project_profile.entity_type
        super().__init__(wb, "Tax")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = NAVY
        self.ws.column_dimensions[get_column_letter(COL_LABEL)].width = 34
        self.ws.column_dimensions[get_column_letter(3)].width = 14
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[
                get_column_letter(TC_DATA_COL_START + yr - 1)].width = 14

    def _write(self):
        self._write_title()
        self._write_entity_status()
        self._write_slab_reference()
        self._write_tax_computation()
        self._register_layout()

    # ── Title ─────────────────────────────────────────────────────────────────

    def _write_title(self):
        last = TC_DATA_COL_START + self.n_years
        self.write_section_header(1, COL_LABEL, last,
            "CALCULATION OF CURRENT TAX")
        self.write_section_header(2, COL_LABEL, last,
            f"{self.store.project_profile.company_name}  |  "
            "All tax rates as per Income Tax Act")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 14

    # ── Entity status ─────────────────────────────────────────────────────────

    def _write_entity_status(self):
        self.write_label(TxL.STATUS_ROW, COL_LABEL,
                         "Entity Type (from Assumption):",
                         bold=True, fill=fill_alt)
        asmp_entity = self._asmp_abs("entity_type")
        self.write_formula(TxL.STATUS_ROW, TC_DATA_COL_START,
                           f"={asmp_entity}",
                           fmt="@", fill=fill_amber, xsheet=True)
        self.ws.row_dimensions[TxL.STATUS_ROW].height = ROW_HEIGHT_DATA

    # ── Tax slab reference table (static) ────────────────────────────────────

    def _write_slab_reference(self):
        last = TC_DATA_COL_START + 2
        self.write_sub_header(TxL.HDR_ROW - 1, COL_LABEL, last,
                               "TAX RATE REFERENCE",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[TxL.HDR_ROW - 1].height = 13

        tc = self.store.tax_config

        # Company
        self.write_label(TxL.HDR_ROW, COL_LABEL,
                         "Company (Pvt Ltd)", bold=True, fill=fill_alt)
        self.write_label(TxL.HDR_ROW + 1, COL_LABEL,
                         "  Basic Tax Rate", fill=fill_white)
        self.write_input(TxL.HDR_ROW + 1, TC_DATA_COL_START,
                         tc.company_basic_rate, fmt=FMT_PCT_2)
        self.write_label(TxL.HDR_ROW + 2, COL_LABEL,
                         "  Health & Education Cess", fill=fill_alt)
        self.write_input(TxL.HDR_ROW + 2, TC_DATA_COL_START,
                         tc.hec_rate, fmt=FMT_PCT_2)
        self.write_label(TxL.HDR_ROW + 3, COL_LABEL,
                         "  Surcharge (₹1Cr–₹10Cr)", fill=fill_white)
        self.write_input(TxL.HDR_ROW + 3, TC_DATA_COL_START,
                         tc.surcharge_rate_1cr_10cr, fmt=FMT_PCT_2)

        # Partnership
        self.write_label(TxL.HDR_ROW + 5, COL_LABEL,
                         "Partnership Firm", bold=True, fill=fill_alt)
        self.write_label(TxL.HDR_ROW + 6, COL_LABEL,
                         "  Tax Rate", fill=fill_white)
        self.write_input(TxL.HDR_ROW + 6, TC_DATA_COL_START,
                         tc.partnership_rate, fmt=FMT_PCT_2)
        self.write_label(TxL.HDR_ROW + 7, COL_LABEL,
                         "  Surcharge (above ₹1Cr)", fill=fill_alt)
        self.write_input(TxL.HDR_ROW + 7, TC_DATA_COL_START,
                         tc.partnership_surcharge_rate, fmt=FMT_PCT_2)

        for r in range(TxL.HDR_ROW, TxL.HDR_ROW + 9):
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        # Store slab references
        self._co_rate_row   = TxL.HDR_ROW + 1
        self._hec_row       = TxL.HDR_ROW + 2
        self._co_sc_row     = TxL.HDR_ROW + 3
        self._part_rate_row = TxL.HDR_ROW + 6
        self._part_sc_row   = TxL.HDR_ROW + 7

    # ── Tax computation ───────────────────────────────────────────────────────

    def _write_tax_computation(self):
        last = TC_DATA_COL_START + self.n_years - 1

        self.write_sub_header(TxL.YEAR_ROW - 2, COL_LABEL, last,
                               "CURRENT TAX CALCULATION",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[TxL.YEAR_ROW - 2].height = 13

        # Year headers
        for yr in range(1, self.n_years + 1):
            col = TC_DATA_COL_START + yr - 1
            self.write_column_header(TxL.YEAR_ROW - 1, col, f"Year {yr}",
                                     fill=fill_solid(NAVY))
        self.ws.row_dimensions[TxL.YEAR_ROW - 1].height = ROW_HEIGHT_SUBHDR

        # Taxable income row (references PL PBT × 100000 to convert Lakhs → INR)
        self.write_label(TxL.TAXABLE_INC_ROW, COL_LABEL,
                         "Taxable Income (₹)", bold=True, fill=fill_alt)

        # PBT row will be set when PL is written; use a placeholder formula
        # that P&L sheet will fill correctly. For now, reference PL!PBT_ROW.
        # We'll use the known PL layout row — import it
        from core.layout_engine import PLLayout
        pbt_row_pl = PLLayout.pbt_row()

        for yr in range(1, self.n_years + 1):
            col = TC_DATA_COL_START + yr - 1
            c   = get_column_letter(col)
            # PL data starts at F (col 6); year 1 = F, year 2 = G ...
            pl_col = get_column_letter(6 + yr - 1)
            formula = f"=PL!{pl_col}{pbt_row_pl}*100000"
            self.write_formula(TxL.TAXABLE_INC_ROW, col, formula,
                               fmt="#,##0", fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[TxL.TAXABLE_INC_ROW].height = ROW_HEIGHT_DATA

        # Company tax computation
        self._write_company_tax()
        # Partnership tax computation
        self._write_partnership_tax()
        # Individual tax computation
        self._write_individual_tax()
        # Active tax row (entity-type selector)
        self._write_active_tax_row()

    def _write_company_tax(self):
        basic_r = TxL.CO_BASIC_ROW
        hec_r   = TxL.CO_HEC_ROW
        total_r = TxL.CO_TOTAL_ROW

        self.write_label(basic_r - 1, COL_LABEL,
                         "Company — Tax Computation", bold=True,
                         fill=fill_solid(BLUE))
        self.ws.cell(basic_r - 1, COL_LABEL).font = Font(
            name="Arial", size=10, bold=True, color=WHITE)

        self.write_label(basic_r, COL_LABEL,
                         "  Basic Tax (30%)", fill=fill_white)
        self.write_label(hec_r,   COL_LABEL,
                         "  + Health & Education Cess (4%)", fill=fill_alt)
        self.write_label(total_r, COL_LABEL,
                         "  TOTAL COMPANY TAX", bold=True,
                         fill=fill_solid("D5F5E3"))

        co_rate_ref = f"${get_column_letter(TC_DATA_COL_START)}${self._co_rate_row}"
        hec_ref     = f"${get_column_letter(TC_DATA_COL_START)}${self._hec_row}"

        for yr in range(1, self.n_years + 1):
            col = TC_DATA_COL_START + yr - 1
            c   = get_column_letter(col)
            inc = f"{c}{TxL.TAXABLE_INC_ROW}"

            basic_formula = f"={inc}*{co_rate_ref}"
            self.write_formula(basic_r, col, basic_formula,
                               fmt=FMT_LAKHS, fill=fill_white)
            hec_formula   = f"={c}{basic_r}*{hec_ref}"
            self.write_formula(hec_r,   col, hec_formula,
                               fmt=FMT_LAKHS, fill=fill_alt)
            total_formula = f"={c}{basic_r}+{c}{hec_r}"
            self.write_total(total_r, col, total_formula, bold=False)

        for r in [basic_r, hec_r, total_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    def _write_partnership_tax(self):
        basic_r = TxL.PART_BASIC_ROW
        hec_r   = TxL.PART_HEC_ROW
        total_r = TxL.PART_TOTAL_ROW

        self.write_label(basic_r - 1, COL_LABEL,
                         "Partnership — Tax Computation", bold=True,
                         fill=fill_solid(BLUE))
        self.ws.cell(basic_r - 1, COL_LABEL).font = Font(
            name="Arial", size=10, bold=True, color=WHITE)

        self.write_label(basic_r, COL_LABEL,
                         "  Basic Tax (30%)", fill=fill_white)
        self.write_label(hec_r,   COL_LABEL,
                         "  + Health & Education Cess (4%)", fill=fill_alt)
        self.write_label(total_r, COL_LABEL,
                         "  TOTAL PARTNERSHIP TAX", bold=True,
                         fill=fill_solid("D5F5E3"))

        pt_rate_ref = f"${get_column_letter(TC_DATA_COL_START)}${self._part_rate_row}"
        hec_ref     = f"${get_column_letter(TC_DATA_COL_START)}${self._hec_row}"

        for yr in range(1, self.n_years + 1):
            col = TC_DATA_COL_START + yr - 1
            c   = get_column_letter(col)
            inc = f"{c}{TxL.TAXABLE_INC_ROW}"

            self.write_formula(basic_r, col, f"={inc}*{pt_rate_ref}",
                               fmt=FMT_LAKHS, fill=fill_white)
            self.write_formula(hec_r,   col, f"={c}{basic_r}*{hec_ref}",
                               fmt=FMT_LAKHS, fill=fill_alt)
            self.write_total(total_r, col, f"={c}{basic_r}+{c}{hec_r}",
                             bold=False)

        for r in [basic_r, hec_r, total_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    def _write_individual_tax(self):
        basic_r = TxL.IND_BASIC_ROW
        hec_r   = TxL.IND_HEC_ROW
        total_r = TxL.IND_TOTAL_ROW

        self.write_label(basic_r - 1, COL_LABEL,
                         "Individual / Proprietor — Tax Computation",
                         bold=True, fill=fill_solid(BLUE))
        self.ws.cell(basic_r - 1, COL_LABEL).font = Font(
            name="Arial", size=10, bold=True, color=WHITE)

        self.write_label(basic_r, COL_LABEL,
                         "  Basic Tax (slab rates)", fill=fill_white)
        self.write_label(hec_r,   COL_LABEL,
                         "  + Health & Education Cess (4%)", fill=fill_alt)
        self.write_label(total_r, COL_LABEL,
                         "  TOTAL INDIVIDUAL TAX", bold=True,
                         fill=fill_solid("D5F5E3"))

        hec_ref = f"${get_column_letter(TC_DATA_COL_START)}${self._hec_row}"
        # Slab-based nested IF (new tax regime slabs for FY25 onwards)
        SLABS = [
            (300000,  0.00),
            (700000,  0.05),
            (1000000, 0.10),
            (1200000, 0.15),
            (1500000, 0.20),
            (float("inf"), 0.30),
        ]

        for yr in range(1, self.n_years + 1):
            col = TC_DATA_COL_START + yr - 1
            c   = get_column_letter(col)
            inc = f"{c}{TxL.TAXABLE_INC_ROW}"

            # Build nested IF for slab calculation
            slabs = [(300000,0),(700000,0.05),(1000000,0.10),
                     (1200000,0.15),(1500000,0.20)]
            formula = f"={inc}*0.30"  # top slab fallback
            prev_limit = 0
            for limit, rate in reversed(slabs):
                tax_below = 0.0
                pl = 0
                for lim2, r2 in slabs:
                    if lim2 <= limit and lim2 > pl:
                        tax_below += (lim2 - pl) * r2
                        pl = lim2
                formula = (f"IF({inc}<={limit},"
                           f"{tax_below}+({inc}-{prev_limit})*{rate},"
                           f"{formula})")
                prev_limit = limit
            formula = "=" + formula

            self.write_formula(basic_r, col, formula,
                               fmt=FMT_LAKHS, fill=fill_white)
            self.write_formula(hec_r,   col, f"={c}{basic_r}*{hec_ref}",
                               fmt=FMT_LAKHS, fill=fill_alt)
            self.write_total(total_r, col, f"={c}{basic_r}+{c}{hec_r}",
                             bold=False)

        for r in [basic_r, hec_r, total_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    def _write_active_tax_row(self):
        """Write the row that P&L uses — selects correct table based on entity type."""
        active_r = TxL.CO_TOTAL_ROW + 3  # a few rows below last table
        last     = TC_DATA_COL_START + self.n_years - 1

        self.write_sub_header(active_r - 1, COL_LABEL, last,
                               "ACTIVE TAX (auto-selected by entity type)",
                               fill=fill_solid(TEAL))
        self.ws.row_dimensions[active_r - 1].height = 13

        self.write_label(active_r, COL_LABEL,
                         "Current Tax Expense (₹ Lakhs)",
                         bold=True, fill=fill_solid("D5F5E3"))

        asmp_entity = f"Assumption!${get_column_letter(TC_DATA_COL_START)}${self._entity_row()}"
        co_total    = TxL.CO_TOTAL_ROW
        part_total  = TxL.PART_TOTAL_ROW
        ind_total   = TxL.IND_TOTAL_ROW

        for yr in range(1, self.n_years + 1):
            col = TC_DATA_COL_START + yr - 1
            c   = get_column_letter(col)
            # Convert from INR to Lakhs (/100000)
            formula = (
                f'=IF({asmp_entity}="Company",{c}{co_total}/100000,'
                f'IF({asmp_entity}="Partnership",{c}{part_total}/100000,'
                f'{c}{ind_total}/100000))'
            )
            self.write_total(active_r, col, formula, bold=True)
        self.ws.row_dimensions[active_r].height = ROW_HEIGHT_DATA + 2

        self.layout._map["Tax"] = self.layout._map.get("Tax", {})
        self.layout._map["Tax"]["active_tax_row"] = (active_r, TC_DATA_COL_START)
        self.layout._map["Tax"]["data_col_start"]  = TC_DATA_COL_START

    def _entity_row(self) -> int:
        """Row in Assumption where entity_type is stored."""
        entry = self.layout._map["Assumption"].get("entity_type")
        if entry:
            return entry[0]
        return 106  # fallback

    # ── Register layout ───────────────────────────────────────────────────────

    def _register_layout(self):
        self.layout._map["Tax"]["taxable_inc_row"] = (
            TxL.TAXABLE_INC_ROW, TC_DATA_COL_START
        )
        self.layout._map["Tax"]["co_total_row"] = (
            TxL.CO_TOTAL_ROW, TC_DATA_COL_START
        )
        self.layout._map["Tax"]["part_total_row"] = (
            TxL.PART_TOTAL_ROW, TC_DATA_COL_START
        )
        self.layout._map["Tax"]["ind_total_row"] = (
            TxL.IND_TOTAL_ROW, TC_DATA_COL_START
        )

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _asmp_abs(self, logical: str) -> str:
        entry = self.layout._map["Assumption"].get(logical)
        if entry:
            r, c = entry
            return f"Assumption!${get_column_letter(c)}${r}"
        return "0"
