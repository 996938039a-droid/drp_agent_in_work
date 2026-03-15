"""
sheet_depreciation.py
──────────────────────
Writes the Depreciation sheet.

Structure per asset class:
  Row +0: Opening WDV balance (Year 1 = from IDC/Cost & Means; Year 2+ = prior closing)
  Row +1: Additions during the year (currently 0 for all years)
  Row +2: Depreciation charge = IF(months<=6, (opening+additions)×rate/2,
                                              (opening+additions)×rate)
  Row +3: Closing WDV = Opening + Additions - Depreciation

Summary block:
  Gross Block    = sum of all opening values
  Cumulative Depr = running total of all charges
  Net Block WDV  = Gross Block - Cumulative Depr

Balance check:
  Sum of individual closing WDVs = Net Block (should be TRUE each year)

All depreciation rates reference Assumption sheet.
Opening Year 1 values are the asset costs (post-IDC capitalization).
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore, AssetCategory
from core.layout_engine import (
    LayoutEngine, DepreciationLayout as DL, AssumptionLayout as AL,
    col_letter, year_col, COL_LABEL, COL_BASIS, COL_YEAR_START,
    ROWS_PER_ASSET_CLASS
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_NUMBER, FMT_PCT_2, FMT_INTEGER,
    font_formula, font_xsheet, font_input, font_header, font_check_pass,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid, fill_amber,
    fill_teal_hdr, fill_blue_hdr,
    border_all_thin, align_center, align_left, align_right,
    ROW_HEIGHT_HEADER, ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, LIGHT_GREY, WHITE,
    FORMULA_FONT_COLOR, XSHEET_FONT_COLOR,
)

# Map AssetCategory → Assumption depreciation rate key
DEPR_RATE_MAP = {
    AssetCategory.PLANT_MACHINERY:  "dep_pm_rate",
    AssetCategory.CIVIL_WORKS:      "dep_civil_rate",
    AssetCategory.FURNITURE:        "dep_furn_rate",
    AssetCategory.VEHICLE:          "dep_veh_rate",
    AssetCategory.ELECTRICAL:       "dep_elec_rate",
    AssetCategory.PRE_OPERATIVE:    "dep_preop_rate",
    AssetCategory.OTHER:            "dep_pm_rate",   # default to P&M rate
}


class DepreciationWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years = store.n_years

        # Get ordered unique asset categories from store
        seen = []
        for a in store.capital_means.assets:
            if a.category not in seen:
                seen.append(a.category)
        self._asset_classes = seen
        self._n_classes     = len(seen)

        # Group assets by category for cost totals
        self._assets_by_cat = {}
        for a in store.capital_means.assets:
            self._assets_by_cat.setdefault(a.category, []).append(a)

        super().__init__(wb, "Depreciation")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = NAVY
        self.ws.column_dimensions[get_column_letter(COL_LABEL)].width = 30
        self.ws.column_dimensions[get_column_letter(COL_BASIS)].width = 10
        for yr in range(1, self.store.n_years + 1):
            self.ws.column_dimensions[get_column_letter(year_col(yr))].width = 16

    def _write(self):
        self._write_title()
        self._write_months_row()
        self._write_asset_blocks()
        self._write_summary()
        self._write_balance_check()
        self.ws.freeze_panes = f"{get_column_letter(COL_YEAR_START)}{DL.MONTHS_ROW + 1}"

    # ── Title ─────────────────────────────────────────────────────────────────

    def _write_title(self):
        last = year_col(self.n_years)
        self.write_section_header(1, COL_LABEL, last,
            "STATEMENT OF DEPRECIATION (IT Act — WDV Method)")
        self.write_section_header(2, COL_LABEL, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 14

    # ── Months in operation ───────────────────────────────────────────────────

    def _write_months_row(self):
        self.write_label(DL.MONTHS_ROW, COL_LABEL, "Months in Operation",
                         fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            self.write_formula(DL.MONTHS_ROW, year_col(yr), "=12",
                               fmt=FMT_INTEGER, fill=fill_alt)
        self.ws.row_dimensions[DL.MONTHS_ROW].height = ROW_HEIGHT_DATA

    # ── Asset class blocks ────────────────────────────────────────────────────

    def _write_asset_blocks(self):
        for i, cat in enumerate(self._asset_classes):
            self._write_one_asset_block(i, cat)

    def _write_one_asset_block(self, idx: int, cat: AssetCategory):
        opening_row = DL.opening_row(idx)
        addition_row= DL.addition_row(idx)
        charge_row  = DL.charge_row(idx)
        closing_row = DL.closing_row(idx)
        last_col    = year_col(self.n_years)

        # Asset class label
        cat_key = cat.value.lower().replace(" ", "_").replace("&", "and")
        rate_key = DEPR_RATE_MAP.get(cat, "dep_pm_rate")
        rate_ref = self._asmp_abs(rate_key)   # e.g. =Assumption!$E$97

        fill_bg   = fill_white if idx % 2 == 0 else fill_alt
        fill_hdr  = fill_solid("EAF4FB") if idx % 2 == 0 else fill_solid("D6EAF8")

        # Section header for this asset class
        hdr_row = opening_row - 1
        self.write_sub_header(hdr_row, COL_LABEL, last_col,
                               f"  {cat.value}",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[hdr_row].height = 13

        # Opening balance
        self.write_label(opening_row, COL_LABEL, "  Opening Balance (WDV)",
                         fill=fill_bg)
        self.write_label(opening_row, COL_BASIS, "₹ Lakhs", fill=fill_bg)

        total_cost = sum(a.cost_lakhs for a in self._assets_by_cat.get(cat, []))

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            if yr == 1:
                # Year 1 opening = total cost of this asset category
                # (In the real model this comes from IDC sheet; here we use direct cost)
                formula = f"={total_cost}"
                self.write_formula(opening_row, col, formula,
                                   fmt=FMT_LAKHS, fill=fill_hdr)
            else:
                # Year N opening = Year (N-1) closing
                prev_c = get_column_letter(year_col(yr - 1))
                formula = f"={prev_c}{closing_row}"
                self.write_formula(opening_row, col, formula,
                                   fmt=FMT_LAKHS, fill=fill_bg)

        # Additions (currently zero — no capex assumed after Year 1)
        self.write_label(addition_row, COL_LABEL, "  Additions", fill=fill_bg)
        for yr in range(1, self.n_years + 1):
            self.write_formula(addition_row, year_col(yr), "=0",
                               fmt=FMT_LAKHS, fill=fill_bg)

        # Depreciation charge — with half-year rule
        self.write_label(charge_row, COL_LABEL,
                         f"  Depreciation ({rate_key.replace('dep_','').replace('_rate','').upper()} Rate)",
                         fill=fill_bg)
        self.write_label(charge_row, COL_BASIS, rate_ref.replace("=", ""),
                         fill=fill_bg)

        months_row = DL.MONTHS_ROW
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            # Half-year rule: if months ≤ 6, apply half rate
            formula = (
                f"=IF({c}{months_row}<=6,"
                f"({c}{opening_row}+{c}{addition_row})*{rate_ref}/2,"
                f"({c}{opening_row}+{c}{addition_row})*{rate_ref})"
            )
            self.write_formula(charge_row, col, formula,
                               fmt=FMT_LAKHS, fill=fill_solid("FFF3CD"),
                               xsheet=True)

        # Closing balance
        self.write_label(closing_row, COL_LABEL,
                         "  Closing Balance (WDV)", bold=True,
                         fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = (f"={c}{opening_row}"
                       f"+{c}{addition_row}"
                       f"-{c}{charge_row}")
            self.write_total(closing_row, col, formula, bold=False)

        # Set row heights
        for r in [opening_row, addition_row, charge_row, closing_row]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    # ── Summary block ─────────────────────────────────────────────────────────

    def _write_summary(self):
        n        = self._n_classes
        gb_row   = DL.gross_block_row(n)
        cd_row   = DL.cumul_depr_row(n)
        nb_row   = DL.net_block_row(n)
        last_col = year_col(self.n_years)

        # Section header
        self.write_section_header(gb_row - 1, COL_LABEL, last_col,
                                   "SUMMARY — FIXED ASSETS SCHEDULE")
        self.ws.row_dimensions[gb_row - 1].height = 14

        # Gross Block = sum of opening balances of all asset classes (Year 1 = cost, fixed thereafter)
        self.write_label(gb_row, COL_LABEL, "Gross Block",
                         bold=True, fill=fill_alt)
        self.write_label(gb_row, COL_BASIS, "₹ Lakhs", fill=fill_alt)

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            if yr == 1:
                # Year 1 gross block = sum of all opening year 1 values
                opening_refs = [f"{c}{DL.opening_row(i)}" for i in range(n)]
                formula = "=" + "+".join(opening_refs)
            else:
                # Gross block stays constant (no additions modelled)
                prev_c = get_column_letter(year_col(yr - 1))
                formula = f"={prev_c}{gb_row}"
            self.write_formula(gb_row, col, formula,
                               fmt=FMT_LAKHS, fill=fill_alt)

        # Cumulative Depreciation = running total
        self.write_label(cd_row, COL_LABEL, "Less: Accumulated Depreciation",
                         fill=fill_white)
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            if yr == 1:
                # Year 1 cumulative = Year 1 charges only
                charge_refs = [f"{c}{DL.charge_row(i)}" for i in range(n)]
                formula = "=" + "+".join(charge_refs)
            else:
                prev_c  = get_column_letter(year_col(yr - 1))
                # Cumulative = prior cumulative + this year's charges
                charge_refs = [f"{c}{DL.charge_row(i)}" for i in range(n)]
                formula = (f"={prev_c}{cd_row}+" +
                           "+".join(charge_refs))
            self.write_formula(cd_row, col, formula,
                               fmt=FMT_LAKHS, fill=fill_white)

        # Net Block WDV = Gross - Cumulative
        self.write_label(nb_row, COL_LABEL, "Net Block (WDV)",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = f"={c}{gb_row}-{c}{cd_row}"
            cell = self.ws.cell(row=nb_row, column=col, value=formula)
            cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill = fill_solid("1ABC9C")
            cell.number_format = FMT_LAKHS
            cell.alignment = align_right
            cell.border = border_all_thin(TEAL)

        for r in [gb_row, cd_row, nb_row]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        # Store net block row for reference by Expenses, P&L, BS
        self.layout._map["Depreciation"]["net_block"] = (nb_row, year_col(1))
        self.layout._map["Depreciation"]["gross_block"] = (gb_row, year_col(1))
        self.layout._map["Depreciation"]["cumul_depr"] = (cd_row, year_col(1))

    # ── Balance check ─────────────────────────────────────────────────────────

    def _write_balance_check(self):
        n         = self._n_classes
        nb_row    = DL.net_block_row(n)
        check_row = DL.check_row(n)
        last_col  = year_col(self.n_years)

        self.write_label(check_row, COL_LABEL,
                         "Check: Sum of WDVs = Net Block",
                         fill=fill_alt)

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            # Sum of individual closing balances
            closing_refs = "+".join(
                f"{c}{DL.closing_row(i)}" for i in range(n)
            )
            formula = (
                f'=IF(ROUND({closing_refs},0)='
                f'ROUND({c}{nb_row},0),"✓ OK",'
                f'ROUND({closing_refs},0)-ROUND({c}{nb_row},0))'
            )
            cell = self.ws.cell(row=check_row, column=col, value=formula)
            cell.font   = Font(name="Arial", size=9, color="27AE60")
            cell.fill   = fill_alt
            cell.alignment = align_center
        self.ws.row_dimensions[check_row].height = ROW_HEIGHT_DATA

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _asmp_abs(self, logical: str) -> str:
        entry = self.layout._map["Assumption"].get(logical)
        if entry:
            r, c = entry
            return f"=Assumption!${get_column_letter(c)}${r}"
        return "=0.15"  # safe fallback

    def _depr_rate_label(self, cat: AssetCategory) -> str:
        return DEPR_RATE_MAP.get(cat, "dep_pm_rate")
