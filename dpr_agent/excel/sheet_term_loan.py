"""
sheet_term_loan.py
───────────────────
Writes the Term Loan sheet.

Structure per loan:
  Monthly schedule (rows 6–89 for 84 months):
    Col B: Year label (1-8)
    Col C: Opening balance = prior row closing
    Col D: Principal repayment = $I$16 (constant EMI after moratorium)
    Col E: Interest = C × rate × (1/12)
    Col F: Total instalment = D + E
    Col G: Closing = C - D

  Annual summary (right side, cols J onwards):
    Row 8:  Repayment months per year
    Row 9:  Annual principal repaid (SUMIF on year label)
    Row 10: Year-end outstanding balance
    Row 11: Annual interest paid (SUMIF)

  OD interest table below monthly schedule.

Key fix vs original sample:
  Repayment months = Tenor - Moratorium (not (12×7)-6)
  EMI = Loan_Amount / Repayment_months
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore, FinanceSourceType
from core.layout_engine import (
    LayoutEngine, TermLoanLayout as TL, AssumptionLayout as AL,
    col_letter, year_col, COL_LABEL
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_NUMBER, FMT_PCT_2, FMT_INTEGER,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid, fill_amber,
    border_all_thin, align_center, align_left, align_right,
    ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, WHITE, LIGHT_GREY,
)

# Monthly schedule column layout
MC_YEAR    = 2   # B
MC_OPENING = 3   # C
MC_PRINC   = 4   # D
MC_INT     = 5   # E
MC_TOTAL   = 6   # F
MC_CLOSING = 7   # G

# Annual summary column layout
AC_START   = 9   # I — year labels start here; data in J onwards
AC_YEAR_LABEL_ROW_OFFSET = 0
AC_MONTHS_ROW_OFFSET     = 1
AC_REPAY_ROW_OFFSET      = 2
AC_PRINCIPAL_ROW_OFFSET  = 3
AC_OUTSTANDING_ROW_OFFSET= 4
AC_INTEREST_ROW_OFFSET   = 5


class TermLoanWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years = store.n_years
        self._loans  = store.capital_means.term_loans
        self._od     = store.capital_means.od_sources
        super().__init__(wb, "Term Loan")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = NAVY
        # Monthly schedule columns
        widths = {2: 8, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14,
                  8: 4,  9: 18}
        for col, w in widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = w
        for yr in range(1, self.n_years + 2):
            self.ws.column_dimensions[get_column_letter(AC_START + yr)].width = 15

    def _write(self):
        self._write_title()
        if self._loans:
            self._write_monthly_schedule(self._loans[0])
            self._write_annual_summary(self._loans[0])
        self._write_od_interest()
        self.ws.freeze_panes = f"{get_column_letter(MC_OPENING)}6"

    # ── Title ─────────────────────────────────────────────────────────────────

    def _write_title(self):
        last_col = AC_START + self.n_years + 1
        self.write_section_header(1, COL_LABEL, last_col,
            "STATEMENT OF REPAYMENT OF TERM LOAN & INTEREST CALCULATION")
        self.write_section_header(2, COL_LABEL, last_col,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 14

    # ── Monthly schedule ──────────────────────────────────────────────────────

    def _write_monthly_schedule(self, loan):
        # Rate row
        r_row = TL.RATE_ROW
        self.write_label(r_row, MC_YEAR, "Rate", fill=fill_alt)
        asmp_rate = self._asmp_loan_abs("fin_rate_l0")
        self.write_formula(r_row, MC_OPENING, f"={asmp_rate}",
                           fmt=FMT_PCT_2, fill=fill_amber, xsheet=True)
        self.ws.row_dimensions[r_row].height = ROW_HEIGHT_DATA

        # Column headers
        hdr = TL.SUBHDR_ROW
        for col, lbl in [(MC_YEAR,"Yr"),(MC_OPENING,"Opening"),
                         (MC_PRINC,"Principal"),(MC_INT,"Interest"),
                         (MC_TOTAL,"Instalment"),(MC_CLOSING,"Closing")]:
            self.write_column_header(hdr, col, lbl, fill=fill_solid(NAVY))
        self.ws.row_dimensions[hdr].height = ROW_HEIGHT_SUBHDR

        # EMI calculation cells (hidden helper area in col I)
        emi_loan_row = hdr + 1
        emi_rep_row  = hdr + 2
        emi_val_row  = hdr + 3

        asmp_amount   = self._asmp_loan_abs("fin_amount_l0")
        asmp_tenor    = self._asmp_loan_abs("fin_tenor_l0")
        asmp_mora     = self._asmp_loan_abs("fin_moratorium_l0")

        self.ws.cell(emi_loan_row, AC_START).value = f"={asmp_amount}"
        self.ws.cell(emi_rep_row,  AC_START).value = (
            f"={asmp_tenor}-{asmp_mora}")
        self.ws.cell(emi_val_row,  AC_START).value = (
            f"={get_column_letter(AC_START)}{emi_loan_row}"
            f"/{get_column_letter(AC_START)}{emi_rep_row}")
        for r in [emi_loan_row, emi_rep_row, emi_val_row]:
            self.ws.cell(r, AC_START).number_format = FMT_LAKHS

        emi_ref = (f"${get_column_letter(AC_START)}${emi_val_row}")

        # Monthly rows
        base         = TL.MONTHLY_BASE
        tenor_months = loan.tenor_months
        mora_months  = loan.moratorium_months
        rate_ref     = f"${get_column_letter(MC_OPENING)}${r_row}"

        for m in range(tenor_months):
            r = base + m
            yr_label = (m // 12) + 1

            # Year label
            self.ws.cell(r, MC_YEAR).value = yr_label
            self.ws.cell(r, MC_YEAR).font  = Font(name="Arial", size=9,
                                                    color="888888")
            self.ws.row_dimensions[r].height = 13

            # Opening
            if m == 0:
                self.ws.cell(r, MC_OPENING).value = f"={asmp_amount}"
            else:
                prev_g = f"{get_column_letter(MC_CLOSING)}{r-1}"
                self.ws.cell(r, MC_OPENING).value = f"={prev_g}"
            self.ws.cell(r, MC_OPENING).number_format = FMT_LAKHS

            # Principal (0 during moratorium, EMI thereafter)
            if m < mora_months:
                self.ws.cell(r, MC_PRINC).value = "=0"
            else:
                self.ws.cell(r, MC_PRINC).value = f"={emi_ref}"
            self.ws.cell(r, MC_PRINC).number_format = FMT_LAKHS

            # Interest = opening × rate / 12
            op_ref = f"{get_column_letter(MC_OPENING)}{r}"
            self.ws.cell(r, MC_INT).value = (
                f"={op_ref}*{rate_ref}/12")
            self.ws.cell(r, MC_INT).number_format = FMT_LAKHS

            # Total instalment
            p_ref = get_column_letter(MC_PRINC)
            i_ref = get_column_letter(MC_INT)
            self.ws.cell(r, MC_TOTAL).value = (
                f"={p_ref}{r}+{i_ref}{r}")
            self.ws.cell(r, MC_TOTAL).number_format = FMT_LAKHS

            # Closing
            self.ws.cell(r, MC_CLOSING).value = (
                f"={op_ref}-{p_ref}{r}")
            self.ws.cell(r, MC_CLOSING).number_format = FMT_LAKHS

        # Store EMI ref for annual summary
        self._emi_val_row = emi_val_row
        self._monthly_base = base
        self._tenor_months = tenor_months

    # ── Annual summary ────────────────────────────────────────────────────────

    def _write_annual_summary(self, loan):
        base     = TL.SUMMARY_YEAR_ROW
        yr_lbl   = base
        months_r = base + 1
        repay_r  = base + 2
        princ_r  = base + 3
        outst_r  = base + 4
        int_r    = base + 5

        # Labels (in col I)
        labels = {
            yr_lbl:  "Year →",
            months_r:"Months",
            repay_r: "Repayment months",
            princ_r: "Principal repaid",
            outst_r: "Outstanding balance",
            int_r:   "Interest paid",
        }
        for row, lbl in labels.items():
            cell = self.ws.cell(row=row, column=AC_START)
            cell.value = lbl
            cell.font  = Font(name="Arial", size=9, bold=True, color="595959")
            self.ws.row_dimensions[row].height = 14

        # Year columns
        monthly_start = TL.MONTHLY_BASE
        month_col_b   = get_column_letter(MC_YEAR)

        for yr in range(1, self.n_years + 1):
            col = AC_START + yr

            # Year label
            self.ws.cell(yr_lbl,  col).value = yr
            self.ws.cell(yr_lbl,  col).font  = Font(name="Arial", size=9, bold=True)

            # Months in year
            self.ws.cell(months_r, col).value = "=12"
            self.ws.cell(months_r, col).number_format = FMT_INTEGER

            # Repayment months (COUNTIF on year label column)
            b_range = (f"{month_col_b}{monthly_start}:"
                       f"{month_col_b}{monthly_start + self._tenor_months - 1}")
            self.ws.cell(repay_r, col).value = (
                f"=COUNTIF({b_range},{yr})")
            self.ws.cell(repay_r, col).number_format = FMT_INTEGER

            # Annual principal (SUMIF)
            d_range = (f"{get_column_letter(MC_PRINC)}{monthly_start}:"
                       f"{get_column_letter(MC_PRINC)}{monthly_start + self._tenor_months - 1}")
            self.ws.cell(princ_r, col).value = (
                f"=SUMIF({b_range},{yr},{d_range})")
            self.ws.cell(princ_r, col).number_format = FMT_LAKHS

            # Outstanding = prior outstanding - this year principal
            asmp_loan = self._asmp_loan_abs("fin_amount_l0")
            if yr == 1:
                self.ws.cell(outst_r, col).value = (
                    f"={asmp_loan}"
                    f"-{get_column_letter(col)}{princ_r}")
            else:
                prev_c = get_column_letter(col - 1)
                self.ws.cell(outst_r, col).value = (
                    f"={prev_c}{outst_r}"
                    f"-{get_column_letter(col)}{princ_r}")
            self.ws.cell(outst_r, col).number_format = FMT_LAKHS

            # Annual interest (SUMIF)
            e_range = (f"{get_column_letter(MC_INT)}{monthly_start}:"
                       f"{get_column_letter(MC_INT)}{monthly_start + self._tenor_months - 1}")
            self.ws.cell(int_r, col).value = (
                f"=SUMIF({b_range},{yr},{e_range})")
            self.ws.cell(int_r, col).number_format = FMT_LAKHS

            for r in [yr_lbl, months_r, repay_r, princ_r, outst_r, int_r]:
                self.ws.row_dimensions[r].height = 14

        # Store annual summary rows in layout for P&L, BS, Ratio
        self.layout._map["Term Loan"] = self.layout._map.get("Term Loan", {})
        self.layout._map["Term Loan"]["annual_principal_row"] = (princ_r, AC_START + 1)
        self.layout._map["Term Loan"]["annual_outstanding_row"] = (outst_r, AC_START + 1)
        self.layout._map["Term Loan"]["annual_interest_row"] = (int_r, AC_START + 1)
        self.layout._map["Term Loan"]["summary_col_start"] = AC_START + 1

    # ── OD Interest table ─────────────────────────────────────────────────────

    def _write_od_interest(self):
        od_list = self._od
        if not od_list:
            return

        od = od_list[0]
        base = TL.od_base_row()

        last_col = AC_START + self.n_years + 1
        self.write_sub_header(base - 1, COL_LABEL, last_col,
                               "OD / Working Capital Interest",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[base - 1].height = 13

        # Headers
        self.write_column_header(base, MC_YEAR,    "Year",    fill=fill_solid(NAVY))
        self.write_column_header(base, MC_OPENING, "OD Limit",fill=fill_solid(NAVY))
        self.write_column_header(base, MC_INT,     "Interest", fill=fill_solid(NAVY))
        self.ws.row_dimensions[base].height = ROW_HEIGHT_SUBHDR

        asmp_od_limit = self._asmp_loan_abs("fin_od_limit")
        asmp_od_rate  = self._asmp_loan_abs("fin_od_rate")

        for yr in range(1, self.n_years + 1):
            r = base + yr
            self.ws.cell(r, MC_YEAR).value    = yr
            self.ws.cell(r, MC_OPENING).value = f"={asmp_od_limit}"
            self.ws.cell(r, MC_INT).value     = (
                f"={get_column_letter(MC_OPENING)}{r}*{asmp_od_rate}")
            self.ws.cell(r, MC_OPENING).number_format = FMT_LAKHS
            self.ws.cell(r, MC_INT).number_format     = FMT_LAKHS
            self.ws.row_dimensions[r].height = 14

        # Store OD interest row range for P&L
        self.layout._map["Term Loan"]["od_interest_base_row"] = (
            base + 1, MC_INT
        )

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _asmp_loan_abs(self, logical: str) -> str:
        entry = self.layout._map["Assumption"].get(logical)
        if entry:
            r, c = entry
            return f"Assumption!${get_column_letter(c)}${r}"
        # Fallback to Finance section OD params
        entry2 = {
            "fin_od_limit": self.layout._map["Assumption"].get("fin_od_limit"),
            "fin_od_rate":  self.layout._map["Assumption"].get("fin_od_rate"),
        }.get(logical)
        if entry2:
            r, c = entry2
            return f"Assumption!${get_column_letter(c)}${r}"
        return "0"
