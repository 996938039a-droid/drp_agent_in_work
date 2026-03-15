"""
sheet_pl.py
────────────
Writes the Profit & Loss Account sheet.

Structure:
  Revenue:
    Sales (per product)
    Total Revenue from Operations

  Cost of Sales:
    Raw Material Cost (from Expenses)
    Total COGS
    Gross Profit

  Operating Expenses:
    R&M, Insurance, Marketing, Power, Manpower, SGA, Transport, Misc
    Total Operating Expenses

  EBIT (Gross Profit - Opex)
  Finance Costs:
    Term Loan Interest
    OD Interest
    WC Interest
    Total Finance Costs

  PBT (EBIT - Finance Costs)
  Less: Current Tax (from Tax sheet)
  PAT

  Retained Profit = PAT (no dividend modelled for MSME DPR)
  EBITDA = EBIT + Depreciation

All rows cross-reference their source sheets. Zero hardcoded values.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore
from core.layout_engine import (
    LayoutEngine, PLLayout as PL,
    RevenueLayout as RL, ExpensesLayout as EL,
    DepreciationLayout as DL, TermLoanLayout as TL,
    WCapLayout as WL, ManPowerLayout as ML, TaxLayout as TxL,
    col_letter, year_col, COL_LABEL, COL_BASIS, COL_YEAR_START
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_PCT_2,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid,
    border_all_thin, align_right,
    ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA,
    NAVY, BLUE, TEAL, WHITE,
)

PL_DATA_COL = 6   # F = Year 1 data column


class PLWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store    = store
        self.layout   = layout
        self.n_years  = store.n_years
        self.n_prods  = store.n_products
        self.n_mats   = store.n_materials
        self.n_emp    = store.n_employee_categories
        self._n_ac    = len(set(a.category for a in store.capital_means.assets))
        super().__init__(wb, "PL")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = TEAL
        self.ws.column_dimensions[get_column_letter(COL_LABEL)].width = 38
        self.ws.column_dimensions[get_column_letter(5)].width = 10
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[
                get_column_letter(PL_DATA_COL + yr - 1)].width = 15

    def _write(self):
        self._write_title()
        self._write_col_headers()
        self._write_revenue_section()
        self._write_cogs_section()
        self._write_gross_profit()
        self._write_opex_section()
        self._write_ebit()
        self._write_finance_costs()
        self._write_pbt()
        self._write_tax_and_pat()
        self._write_appropriations()
        self._write_ebitda()
        self._register_layout()
        self.ws.freeze_panes = f"{get_column_letter(PL_DATA_COL)}8"

    # ── Title ─────────────────────────────────────────────────────────────────

    def _write_title(self):
        last = PL_DATA_COL + self.n_years - 1
        self.write_section_header(1, COL_LABEL, last,
            "PROJECTED PROFIT & LOSS ACCOUNT")
        self.write_section_header(2, COL_LABEL, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        for r in [1, 2]:
            self.ws.row_dimensions[r].height = 18 if r == 1 else 14
        # Blank + note rows
        self.write_label(3, COL_LABEL, "", fill=fill_white)
        self.ws.row_dimensions[3].height = 5

    def _write_col_headers(self):
        self.write_column_header(4, COL_LABEL, "Particulars",
                                  fill=fill_solid(NAVY))
        self.write_column_header(4, 5, "Sch/Ref",
                                  fill=fill_solid(NAVY))
        for yr in range(1, self.n_years + 1):
            self.write_column_header(4, PL_DATA_COL + yr - 1, f"Year {yr}",
                                      fill=fill_solid(NAVY))
        self.ws.row_dimensions[4].height = ROW_HEIGHT_SUBHDR

        # Blank spacer
        self.ws.row_dimensions[5].height = 5
        self.write_label(5, COL_LABEL, "", fill=fill_white)

    # ── Revenue ───────────────────────────────────────────────────────────────

    def _write_revenue_section(self):
        last = PL_DATA_COL + self.n_years - 1
        self.write_sub_header(6, COL_LABEL, last,
                               "I. REVENUE FROM OPERATIONS",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[6].height = 13

        # One row per product
        for i, prod in enumerate(self.store.revenue_model.products):
            r    = PL.SALES_ROW + i
            rev_r= RL.prod_revenue_row(i)
            fill = fill_white if i % 2 == 0 else fill_alt
            self.write_label(r, COL_LABEL,
                             f"  Sales — {prod.name}", fill=fill)
            self.write_label(r, 5, "Revenue", fill=fill)
            for yr in range(1, self.n_years + 1):
                c = get_column_letter(PL_DATA_COL + yr - 1)
                self.write_formula(r, PL_DATA_COL + yr - 1,
                    f"=Revenue!{c}{rev_r}",
                    fmt=FMT_LAKHS, fill=fill, xsheet=True)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        # Total Revenue
        r_total = PL.TOTAL_REV_ROW
        tr_src  = RL.total_revenue_row(self.n_prods)
        self.write_label(r_total, COL_LABEL, "  Total Revenue from Operations",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            c = get_column_letter(PL_DATA_COL + yr - 1)
            self._bold_teal(r_total, PL_DATA_COL + yr - 1,
                            f"=Revenue!{c}{tr_src}")
        self.ws.row_dimensions[r_total].height = ROW_HEIGHT_DATA

    # ── COGS ──────────────────────────────────────────────────────────────────

    def _write_cogs_section(self):
        last = PL_DATA_COL + self.n_years - 1
        self.write_sub_header(PL.COGS_ROW - 1, COL_LABEL, last,
                               "II. COST OF SALES",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[PL.COGS_ROW - 1].height = 13

        cogs_src = EL.total_cogs_row(self.n_mats)
        self.write_label(PL.COGS_ROW, COL_LABEL,
                         "  Raw Material Cost of Sales",
                         fill=fill_alt)
        self.write_label(PL.COGS_ROW, 5, "Expenses", fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            c = get_column_letter(PL_DATA_COL + yr - 1)
            self.write_formula(PL.COGS_ROW, PL_DATA_COL + yr - 1,
                f"=Expenses!{c}{cogs_src}",
                fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[PL.COGS_ROW].height = ROW_HEIGHT_DATA

        # Total COGS = same single row for now
        self.write_label(PL.TOTAL_COGS_ROW, COL_LABEL,
                         "  Total Cost of Sales",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = PL_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(PL.TOTAL_COGS_ROW, col,
                            f"={c}{PL.COGS_ROW}")
        self.ws.row_dimensions[PL.TOTAL_COGS_ROW].height = ROW_HEIGHT_DATA

    # ── Gross Profit ──────────────────────────────────────────────────────────

    def _write_gross_profit(self):
        r = PL.GROSS_PROFIT_ROW
        self.write_label(r, COL_LABEL, "GROSS PROFIT  (Revenue – COGS)",
                         bold=True, fill=fill_solid(TEAL))
        for yr in range(1, self.n_years + 1):
            col = PL_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(r, col,
                f"={c}{PL.TOTAL_REV_ROW}-{c}{PL.TOTAL_COGS_ROW}")
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA + 1

    # ── Operating Expenses ────────────────────────────────────────────────────

    def _write_opex_section(self):
        nm   = self.n_mats
        last = PL_DATA_COL + self.n_years - 1
        self.write_sub_header(PL.OE_BASE_ROW - 1, COL_LABEL, last,
                               "III. OPERATING EXPENSES",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[PL.OE_BASE_ROW - 1].height = 13

        # Manpower salary projection row
        sal_proj_row = ML.BASE_ROW + self.n_emp + 4

        lines = [
            ("R&M Expenses",           EL.rm_amount_row(nm),    "Expenses"),
            ("Insurance",              EL.ins_amount_row(nm),   "Expenses"),
            ("Marketing Expenses",     EL.mkt_amount_row(nm),   "Expenses"),
            ("Power & Fuel",           EL.power_amount_row(nm), "Expenses"),
            ("Manpower / Salaries",    sal_proj_row,             "ManPower"),
            ("Depreciation",           DL.charge_row(0),         "Dep_Sum"),
            ("SG&A Expenses",          EL.sga_amount_row(nm),   "Expenses"),
            ("Transportation",         EL.transport_row(nm),    "Expenses"),
            ("Miscellaneous",          EL.misc_row(nm),         "Expenses"),
        ]

        for i, (label, src_row, src_sheet) in enumerate(lines):
            r = PL.OE_BASE_ROW + i
            fill = fill_white if i % 2 == 0 else fill_alt
            self.write_label(r, COL_LABEL, f"  {label}", fill=fill)
            self.write_label(r, 5, src_sheet, fill=fill)

            for yr in range(1, self.n_years + 1):
                col = PL_DATA_COL + yr - 1
                c   = get_column_letter(col)

                if src_sheet == "ManPower":
                    # Salary proj row is n_emp rows below BASE_ROW + buffer
                    formula = f"=ManPower!{c}{src_row}"
                elif src_sheet == "Dep_Sum":
                    # Sum all depreciation charge rows
                    charge_refs = "+".join(
                        f"Depreciation!{c}{DL.charge_row(j)}"
                        for j in range(self._n_ac)
                    )
                    formula = f"={charge_refs}"
                else:
                    formula = f"={src_sheet}!{c}{src_row}"

                self.write_formula(r, col, formula,
                                   fmt=FMT_LAKHS, fill=fill, xsheet=True)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        # Store depreciation row index for EBITDA
        self._depr_opex_row = PL.OE_BASE_ROW + 5  # index 5 = Depreciation

        # Total OPEX
        opex_r = PL.total_opex_row()
        self.write_label(opex_r, COL_LABEL, "  Total Operating Expenses",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = PL_DATA_COL + yr - 1
            c   = get_column_letter(col)
            first = PL.OE_BASE_ROW; last_r = opex_r - 1
            self._bold_teal(opex_r, col,
                            f"=SUM({c}{first}:{c}{last_r})")
        self.ws.row_dimensions[opex_r].height = ROW_HEIGHT_DATA

    # ── EBIT ──────────────────────────────────────────────────────────────────

    def _write_ebit(self):
        r = PL.ebit_row()
        self.write_label(r, COL_LABEL,
                         "EARNINGS BEFORE INTEREST & TAX  (EBIT)",
                         bold=True, fill=fill_solid(TEAL))
        for yr in range(1, self.n_years + 1):
            col = PL_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(r, col,
                f"={c}{PL.GROSS_PROFIT_ROW}-{c}{PL.total_opex_row()}")
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA + 1

    # ── Finance Costs ─────────────────────────────────────────────────────────

    def _write_finance_costs(self):
        last = PL_DATA_COL + self.n_years - 1
        self.write_sub_header(PL.interest_tl_row() - 1, COL_LABEL, last,
                               "IV. FINANCE COSTS",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[PL.interest_tl_row() - 1].height = 13

        tl_int_r  = PL.interest_tl_row()
        od_int_r  = PL.interest_tl_row() + 1   # one row below TL interest
        wc_int_r  = PL.interest_wc_row()
        tot_int_r = PL.total_interest_row()

        # TL annual interest — from Term Loan annual summary
        tl_ann_int_row = TL.SUMMARY_YEAR_ROW + 5
        od_int_base    = TL.od_base_row()

        self.write_label(tl_int_r, COL_LABEL,
                         "  Term Loan Interest", fill=fill_alt)
        self.write_label(od_int_r, COL_LABEL,
                         "  OD / CC Interest", fill=fill_white)
        self.write_label(wc_int_r, COL_LABEL,
                         "  Working Capital Interest", fill=fill_alt)
        self.write_label(tot_int_r, COL_LABEL,
                         "  Total Finance Costs",
                         bold=True, fill=fill_solid("D5F5E3"))

        for yr in range(1, self.n_years + 1):
            col   = PL_DATA_COL + yr - 1
            c     = get_column_letter(col)
            tl_c  = get_column_letter(10 + yr - 1)   # AC_START+yr = col J onward

            # TL interest from Term Loan annual summary
            self.write_formula(tl_int_r, col,
                f"='Term Loan'!{tl_c}{tl_ann_int_row}",
                fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            # OD interest
            od_r = od_int_base + yr
            self.write_formula(od_int_r, col,
                f"='Term Loan'!{get_column_letter(5)}{od_r}",
                fmt=FMT_LAKHS, fill=fill_white, xsheet=True)

            # WC interest
            self.write_formula(wc_int_r, col,
                f"='W Cap'!{c}{WL.WC_INTEREST_ROW}",
                fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            # Total interest
            self._bold_teal(tot_int_r, col,
                f"={c}{tl_int_r}+{c}{od_int_r}+{c}{wc_int_r}")

        for r in [tl_int_r, od_int_r, wc_int_r, tot_int_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    # ── PBT ───────────────────────────────────────────────────────────────────

    def _write_pbt(self):
        r = PL.pbt_row()
        self.write_label(r, COL_LABEL,
                         "PROFIT BEFORE TAX  (PBT)",
                         bold=True, fill=fill_solid(TEAL))
        for yr in range(1, self.n_years + 1):
            col = PL_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(r, col,
                f"={c}{PL.ebit_row()}-{c}{PL.total_interest_row()}")
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA + 1

    # ── Tax & PAT ─────────────────────────────────────────────────────────────

    def _write_tax_and_pat(self):
        last = PL_DATA_COL + self.n_years - 1
        self.write_sub_header(PL.current_tax_row() - 1, COL_LABEL, last,
                               "V. TAX",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[PL.current_tax_row() - 1].height = 13

        tax_r = PL.current_tax_row()
        pat_r = PL.pat_row()

        active_tax_r = self.layout._map.get("Tax", {}).get(
            "active_tax_row", (TxL.CO_TOTAL_ROW + 3, 3))[0]

        self.write_label(tax_r, COL_LABEL, "  Less: Current Tax",
                         fill=fill_alt)
        self.write_label(tax_r, 5, "Tax", fill=fill_alt)

        self.write_label(pat_r, COL_LABEL,
                         "PROFIT AFTER TAX  (PAT)",
                         bold=True, fill=fill_solid(TEAL))

        for yr in range(1, self.n_years + 1):
            col   = PL_DATA_COL + yr - 1
            c     = get_column_letter(col)
            tax_c = get_column_letter(3 + yr - 1)   # TC_DATA_COL_START + yr - 1

            # Tax from Tax sheet active row
            self.write_formula(tax_r, col,
                f"=Tax!{tax_c}{active_tax_r}",
                fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            # PAT = PBT - Tax
            self._bold_teal(pat_r, col,
                f"={c}{PL.pbt_row()}-{c}{tax_r}")

        for r in [tax_r, pat_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    # ── Appropriations & EBITDA ───────────────────────────────────────────────

    def _write_appropriations(self):
        r = PL.retained_profit_row()
        self.write_label(r, COL_LABEL,
                         "  Retained Profit (carried forward)",
                         fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            col = PL_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self.write_formula(r, col,
                f"={c}{PL.pat_row()}",
                fmt=FMT_LAKHS, fill=fill_alt)
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    def _write_ebitda(self):
        last = PL_DATA_COL + self.n_years - 1
        self.write_sub_header(PL.ebitda_row() - 1, COL_LABEL, last,
                               "EBITDA (for Ratio Analysis)",
                               fill=fill_solid(NAVY))
        self.ws.row_dimensions[PL.ebitda_row() - 1].height = 13

        r     = PL.ebitda_row()
        d_row = self._depr_opex_row
        self.write_label(r, COL_LABEL,
                         "EBITDA  (EBIT + Depreciation)",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = PL_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(r, col,
                f"={c}{PL.ebit_row()}+{c}{d_row}")
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    # ── Register ──────────────────────────────────────────────────────────────

    def _register_layout(self):
        self.layout._map["PL"] = {
            "total_rev_row":    (PL.TOTAL_REV_ROW,    PL_DATA_COL),
            "total_cogs_row":   (PL.TOTAL_COGS_ROW,   PL_DATA_COL),
            "gross_profit_row": (PL.GROSS_PROFIT_ROW, PL_DATA_COL),
            "total_opex_row":   (PL.total_opex_row(),  PL_DATA_COL),
            "ebit_row":         (PL.ebit_row(),         PL_DATA_COL),
            "pbt_row":          (PL.pbt_row(),           PL_DATA_COL),
            "tax_row":          (PL.current_tax_row(), PL_DATA_COL),
            "pat_row":          (PL.pat_row(),           PL_DATA_COL),
            "ebitda_row":       (PL.ebitda_row(),        PL_DATA_COL),
            "retained_row":     (PL.retained_profit_row(), PL_DATA_COL),
            "total_interest_row": (PL.total_interest_row(), PL_DATA_COL),
            "data_col_start":   PL_DATA_COL,
            "depr_opex_row":    self._depr_opex_row,
        }

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _bold_teal(self, row: int, col: int, formula: str):
        cell = self.ws.cell(row=row, column=col, value=formula)
        cell.font          = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill          = fill_solid("1ABC9C")
        cell.number_format = FMT_LAKHS
        cell.alignment     = align_right
        cell.border        = border_all_thin(TEAL)
