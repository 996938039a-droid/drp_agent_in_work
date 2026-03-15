"""
sheet_wcap.py
──────────────
Writes the Working Capital sheet.

Structure:
  Current Liabilities:
    Creditors (RM) = Cost of Sales / 365 × creditor_days_rm
    Creditors (Admin) = Admin expenses / 365 × creditor_days_admin
    Total CL

  Current Assets:
    Consumables stock = Cost of Sales / 365 × stock_days_rm
    Debtors = Revenue / 365 × debtor_days
    Cash = from Balance Sheet (circular — resolved by iterative calc)
    Total CA

  Working Capital Requirement = Total CA - Total CL
  WC Bank Finance (if any)
  WC Interest

All days reference Assumption sheet.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore
from core.layout_engine import (
    LayoutEngine, WCapLayout as WL, ExpensesLayout as EL,
    RevenueLayout as RL,
    col_letter, year_col, COL_LABEL, COL_BASIS
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_NUMBER, FMT_PCT_2, FMT_INTEGER,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid, fill_amber,
    border_all_thin, align_center, align_left, align_right,
    ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, WHITE, LIGHT_GREY,
)


class WCapWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years = store.n_years
        self._n_asset_classes = len(set(
            a.category for a in store.capital_means.assets
        ))
        super().__init__(wb, "W Cap")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = TEAL
        self.ws.column_dimensions[get_column_letter(COL_LABEL)].width = 34
        self.ws.column_dimensions[get_column_letter(COL_BASIS)].width = 12
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[get_column_letter(year_col(yr))].width = 15

    def _write(self):
        self._write_title()
        self._write_col_headers()
        self._write_current_liabilities()
        self._write_current_assets()
        self._write_wc_requirement()
        self.ws.freeze_panes = f"{get_column_letter(year_col(1))}5"

    # ── Title & headers ───────────────────────────────────────────────────────

    def _write_title(self):
        last = year_col(self.n_years)
        self.write_section_header(1, COL_LABEL, last,
            "STATEMENT OF WORKING CAPITAL REQUIREMENT")
        self.write_section_header(2, COL_LABEL, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 14

    def _write_col_headers(self):
        self.write_column_header(3, COL_LABEL, "Particulars", fill=fill_solid(NAVY))
        self.write_column_header(3, COL_BASIS, "Basis",       fill=fill_solid(NAVY))
        for yr in range(1, self.n_years + 1):
            self.write_column_header(3, year_col(yr), f"Year {yr}",
                                     fill=fill_solid(NAVY))
        self.ws.row_dimensions[3].height = ROW_HEIGHT_SUBHDR

    # ── Current Liabilities ───────────────────────────────────────────────────

    def _write_current_liabilities(self):
        last = year_col(self.n_years)
        self.write_sub_header(WL.CL_HDR_ROW - 1, COL_LABEL, last,
                               "A  |  CURRENT LIABILITIES",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[WL.CL_HDR_ROW - 1].height = 13

        # Creditors for Raw Materials
        self._write_creditors_rm()
        # Creditors for Admin Expenses
        self._write_creditors_admin()
        # Total CL
        self._write_total_cl()

    def _write_creditors_rm(self):
        r         = WL.CREDITORS_ROW
        days_r    = WL.CREDITOR_DAYS_ROW
        asmp_days = self._asmp_abs("wc_creditor_rm")
        n_mats    = self.store.n_materials

        self.write_label(r, COL_LABEL, "  Creditors — Raw Materials",
                         fill=fill_alt)
        self.write_label(r, COL_BASIS, "₹ Lakhs", fill=fill_alt)
        self.write_label(days_r, COL_LABEL,
                         "    Days outstanding", fill=fill_white)
        self.write_formula(days_r, COL_BASIS,
                           f"={asmp_days}", fmt=FMT_INTEGER,
                           fill=fill_white, xsheet=True)

        cogs_row = EL.total_cogs_row(n_mats)
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            # Creditors = COGS / 365 × creditor_days
            formula = (f"=Expenses!{c}{cogs_row}"
                       f"/{get_column_letter(COL_BASIS)}{days_r}")
            self.write_formula(r, col, formula,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        for row in [r, days_r]:
            self.ws.row_dimensions[row].height = ROW_HEIGHT_DATA

        self.layout._map["W Cap"]["creditors_rm_row"] = (r, year_col(1))

    def _write_creditors_admin(self):
        r         = WL.ADMIN_CRED_ROW
        days_r    = WL.ADMIN_DAYS_ROW
        asmp_days = self._asmp_abs("wc_creditor_admin")

        self.write_label(r, COL_LABEL, "  Creditors — Admin Expenses",
                         fill=fill_alt)
        self.write_label(r, COL_BASIS, "₹ Lakhs", fill=fill_alt)
        self.write_label(days_r, COL_LABEL,
                         "    Days outstanding", fill=fill_white)
        self.write_formula(days_r, COL_BASIS,
                           f"={asmp_days}", fmt=FMT_INTEGER,
                           fill=fill_white, xsheet=True)

        # Admin expenses = SGA + Transport + Misc from Expenses sheet
        n_mats = self.store.n_materials
        sga_row = self.layout._map["Expenses"].get("sga_amount_pl_row",
                  (EL.sga_amount_row(n_mats), year_col(1)))[0]
        trn_row = self.layout._map["Expenses"].get("transport_pl_row",
                  (EL.transport_row(n_mats), year_col(1)))[0]
        msc_row = self.layout._map["Expenses"].get("misc_pl_row",
                  (EL.misc_row(n_mats), year_col(1)))[0]

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = (
                f"=(Expenses!{c}{sga_row}"
                f"+Expenses!{c}{trn_row}"
                f"+Expenses!{c}{msc_row})"
                f"/{get_column_letter(COL_BASIS)}{days_r}"
            )
            self.write_formula(r, col, formula,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        for row in [r, days_r]:
            self.ws.row_dimensions[row].height = ROW_HEIGHT_DATA

        self.layout._map["W Cap"]["creditors_admin_row"] = (r, year_col(1))

    def _write_total_cl(self):
        r      = WL.TOTAL_CL_ROW
        cr_rm  = WL.CREDITORS_ROW
        cr_adm = WL.ADMIN_CRED_ROW

        self.write_label(r, COL_LABEL, "TOTAL CURRENT LIABILITIES",
                         bold=True, fill=fill_solid(TEAL))
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = f"={c}{cr_rm}+{c}{cr_adm}"
            cell = self.ws.cell(row=r, column=col, value=formula)
            cell.font  = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill  = fill_solid(TEAL)
            cell.number_format = FMT_LAKHS
            cell.alignment = align_right
            cell.border = border_all_thin(TEAL)
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA + 2

        self.layout._map["W Cap"]["total_cl_row"] = (r, year_col(1))

    # ── Current Assets ────────────────────────────────────────────────────────

    def _write_current_assets(self):
        last = year_col(self.n_years)
        self.write_sub_header(WL.CA_HDR_ROW - 1, COL_LABEL, last,
                               "B  |  CURRENT ASSETS",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[WL.CA_HDR_ROW - 1].height = 13

        self._write_stock()
        self._write_debtors()
        self._write_total_ca()

    def _write_stock(self):
        r         = WL.STOCK_ROW
        days_r    = WL.STOCK_DAYS_ROW
        asmp_days = self._asmp_abs("wc_stock_rm")
        n_mats    = self.store.n_materials
        cogs_row  = EL.total_cogs_row(n_mats)

        self.write_label(r, COL_LABEL, "  Raw Material Stock",
                         fill=fill_alt)
        self.write_label(r, COL_BASIS, "₹ Lakhs", fill=fill_alt)
        self.write_label(days_r, COL_LABEL,
                         "    Stock days", fill=fill_white)
        self.write_formula(days_r, COL_BASIS,
                           f"={asmp_days}", fmt=FMT_INTEGER,
                           fill=fill_white, xsheet=True)

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = (f"=Expenses!{c}{cogs_row}"
                       f"/{get_column_letter(COL_BASIS)}{days_r}")
            self.write_formula(r, col, formula,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        for row in [r, days_r]:
            self.ws.row_dimensions[row].height = ROW_HEIGHT_DATA

        self.layout._map["W Cap"]["stock_row"] = (r, year_col(1))

    def _write_debtors(self):
        r         = WL.DEBTORS_ROW
        days_r    = WL.DEBTOR_DAYS_ROW
        asmp_days = self._asmp_abs("wc_debtor_days")

        self.write_label(r, COL_LABEL, "  Debtors (Receivables)",
                         fill=fill_alt)
        self.write_label(r, COL_BASIS, "₹ Lakhs", fill=fill_alt)
        self.write_label(days_r, COL_LABEL,
                         "    Debtor days", fill=fill_white)
        self.write_formula(days_r, COL_BASIS,
                           f"={asmp_days}", fmt=FMT_INTEGER,
                           fill=fill_white, xsheet=True)

        tr_row = RL.total_revenue_row(self.store.n_products)
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = (f"=Revenue!{c}{tr_row}"
                       f"/{get_column_letter(COL_BASIS)}{days_r}")
            self.write_formula(r, col, formula,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        for row in [r, days_r]:
            self.ws.row_dimensions[row].height = ROW_HEIGHT_DATA

        self.layout._map["W Cap"]["debtors_row"] = (r, year_col(1))

    def _write_total_ca(self):
        r       = WL.TOTAL_CA_ROW
        stk_r   = WL.STOCK_ROW
        dbt_r   = WL.DEBTORS_ROW

        self.write_label(r, COL_LABEL, "TOTAL CURRENT ASSETS",
                         bold=True, fill=fill_solid(TEAL))
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = f"={c}{stk_r}+{c}{dbt_r}"
            cell = self.ws.cell(row=r, column=col, value=formula)
            cell.font  = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill  = fill_solid(TEAL)
            cell.number_format = FMT_LAKHS
            cell.alignment = align_right
            cell.border = border_all_thin(TEAL)
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA + 2

        self.layout._map["W Cap"]["total_ca_row"] = (r, year_col(1))

    # ── WC Requirement ────────────────────────────────────────────────────────

    def _write_wc_requirement(self):
        r       = WL.WC_REQ_ROW
        ca_r    = WL.TOTAL_CA_ROW
        cl_r    = WL.TOTAL_CL_ROW

        last = year_col(self.n_years)
        self.write_sub_header(r - 1, COL_LABEL, last,
                               "C  |  NET WORKING CAPITAL REQUIREMENT",
                               fill=fill_solid(NAVY))
        self.ws.row_dimensions[r - 1].height = 13

        self.write_label(r, COL_LABEL,
                         "Net Working Capital Requirement (CA - CL)",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = f"={c}{ca_r}-{c}{cl_r}"
            cell = self.ws.cell(row=r, column=col, value=formula)
            cell.font  = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill  = fill_solid("1ABC9C")
            cell.number_format = FMT_LAKHS
            cell.alignment = align_right
            cell.border = border_all_thin(TEAL)
        self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA + 2

        self.layout._map["W Cap"]["wc_req_row"] = (r, year_col(1))

        # WC Interest (if applicable)
        wc_int_r    = WL.WC_INTEREST_ROW
        asmp_wc_int = self._asmp_abs("wc_interest_rate")

        self.write_label(wc_int_r, COL_LABEL,
                         "Working Capital Interest", fill=fill_white)
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            formula = f"={asmp_wc_int}*{c}{WL.BANK_LOAN_ROW}"
            self.write_formula(wc_int_r, col, formula,
                               fmt=FMT_LAKHS, fill=fill_white, xsheet=True)
        self.ws.row_dimensions[wc_int_r].height = ROW_HEIGHT_DATA
        self.layout._map["W Cap"]["wc_interest_row"] = (wc_int_r, year_col(1))

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _asmp_abs(self, logical: str) -> str:
        entry = self.layout._map["Assumption"].get(logical)
        if entry:
            r, c = entry
            return f"Assumption!${get_column_letter(c)}${r}"
        return "0"
