"""
sheet_revenue.py
─────────────────
Writes the Revenue sheet.

Structure:
  - Header rows: dates, months, days, capacity, utilisation
  - Dynamic product blocks: one block per product (5 rows each)
    Row 1: Total production volume (months × days × capacity × output_ratio × util)
    Row 2: Total production in liters (tons → liters)
    Row 3: Product liters (total liters × split_percent)
    Row 4: Price per unit (base price escalated annually)
    Row 5: Revenue (product_liters × price) / 10^5 → Lakhs

  - Total Revenue row: SUM of all product revenue rows

All constants (capacity, yield, utilisation ramp, price) come from Assumption.
Year columns: D = Year 1, E = Year 2, ... J = Year 7 (or further for N years).
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore, RevenueModelType
from core.layout_engine import (
    LayoutEngine, RevenueLayout as RL, AssumptionLayout as AL,
    col_letter, year_col, COL_LABEL, COL_BASIS, COL_YEAR_START
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_NUMBER, FMT_PCT_2, FMT_INTEGER,
    FMT_DATE, FMT_DATE_SHORT,
    font_formula, font_xsheet, font_input, font_subhdr, font_header,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid,
    fill_teal_hdr, fill_blue_hdr,
    border_all_thin, align_center, align_left, align_right,
    ROW_HEIGHT_HEADER, ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, LIGHT_GREY, WHITE,
    FORMULA_FONT_COLOR, XSHEET_FONT_COLOR,
)

ROWS_PER_PRODUCT = 5   # must match layout_engine.ROWS_PER_PRODUCT


class RevenueWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years    = store.n_years
        self.n_products = store.n_products
        super().__init__(wb, "Revenue")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = TEAL
        self.ws.column_dimensions[get_column_letter(COL_LABEL)].width = 32
        self.ws.column_dimensions[get_column_letter(COL_BASIS)].width = 12
        for yr in range(1, self.store.n_years + 1):
            self.ws.column_dimensions[get_column_letter(year_col(yr))].width = 16

    def _write(self):
        self._write_title()
        self._write_header_rows()
        self._write_product_blocks()
        self._write_total_revenue()
        self.ws.freeze_panes = f"{get_column_letter(COL_YEAR_START)}7"

    # ── Title ─────────────────────────────────────────────────────────────────

    def _write_title(self):
        last = year_col(self.n_years)
        self.write_section_header(1, COL_LABEL, last,
            "STATEMENT OF TOTAL REVENUE")
        self.write_section_header(2, COL_LABEL, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 14

    # ── Header rows (rows 3–6) ────────────────────────────────────────────────

    def _write_header_rows(self):
        """Write rows for Particulars, Basis, Year dates, Months, Days."""
        # Row 3 — blank separator
        self.ws.row_dimensions[3].height = 6

        # Row 4 — column headers: Particulars | Basis | Year dates
        self.write_column_header(4, COL_LABEL, "Particulars", fill=fill_solid(NAVY))
        self.write_column_header(4, COL_BASIS, "Basis",       fill=fill_solid(NAVY))
        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            # Date formula: Year 1 is hardcoded, subsequent years use EOMONTH
            if yr == 1:
                # First year end date — reference from store operation start
                start = self.store.project_profile.operation_start_date or "2027-03"
                year_part = int(start.split("-")[0]) if start else 2027
                month_part = int(start.split("-")[1]) if start else 3
                # Build a date serial for EOMONTH — use DATE formula
                date_formula = f'=DATE({year_part},{month_part},1)+365-1'
                self.write_formula(4, col, date_formula,
                                   fmt=FMT_DATE_SHORT,
                                   fill=fill_solid(BLUE))
                cell = self.ws.cell(row=4, column=col)
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color=WHITE)
            else:
                prev_col = get_column_letter(year_col(yr - 1))
                self.write_formula(4, col,
                                   f"=EOMONTH({prev_col}4,12)",
                                   fmt=FMT_DATE_SHORT,
                                   fill=fill_solid(BLUE))
                cell = self.ws.cell(row=4, column=col)
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color=WHITE)
        self.ws.row_dimensions[4].height = ROW_HEIGHT_SUBHDR

        # Row 5 — Months in operation (hardcoded 12 each year)
        self.write_label(5, COL_LABEL, "Months in Operation", fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            self.write_formula(5, year_col(yr), "=12",
                               fmt=FMT_INTEGER, fill=fill_alt)

        # Row 6 — Working days per month (from Assumption)
        asmp_days_ref = self._asmp_abs("cap_working_days")
        self.write_label(6, COL_LABEL, "Working Days per Month", fill=fill_white)
        for yr in range(1, self.n_years + 1):
            self.write_formula(6, year_col(yr),
                               f"={asmp_days_ref}",
                               fmt=FMT_INTEGER, fill=fill_white, xsheet=True)

        # Row 7 — blank separator before capacity section
        self.ws.row_dimensions[7].height = 6

        # Row 8 — Total capacity per day (from Assumption: capacity_per_day of first product)
        # For multi-product from same input, capacity is shared
        if self.n_products > 0:
            prod0 = self.store.revenue_model.products[0]
            self.write_label(8, COL_LABEL, "Total Production Capacity per Day")
            self.write_label(8, COL_BASIS, "input units/day")
            for yr in range(1, self.n_years + 1):
                self.write_formula(8, year_col(yr),
                                   f"={prod0.capacity_per_day}",
                                   fmt=FMT_NUMBER)

        # Row 9 — blank separator
        self.ws.row_dimensions[9].height = 6

        # Row 10 — Operational capacity (utilisation) — ramps up each year
        self.write_label(10, COL_LABEL, "Operational Capacity (Utilisation %)",
                         fill=fill_solid("FFF3CD"))
        self.write_label(10, COL_BASIS, "fraction", fill=fill_solid("FFF3CD"))

        asmp_util_y1  = self._asmp_abs("cap_year1_util")
        asmp_increment= self._asmp_abs("cap_annual_increment")
        asmp_max_util = self._asmp_abs("cap_max_util")

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            if yr == 1:
                # Year 1: direct reference to Assumption
                formula = f"={asmp_util_y1}"
            else:
                # Year N: IF(prev < max, prev + increment, prev)
                prev_col = get_column_letter(year_col(yr - 1))
                formula = (f"=IF({prev_col}10<{asmp_max_util},"
                           f"{prev_col}10+{asmp_increment},"
                           f"{prev_col}10)")
            self.write_formula(10, col, formula, fmt=FMT_PCT_2,
                               fill=fill_solid("FFF3CD"), xsheet=(yr == 1))

        self.ws.row_dimensions[10].height = ROW_HEIGHT_DATA

    # ── Product blocks ────────────────────────────────────────────────────────

    def _write_product_blocks(self):
        """Write one 5-row block per product."""
        products = self.store.revenue_model.products

        for i, prod in enumerate(products):
            base_row = RL.prod_volume_row(i)    # rows: base, base+1, ..., base+4
            vol_row  = base_row
            lit_row  = base_row + 1
            spl_row  = base_row + 2
            prc_row  = base_row + 3
            rev_row  = base_row + 4

            fill_bg = fill_white if i % 2 == 0 else fill_alt

            # ── Section header for this product ──────────────────────────────
            last_col = year_col(self.n_years)
            self.write_sub_header(vol_row - 1, COL_LABEL, last_col,
                                  f"  {prod.name}",
                                  fill=fill_solid("2E75B6"))
            self.ws.row_dimensions[vol_row - 1].height = 13

            # ── Row vol_row: Total production volume ─────────────────────────
            self.write_label(vol_row, COL_LABEL,
                             f"  Total Production ({prod.unit})")
            self.write_label(vol_row, COL_BASIS, prod.unit)

            # Formula: months × days × capacity × output_ratio × split × util
            # = row5 * row6 * capacity_per_day * output_ratio * split_percent * util_row10
            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                c   = get_column_letter(col)
                formula = (
                    f"={c}5*{c}6"
                    f"*{prod.capacity_per_day}"
                    f"*{prod.output_ratio}"
                    f"*{prod.split_percent}"
                    f"*{c}10"
                )
                self.write_formula(vol_row, col, formula,
                                   fmt=FMT_NUMBER, fill=fill_bg)
            self.ws.row_dimensions[vol_row].height = ROW_HEIGHT_DATA

            # ── Row lit_row: Convert to liters (× 1000 for tons) ─────────────
            self.write_label(lit_row, COL_LABEL,
                             f"  {prod.name} — Liters")
            self.write_label(lit_row, COL_BASIS, "liters")

            unit_multiplier = 1000 if prod.unit.lower() in ["tons","tonnes","ton","tonne"] else 1

            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                c   = get_column_letter(col)
                formula = f"={c}{vol_row}*{unit_multiplier}"
                self.write_formula(lit_row, col, formula,
                                   fmt=FMT_NUMBER, fill=fill_bg)
            self.ws.row_dimensions[lit_row].height = ROW_HEIGHT_DATA

            # ── Row spl_row: Split — same as vol for single-input products ───
            # (kept for compatibility — vol already incorporates split)
            self.write_label(spl_row, COL_LABEL,
                             f"  {prod.name} — Net Output")
            self.write_label(spl_row, COL_BASIS, prod.unit)
            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                c   = get_column_letter(col)
                formula = f"={c}{lit_row}"
                self.write_formula(spl_row, col, formula,
                                   fmt=FMT_NUMBER, fill=fill_bg)
            self.ws.row_dimensions[spl_row].height = ROW_HEIGHT_DATA

            # ── Row prc_row: Price per unit (escalated annually) ─────────────
            self.write_label(prc_row, COL_LABEL,
                             f"  Price per {prod.unit} (₹)")
            self.write_label(prc_row, COL_BASIS, f"₹/{prod.unit}")

            # Get price from Assumption sheet
            asmp_price = f"=Assumption!$E${AL.rev_price_row(i)}"
            asmp_esc   = f"Assumption!$E${AL.rev_escalation_row(i)}"

            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                if yr == 1:
                    formula = asmp_price
                    self.write_formula(prc_row, col, formula,
                                       fmt=FMT_NUMBER, fill=fill_solid("E8F8F5"),
                                       xsheet=True)
                else:
                    prev_c = get_column_letter(year_col(yr - 1))
                    formula = f"={prev_c}{prc_row}*(1+{asmp_esc})"
                    self.write_formula(prc_row, col, formula,
                                       fmt=FMT_NUMBER, fill=fill_solid("E8F8F5"),
                                       xsheet=True)
            self.ws.row_dimensions[prc_row].height = ROW_HEIGHT_DATA

            # ── Row rev_row: Revenue = (output × price) / 10^5 → Lakhs ──────
            self.write_label(rev_row, COL_LABEL,
                             f"  {prod.name} Revenue",
                             bold=True, fill=fill_solid("D5F5E3"))
            self.write_label(rev_row, COL_BASIS, "₹ Lakhs",
                             fill=fill_solid("D5F5E3"))

            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                c   = get_column_letter(col)
                formula = f"=({c}{spl_row}*{c}{prc_row})/100000"
                self.write_total(rev_row, col, formula, bold=False)
            self.ws.row_dimensions[rev_row].height = ROW_HEIGHT_DATA

            # Blank row between products
            if i < self.n_products - 1:
                self.ws.row_dimensions[rev_row + 1].height = 4

    # ── Total Revenue ─────────────────────────────────────────────────────────

    def _write_total_revenue(self):
        total_row = RL.total_revenue_row(self.n_products)
        last_col  = year_col(self.n_years)

        # Section header
        self.write_section_header(total_row - 1, COL_LABEL, last_col,
                                   "TOTAL REVENUE")
        self.ws.row_dimensions[total_row - 1].height = 14

        self.write_label(total_row, COL_LABEL, "Total Revenue from Operations",
                         bold=True, fill=fill_solid("1ABC9C"))
        self.write_label(total_row, COL_BASIS, "₹ Lakhs",
                         fill=fill_solid("1ABC9C"))

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)

            # Sum all product revenue rows
            rev_rows = [RL.prod_revenue_row(i) for i in range(self.n_products)]
            sum_parts = "+".join(f"{c}{r}" for r in rev_rows)
            formula = f"={sum_parts}"

            cell = self.ws.cell(row=total_row, column=col, value=formula)
            cell.font = Font(name="Arial", size=10, bold=True,
                             color=WHITE)
            cell.fill = fill_solid("1ABC9C")
            cell.number_format = FMT_LAKHS
            cell.alignment = align_right
            cell.border = border_all_thin(TEAL)
        self.ws.row_dimensions[total_row].height = ROW_HEIGHT_DATA + 2

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _asmp_abs(self, logical: str) -> str:
        """Return absolute Assumption reference like Assumption!$E$4."""
        r, c = self.layout._map["Assumption"][logical]
        return f"Assumption!${get_column_letter(c)}${r}"
