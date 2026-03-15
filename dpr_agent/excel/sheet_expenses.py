"""
sheet_expenses.py
──────────────────
Writes the Expenses sheet.

Two sections:
  A) Cost of Sales — raw material consumption × price per unit, escalated annually
  B) Operating Overheads — R&M, insurance, marketing, power, SGA, transport, misc

All rates reference Assumption. No hardcoded constants.
Revenue and Net Block are pulled from Revenue and Depreciation sheets.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore
from core.layout_engine import (
    LayoutEngine, ExpensesLayout as EL, AssumptionLayout as AL,
    RevenueLayout as RL, DepreciationLayout as DL,
    col_letter, year_col, COL_LABEL, COL_BASIS, COL_YEAR_START,
    ROWS_PER_MATERIAL
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_NUMBER, FMT_PCT_2, FMT_INTEGER,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid, fill_amber,
    border_all_thin, align_center, align_left, align_right,
    ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA, ROW_HEIGHT_SECTION,
    NAVY, BLUE, TEAL, WHITE, LIGHT_GREY,
)


class ExpensesWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years = store.n_years
        self.n_mats  = store.n_materials
        self._n_asset_classes = len(set(
            a.category for a in store.capital_means.assets
        ))
        super().__init__(wb, "Expenses")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = BLUE
        self.ws.column_dimensions[get_column_letter(COL_LABEL)].width = 36
        self.ws.column_dimensions[get_column_letter(COL_BASIS)].width = 10
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[get_column_letter(year_col(yr))].width = 15

    def _write(self):
        self._write_title()
        self._write_col_headers()
        self._write_raw_materials()
        self._write_total_cogs()
        self._write_overhead_refs()
        self._write_overheads()
        self.ws.freeze_panes = f"{get_column_letter(COL_YEAR_START)}5"

    # ── Title & col headers ───────────────────────────────────────────────────

    def _write_title(self):
        last = year_col(self.n_years)
        self.write_section_header(1, COL_LABEL, last,
            "STATEMENT OF CALCULATION OF EXPENSES")
        self.write_section_header(2, COL_LABEL, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 14

    def _write_col_headers(self):
        self.write_column_header(3, COL_LABEL, "Particulars", fill=fill_solid(NAVY))
        self.write_column_header(3, COL_BASIS, "Basis",       fill=fill_solid(NAVY))
        for yr in range(1, self.n_years + 1):
            self.write_column_header(3, year_col(yr), f"Year {yr}",
                                     fill=fill_solid(NAVY))
        self.ws.row_dimensions[3].height = ROW_HEIGHT_SUBHDR

    # ── A: Raw Material Cost of Sales ─────────────────────────────────────────

    def _write_raw_materials(self):
        last = year_col(self.n_years)
        self.write_sub_header(EL.RM_BASE_ROW - 1, COL_LABEL, last,
                               "A  |  COST OF SALES — RAW MATERIALS",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[EL.RM_BASE_ROW - 1].height = 13

        for i, mat in enumerate(self.store.cost_structure.raw_materials):
            qty_row  = EL.rm_qty_row(i)
            prc_row  = EL.rm_price_row(i)
            cost_row = EL.rm_cost_row(i)
            fill_bg  = fill_white if i % 2 == 0 else fill_alt

            # ── Quantity row: output_liters × input_per_output ───────────────
            self.write_label(qty_row, COL_LABEL,
                             f"  {mat.name} — Quantity Used",
                             fill=fill_bg)
            self.write_label(qty_row, COL_BASIS, mat.unit, fill=fill_bg)

            # Pull total output from Revenue total_revenue_row logic
            # Use production volume: Revenue!vol_row × input_per_output
            # vol_row of product 0 gives total tons produced
            rev_vol_row = RL.prod_volume_row(0)
            asmp_ipu = self._asmp_abs(f"rm_input_per_output_m{i}")

            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                c   = get_column_letter(col)
                # quantity = production_volume (all products combined) × input_per_output
                # production volume = sum of all product volume rows × 1000 (tons→kg)
                vol_refs = "+".join(
                    f"Revenue!{c}{RL.prod_volume_row(j)}"
                    for j in range(self.store.n_products)
                )
                formula = f"=({vol_refs})*1000*{asmp_ipu}"
                self.write_formula(qty_row, col, formula,
                                   fmt=FMT_NUMBER, fill=fill_bg, xsheet=True)

            # ── Price row: base price escalated annually ──────────────────────
            self.write_label(prc_row, COL_LABEL,
                             f"  {mat.name} — Price per {mat.unit}",
                             fill=fill_bg)
            self.write_label(prc_row, COL_BASIS,
                             f"₹/{mat.unit}", fill=fill_bg)

            asmp_price = self._asmp_abs(f"rm_price_m{i}")
            asmp_esc   = self._asmp_abs(f"rm_escalation_m{i}")

            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                if yr == 1:
                    formula = f"={asmp_price}"
                    self.write_formula(prc_row, col, formula,
                                       fmt=FMT_NUMBER, fill=fill_solid("E8F8F5"),
                                       xsheet=True)
                else:
                    prev_c  = get_column_letter(year_col(yr - 1))
                    formula = f"={prev_c}{prc_row}*(1+{asmp_esc})"
                    self.write_formula(prc_row, col, formula,
                                       fmt=FMT_NUMBER, fill=fill_solid("E8F8F5"),
                                       xsheet=True)

            # ── Cost row: (qty × price) / 100000 → Lakhs ─────────────────────
            self.write_label(cost_row, COL_LABEL,
                             f"  {mat.name} — Cost",
                             bold=True, fill=fill_solid("D5F5E3"))
            self.write_label(cost_row, COL_BASIS, "₹ Lakhs",
                             fill=fill_solid("D5F5E3"))

            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                c   = get_column_letter(col)
                formula = f"=({c}{qty_row}*{c}{prc_row})/100000"
                self.write_total(cost_row, col, formula, bold=False)

            # Spacing
            for r in [qty_row, prc_row, cost_row]:
                self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    # ── Total COGS ────────────────────────────────────────────────────────────

    def _write_total_cogs(self):
        n    = self.n_mats
        row  = EL.total_cogs_row(n)
        last = year_col(self.n_years)

        self.write_label(row, COL_LABEL, "TOTAL COST OF SALES",
                         bold=True, fill=fill_solid(TEAL))
        self.write_label(row, COL_BASIS, "₹ Lakhs", fill=fill_solid(TEAL))

        for yr in range(1, self.n_years + 1):
            col = year_col(yr)
            c   = get_column_letter(col)
            cost_rows = [EL.rm_cost_row(i) for i in range(n)]
            formula   = "=" + "+".join(f"{c}{r}" for r in cost_rows)
            cell = self.ws.cell(row=row, column=col, value=formula)
            cell.font   = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill   = fill_solid(TEAL)
            cell.number_format = FMT_LAKHS
            cell.alignment = align_right
            cell.border = border_all_thin(TEAL)
        self.ws.row_dimensions[row].height = ROW_HEIGHT_DATA + 2

        # Store for P&L reference
        self.layout._map["Expenses"]["total_cogs_row"] = (row, year_col(1))

    # ── Reference rows (Revenue and Net Block) ────────────────────────────────

    def _write_overhead_refs(self):
        """Write helper rows: Revenue pull and Net Block pull for ratio calculations."""
        n        = self.n_mats
        rev_row  = EL.revenue_ref_row(n)
        nb_row   = EL.net_block_ref_row(n)

        # Revenue reference
        self.write_label(rev_row, COL_LABEL,
                         "  [Ref] Total Revenue", fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            col    = year_col(yr)
            c      = get_column_letter(col)
            tr_row = RL.total_revenue_row(self.store.n_products)
            formula = f"=Revenue!{c}{tr_row}"
            self.write_formula(rev_row, col, formula,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[rev_row].height = ROW_HEIGHT_DATA

        # Net Block reference from Depreciation
        self.write_label(nb_row, COL_LABEL,
                         "  [Ref] Net Fixed Assets (WDV)", fill=fill_alt)
        dep_nb_row = DL.net_block_row(self._n_asset_classes)
        for yr in range(1, self.n_years + 1):
            col  = year_col(yr)
            c    = get_column_letter(col)
            formula = f"=Depreciation!{c}{dep_nb_row}"
            self.write_formula(nb_row, col, formula,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[nb_row].height = ROW_HEIGHT_DATA

        # Store reference rows
        self.layout._map["Expenses"]["revenue_ref_row"] = (rev_row, year_col(1))
        self.layout._map["Expenses"]["net_block_ref_row"] = (nb_row, year_col(1))

    # ── B: Operating Overheads ────────────────────────────────────────────────

    def _write_overheads(self):
        n    = self.n_mats
        last = year_col(self.n_years)

        self.write_sub_header(EL.overhead_base_row(n) - 1,
                               COL_LABEL, last,
                               "B  |  OPERATING OVERHEAD EXPENSES",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[EL.overhead_base_row(n) - 1].height = 13

        rev_row = EL.revenue_ref_row(n)
        nb_row  = EL.net_block_ref_row(n)

        # Each overhead item: rate row + amount row
        items = [
            # (rate_row_key, amt_row_key, label, basis, asmp_rate_key, asmp_esc_key, base_on)
            ("rm_rate",   "rm_amount",   "Repair & Maintenance",
             "% of Net Fixed Assets",
             "exp_rm_pct_fa", "exp_rm_escalation", "netblock"),

            ("ins_rate",  "ins_amount",  "Insurance Expenses",
             "% of Net Fixed Assets",
             "exp_ins_pct_fa", "exp_ins_escalation", "netblock"),

            ("mkt_rate",  "mkt_amount",  "Marketing Expenses",
             "% of Revenue",
             "exp_mkt_pct_rev", "exp_mkt_escalation", "revenue"),

            ("power_rate","power_amount","Power & Fuel",
             "% of Revenue",
             "exp_power_pct_rev", "exp_power_escalation", "revenue"),

            ("sga_base",  "sga_amount",  "Selling, General & Admin",
             "₹ Lakhs (base)",
             "exp_sga_base", "exp_sga_esc", "absolute"),

            ("transport", None,          "Transportation Cost",
             "₹ Lakhs",
             "exp_transport_base", "exp_transport_esc", "absolute_escalate"),

            ("misc",      None,          "Miscellaneous Expenses",
             "₹ Lakhs",
             "exp_misc_base", "exp_misc_esc", "absolute_escalate"),
        ]

        for rate_key, amt_key, label, basis, asmp_rate, asmp_esc, calc_type in items:
            r_row = self.layout.row("Expenses", rate_key)
            fill_bg = fill_alt if r_row % 2 == 0 else fill_white

            asmp_rate_ref = self._asmp_abs(asmp_rate)
            asmp_esc_ref  = self._asmp_abs(asmp_esc)

            # Rate / base value row
            self.write_label(r_row, COL_LABEL, f"  {label} — Rate/Base",
                             fill=fill_bg)
            self.write_label(r_row, COL_BASIS, basis, fill=fill_bg)

            for yr in range(1, self.n_years + 1):
                col = year_col(yr)
                if yr == 1:
                    formula = f"={asmp_rate_ref}"
                else:
                    prev_c  = get_column_letter(year_col(yr - 1))
                    formula = f"={prev_c}{r_row}*(1+{asmp_esc_ref})"
                self.write_formula(r_row, col, formula,
                                   fmt=FMT_PCT_2 if calc_type not in ("absolute","absolute_escalate") else FMT_LAKHS,
                                   fill=fill_solid("E8F8F5"), xsheet=True)
            self.ws.row_dimensions[r_row].height = ROW_HEIGHT_DATA

            # Amount row (if separate)
            if amt_key:
                a_row = self.layout.row("Expenses", amt_key)
                self.write_label(a_row, COL_LABEL,
                                 f"  {label}", bold=True,
                                 fill=fill_solid("D5F5E3"))
                self.write_label(a_row, COL_BASIS, "₹ Lakhs",
                                 fill=fill_solid("D5F5E3"))

                for yr in range(1, self.n_years + 1):
                    col = year_col(yr)
                    c   = get_column_letter(col)
                    if calc_type == "netblock":
                        formula = f"={c}{r_row}*{c}{nb_row}"
                    elif calc_type == "revenue":
                        formula = f"={c}{r_row}*{c}{rev_row}"
                    elif calc_type == "absolute":
                        formula = f"={c}{r_row}"
                    else:
                        formula = f"={c}{r_row}"
                    self.write_total(a_row, col, formula, bold=False)
                self.ws.row_dimensions[a_row].height = ROW_HEIGHT_DATA

                # Store amount row for P&L reference
                self.layout._map["Expenses"][f"{amt_key}_pl_row"] = (
                    a_row, year_col(1)
                )
            else:
                # transport and misc: the rate row IS the amount
                self.layout._map["Expenses"][f"{rate_key}_pl_row"] = (
                    r_row, year_col(1)
                )

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _asmp_abs(self, logical: str) -> str:
        entry = self.layout._map["Assumption"].get(logical)
        if entry:
            r, c = entry
            return f"Assumption!${get_column_letter(c)}${r}"
        return "0"
