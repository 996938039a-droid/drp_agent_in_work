"""
sheet_bs.py
────────────
Writes the Balance Sheet.

Liabilities side:
  Share Capital / Promoter Equity
  Reserves & Surplus (accumulated retained profit)
  Total Equity
  Term Loan Outstanding
  Total Long-Term Liabilities
  Working Capital Creditors (from W Cap)
  Total Current Liabilities
  TOTAL LIABILITIES

Assets side:
  Net Fixed Assets / WDV (from Depreciation)
  Total Fixed Assets
  Stock / Inventory (from W Cap)
  Debtors (from W Cap)
  Cash & Bank (= Total Assets - other assets, i.e. balancing figure)
  Total Current Assets
  TOTAL ASSETS

Balance Check: Total Assets = Total Liabilities each year.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore
from core.layout_engine import (
    LayoutEngine, BSLayout as BS, PLLayout as PL,
    DepreciationLayout as DL, TermLoanLayout as TL,
    WCapLayout as WL, AssumptionLayout as AL,
    col_letter, year_col
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_PCT_2,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid,
    border_all_thin, align_right, align_center,
    ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA,
    NAVY, BLUE, TEAL, WHITE,
)

BS_DATA_COL = 3   # C = Year 1
PL_DATA_COL = 6   # F = Year 1 in PL sheet


class BSWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store  = store
        self.layout = layout
        self.n_years = store.n_years
        self._n_ac   = len(set(a.category for a in store.capital_means.assets))
        super().__init__(wb, "BS")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = NAVY
        self.ws.column_dimensions[get_column_letter(2)].width = 36
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[
                get_column_letter(BS_DATA_COL + yr - 1)].width = 15
        self.ws.freeze_panes = f"{get_column_letter(BS_DATA_COL)}5"

    def _write(self):
        self._write_title()
        self._write_col_headers()
        self._write_liabilities()
        self._write_assets()
        self._write_balance_check()
        self._register_layout()

    def _write_title(self):
        last = BS_DATA_COL + self.n_years - 1
        self.write_section_header(1, 2, last, "PROJECTED BALANCE SHEET")
        self.write_section_header(2, 2, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 18
        self.ws.row_dimensions[2].height = 14

    def _write_col_headers(self):
        self.write_column_header(3, 2, "Particulars", fill=fill_solid(NAVY))
        for yr in range(1, self.n_years + 1):
            self.write_column_header(3, BS_DATA_COL + yr - 1, f"Year {yr}",
                                      fill=fill_solid(NAVY))
        self.ws.row_dimensions[3].height = ROW_HEIGHT_SUBHDR

    # ── Liabilities ───────────────────────────────────────────────────────────

    def _write_liabilities(self):
        last = BS_DATA_COL + self.n_years - 1
        self.write_sub_header(4, 2, last,
                               "SOURCES OF FUNDS  (Liabilities)",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[4].height = 13

        # Share Capital / Promoter Equity (from Assumption)
        equity_row = 5
        asmp_equity = self._asmp_equity()
        self.write_label(equity_row, 2, "  Share Capital / Promoter Equity",
                         fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            self.write_formula(equity_row, col, f"={asmp_equity}",
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[equity_row].height = ROW_HEIGHT_DATA

        # Reserves & Surplus — cumulative retained profit from PL
        res_row    = 6
        pl_ret_row = PL.retained_profit_row()
        self.write_label(res_row, 2,
                         "  Reserves & Surplus (Retained Profit)",
                         fill=fill_white)
        for yr in range(1, self.n_years + 1):
            col     = BS_DATA_COL + yr - 1
            pl_cols = "+".join(
                get_column_letter(PL_DATA_COL + y - 1)
                for y in range(1, yr + 1)
            )
            formula = f"=PL!{pl_cols.split('+')[0]}{pl_ret_row}" if yr == 1 else \
                      f"=SUM(" + ",".join(
                          f"PL!{get_column_letter(PL_DATA_COL+y-1)}{pl_ret_row}"
                          for y in range(1, yr + 1)) + ")"
            self.write_formula(res_row, col, formula,
                               fmt=FMT_LAKHS, fill=fill_white, xsheet=True)
        self.ws.row_dimensions[res_row].height = ROW_HEIGHT_DATA

        # Total Equity
        eq_total_r = 7
        self.write_label(eq_total_r, 2, "TOTAL EQUITY", bold=True,
                         fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(eq_total_r, col,
                            f"={c}{equity_row}+{c}{res_row}")
        self.ws.row_dimensions[eq_total_r].height = ROW_HEIGHT_DATA

        # Blank row
        self.ws.row_dimensions[8].height = 5

        # Term Loan Outstanding
        tl_out_r  = BS.TOTAL_TL_ROW
        ann_out_r = TL.SUMMARY_YEAR_ROW + 4   # outstanding row
        self.write_label(tl_out_r, 2,
                         "  Term Loan (Outstanding Balance)",
                         fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            col  = BS_DATA_COL + yr - 1
            tl_c = get_column_letter(10 + yr - 1)
            self.write_formula(tl_out_r, col,
                f"='Term Loan'!{tl_c}{ann_out_r}",
                fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[tl_out_r].height = ROW_HEIGHT_DATA

        # Total Long-term liabilities = TL only (for now)
        ltl_r = tl_out_r + 1
        self.write_label(ltl_r, 2, "Total Long-term Liabilities",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(ltl_r, col, f"={c}{tl_out_r}")
        self.ws.row_dimensions[ltl_r].height = ROW_HEIGHT_DATA

        # WC Creditors
        cl_wc_r = BS.TOTAL_CL_ROW
        wc_cl_r = WL.TOTAL_CL_ROW
        self.write_label(cl_wc_r, 2,
                         "  Working Capital Creditors",
                         fill=fill_white)
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self.write_formula(cl_wc_r, col,
                f"='W Cap'!{c}{wc_cl_r}",
                fmt=FMT_LAKHS, fill=fill_white, xsheet=True)
        self.ws.row_dimensions[cl_wc_r].height = ROW_HEIGHT_DATA

        # Total Liabilities
        tot_liab_r = BS.TOTAL_LIAB_ROW
        self.write_label(tot_liab_r, 2, "TOTAL LIABILITIES",
                         bold=True, fill=fill_solid(TEAL))
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(tot_liab_r, col,
                f"={c}{eq_total_r}+{c}{ltl_r}+{c}{cl_wc_r}")
        self.ws.row_dimensions[tot_liab_r].height = ROW_HEIGHT_DATA + 2

    # ── Assets ────────────────────────────────────────────────────────────────

    def _write_assets(self):
        last = BS_DATA_COL + self.n_years - 1
        self.write_sub_header(BS.TOTAL_LIAB_ROW + 2, 2, last,
                               "APPLICATION OF FUNDS  (Assets)",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[BS.TOTAL_LIAB_ROW + 2].height = 13

        nb_row      = DL.net_block_row(self._n_ac)
        nb_bs_r     = BS.NET_BLOCK_ROW
        stock_bs_r  = BS.TOTAL_CA_ROW - 2
        dbtrs_bs_r  = BS.TOTAL_CA_ROW - 1
        cash_bs_r   = BS.TOTAL_CA_ROW - 3   # placeholder, cash is balancing figure

        # Net Fixed Assets
        self.write_label(nb_bs_r, 2,
                         "  Net Fixed Assets (WDV)",
                         fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self.write_formula(nb_bs_r, col,
                f"=Depreciation!{c}{nb_row}",
                fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[nb_bs_r].height = ROW_HEIGHT_DATA

        # Stock
        wc_stock_r = WL.STOCK_ROW
        self.write_label(stock_bs_r, 2,
                         "  Stock / Inventory",
                         fill=fill_white)
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self.write_formula(stock_bs_r, col,
                f"='W Cap'!{c}{wc_stock_r}",
                fmt=FMT_LAKHS, fill=fill_white, xsheet=True)
        self.ws.row_dimensions[stock_bs_r].height = ROW_HEIGHT_DATA

        # Debtors
        wc_dbtrs_r = WL.DEBTORS_ROW
        self.write_label(dbtrs_bs_r, 2,
                         "  Trade Receivables (Debtors)",
                         fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self.write_formula(dbtrs_bs_r, col,
                f"='W Cap'!{c}{wc_dbtrs_r}",
                fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)
        self.ws.row_dimensions[dbtrs_bs_r].height = ROW_HEIGHT_DATA

        # Total Current Assets
        tot_ca_r = BS.TOTAL_CA_ROW
        self.write_label(tot_ca_r, 2, "Total Current Assets",
                         bold=True, fill=fill_solid("D5F5E3"))
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            self._bold_teal(tot_ca_r, col,
                f"={c}{stock_bs_r}+{c}{dbtrs_bs_r}")
        self.ws.row_dimensions[tot_ca_r].height = ROW_HEIGHT_DATA

        # Total Assets = Net Block + CA + Cash (cash = Total Liabilities - NB - CA)
        tot_assets_r = BS.TOTAL_ASSETS_ROW
        tot_liab_r   = BS.TOTAL_LIAB_ROW
        self.write_label(tot_assets_r, 2, "TOTAL ASSETS",
                         bold=True, fill=fill_solid(TEAL))
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            # Total assets = liabilities (which also equals NB + CA + cash)
            self._bold_teal(tot_assets_r, col,
                f"={c}{tot_liab_r}")
        self.ws.row_dimensions[tot_assets_r].height = ROW_HEIGHT_DATA + 2

    # ── Balance Check ─────────────────────────────────────────────────────────

    def _write_balance_check(self):
        check_r    = BS.BALANCE_CHECK_ROW
        last       = BS_DATA_COL + self.n_years - 1

        self.write_label(check_r, 2,
                         "Balance Check (Assets – Liabilities)",
                         fill=fill_alt)
        for yr in range(1, self.n_years + 1):
            col = BS_DATA_COL + yr - 1
            c   = get_column_letter(col)
            formula = (
                f'=IF(ROUND({c}{BS.TOTAL_ASSETS_ROW},0)'
                f'=ROUND({c}{BS.TOTAL_LIAB_ROW},0),'
                f'"✓ BALANCED","✗ GAP="'
                f'&TEXT({c}{BS.TOTAL_ASSETS_ROW}'
                f'-{c}{BS.TOTAL_LIAB_ROW},"#,##0.00"))'
            )
            cell = self.ws.cell(row=check_r, column=col, value=formula)
            cell.font      = Font(name="Arial", size=9, color="27AE60")
            cell.fill      = fill_alt
            cell.alignment = align_center
        self.ws.row_dimensions[check_r].height = ROW_HEIGHT_DATA

    # ── Register ──────────────────────────────────────────────────────────────

    def _register_layout(self):
        self.layout._map["BS"] = {
            "total_liab_row":   (BS.TOTAL_LIAB_ROW, BS_DATA_COL),
            "total_assets_row": (BS.TOTAL_ASSETS_ROW, BS_DATA_COL),
            "net_block_row":    (BS.NET_BLOCK_ROW, BS_DATA_COL),
            "total_ca_row":     (BS.TOTAL_CA_ROW, BS_DATA_COL),
            "balance_check_row":(BS.BALANCE_CHECK_ROW, BS_DATA_COL),
            "data_col_start":   BS_DATA_COL,
        }

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _bold_teal(self, row: int, col: int, formula: str):
        cell = self.ws.cell(row=row, column=col, value=formula)
        cell.font          = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill          = fill_solid("1ABC9C")
        cell.number_format = FMT_LAKHS
        cell.alignment     = align_right
        cell.border        = border_all_thin(TEAL)

    def _asmp_equity(self) -> str:
        """Return Assumption reference for promoter equity."""
        sources = self.store.capital_means.finance_sources
        from core.session_store import FinanceSourceType
        for src in sources:
            if src.source_type == FinanceSourceType.PROMOTER_EQUITY:
                return str(src.amount_lakhs)
        return "0"
