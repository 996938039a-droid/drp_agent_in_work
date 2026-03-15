"""
styles.py
─────────
Central styling definitions for the DPR workbook.
All sheets import from here — no inline style definitions allowed.

Color convention (industry standard):
  BLUE text   (0000FF) → hardcoded input cells (user-entered values)
  BLACK text  (000000) → formula / calculated cells
  GREEN text  (008000) → cross-sheet link cells
  RED text    (FF0000) → external link cells (not used in this model)
  YELLOW bg   (FFFF00) → key assumption cells needing attention

Section header rows: navy background, white bold text
Sub-header rows: light blue background, dark bold text
Data rows: alternating white / very light grey
"""

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as num_format
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ─── Colour palette ──────────────────────────────────────────────────────────
NAVY        = "1F3864"
BLUE        = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
TEAL        = "1ABC9C"
LIGHT_TEAL  = "D1F5EC"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F2F2F2"
MID_GREY    = "D9D9D9"
DARK_GREY   = "595959"
BLACK       = "000000"
AMBER       = "FFF2CC"   # yellow highlight bg
AMBER_BORDER= "FFD966"

# Input / formula colour convention
INPUT_FONT_COLOR   = "FF0000FF"  # blue  → user input (ARGB full format)
FORMULA_FONT_COLOR = "FF000000"  # black → calculated
XSHEET_FONT_COLOR  = "FF008000"  # green → cross-sheet link
HEADER_BG          = NAVY
HEADER_FG          = WHITE
SUBHDR_BG          = LIGHT_BLUE
SUBHDR_FG          = NAVY
ALT_ROW_BG         = LIGHT_GREY

# ─── Font factories ──────────────────────────────────────────────────────────
def font_input(size=10, bold=False):
    return Font(name="Arial", size=size, bold=bold, color=INPUT_FONT_COLOR)

def font_formula(size=10, bold=False):
    return Font(name="Arial", size=size, bold=bold, color=FORMULA_FONT_COLOR)

def font_xsheet(size=10, bold=False):
    return Font(name="Arial", size=size, bold=bold, color=XSHEET_FONT_COLOR)

def font_header(size=11, bold=True):
    return Font(name="Arial", size=size, bold=bold, color=HEADER_FG)

def font_subhdr(size=10, bold=True):
    return Font(name="Arial", size=size, bold=bold, color=SUBHDR_FG)

def font_label(size=10, bold=False):
    return Font(name="Arial", size=size, bold=bold, color=DARK_GREY)

def font_check_pass():
    return Font(name="Arial", size=10, bold=True, color="27AE60")

def font_check_fail():
    return Font(name="Arial", size=10, bold=True, color="C0392B")

# ─── Fill factories ──────────────────────────────────────────────────────────
def fill_solid(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

fill_header   = fill_solid(HEADER_BG)
fill_subhdr   = fill_solid(SUBHDR_BG)
fill_alt      = fill_solid(ALT_ROW_BG)
fill_white    = fill_solid(WHITE)
fill_amber    = fill_solid(AMBER)
fill_teal_hdr = fill_solid(TEAL)
fill_blue_hdr = fill_solid(BLUE)

# ─── Border factories ────────────────────────────────────────────────────────
def thin_side(color="CCCCCC"):
    return Side(style="thin", color=color)

def thick_side(color=NAVY):
    return Side(style="medium", color=color)

def border_all_thin(color="CCCCCC"):
    s = thin_side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom_thick(color=NAVY):
    return Border(bottom=thick_side(color))

def border_none():
    return Border()

# ─── Alignment presets ───────────────────────────────────────────────────────
align_center   = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_left     = Alignment(horizontal="left",   vertical="center", wrap_text=False)
align_right    = Alignment(horizontal="right",  vertical="center", wrap_text=False)
align_wrap_l   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
align_header   = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ─── Number format strings ────────────────────────────────────────────────────
FMT_NUMBER     = '#,##0.00'          # 1,234.56
FMT_NUMBER_1   = '#,##0.0'           # 1,234.5
FMT_INTEGER    = '#,##0'             # 1,234
FMT_LAKHS      = '#,##0.00'          # standard lakhs display
FMT_PCT_1      = '0.0%'              # 12.5%
FMT_PCT_2      = '0.00%'             # 12.50%
FMT_DATE       = 'DD-MMM-YYYY'       # 31-Mar-2027
FMT_DATE_SHORT = 'MMM-YY'            # Mar-27
FMT_TEXT       = '@'                 # force text
FMT_ZERO_DASH  = '#,##0.00;(#,##0.00);"-"'  # negatives in parens, zero as -
FMT_RATIO      = '0.00x'             # for multiples


# ─── Row height presets (in points) ──────────────────────────────────────────
ROW_HEIGHT_HEADER  = 20
ROW_HEIGHT_SUBHDR  = 16
ROW_HEIGHT_DATA    = 15
ROW_HEIGHT_SECTION = 18
ROW_HEIGHT_HIDDEN  = 0.1   # "invisible" rows for unused dynamic slots


# ═══════════════════════════════════════════════════════════════════════════════
#  BaseSheetWriter
# ═══════════════════════════════════════════════════════════════════════════════

class BaseSheetWriter:
    """
    Base class for all per-sheet writers.

    Subclasses call self.ws to get the worksheet, then use the
    helper methods below. No raw openpyxl styling calls should
    appear in subclasses — always go through these helpers.
    """

    def __init__(self, wb: Workbook, sheet_name: str):
        if sheet_name in wb.sheetnames:
            self.ws = wb[sheet_name]
        else:
            self.ws = wb.create_sheet(title=sheet_name)
        self.sheet_name = sheet_name
        self._setup_sheet()

    def _setup_sheet(self):
        """Override to set tab color, zoom, freeze panes, etc."""
        self.ws.sheet_view.showGridLines = True
        self.ws.sheet_properties.tabColor = NAVY

    # ── Core write helpers ────────────────────────────────────────────────────

    def w(self, row: int, col: int, value,
          font=None, fill=None, alignment=None,
          border=None, number_format=None, height=None):
        """Write a value to a cell with optional styling."""
        cell = self.ws.cell(row=row, column=col, value=value)
        if font:            cell.font          = font
        if fill:            cell.fill          = fill
        if alignment:       cell.alignment     = alignment
        if border:          cell.border        = border
        if number_format:   cell.number_format = number_format
        if height is not None:
            self.ws.row_dimensions[row].height = height
        return cell

    def write_input(self, row: int, col: int, value,
                    fmt=FMT_LAKHS, bold=False, height=ROW_HEIGHT_DATA,
                    fill=None):
        """Write a blue input cell (user-entered value)."""
        return self.w(row, col, value,
                      font=font_input(bold=bold),
                      fill=fill or fill_white,
                      alignment=align_right,
                      border=border_all_thin(),
                      number_format=fmt,
                      height=height)

    def write_formula(self, row: int, col: int, formula: str,
                      fmt=FMT_LAKHS, bold=False, height=ROW_HEIGHT_DATA,
                      fill=None, xsheet=False):
        """Write a black formula cell."""
        fnt = font_xsheet(bold=bold) if xsheet else font_formula(bold=bold)
        return self.w(row, col, formula,
                      font=fnt,
                      fill=fill or fill_white,
                      alignment=align_right,
                      border=border_all_thin(),
                      number_format=fmt,
                      height=height)

    def write_label(self, row: int, col: int, text: str,
                    bold=False, fill=None, height=ROW_HEIGHT_DATA,
                    alignment=None, indent=0):
        """Write a text label cell."""
        cell = self.w(row, col, "  " * indent + text,
                      font=font_label(bold=bold),
                      fill=fill or fill_white,
                      alignment=alignment or align_left,
                      border=border_all_thin(),
                      height=height)
        return cell

    def write_section_header(self, row: int, col_start: int, col_end: int,
                              text: str):
        """Write a full-width navy section header with merged cells."""
        self.w(row, col_start, text,
               font=font_header(size=11),
               fill=fill_header,
               alignment=align_center,
               height=ROW_HEIGHT_SECTION)
        if col_end > col_start:
            self.ws.merge_cells(
                start_row=row, start_column=col_start,
                end_row=row,   end_column=col_end
            )

    def write_sub_header(self, row: int, col_start: int, col_end: int,
                          text: str, fill=None):
        """Write a blue sub-header row."""
        self.w(row, col_start, text,
               font=font_subhdr(),
               fill=fill or fill_subhdr,
               alignment=align_center,
               height=ROW_HEIGHT_SUBHDR)
        if col_end > col_start:
            self.ws.merge_cells(
                start_row=row, start_column=col_start,
                end_row=row,   end_column=col_end
            )

    def write_column_header(self, row: int, col: int, text: str,
                             fill=None, width=None):
        """Write a single column header cell."""
        self.w(row, col, text,
               font=font_subhdr(),
               fill=fill or fill_subhdr,
               alignment=align_header,
               border=border_all_thin(NAVY),
               height=ROW_HEIGHT_SUBHDR)
        if width:
            self.ws.column_dimensions[get_column_letter(col)].width = width

    def write_total(self, row: int, col: int, formula: str,
                    fmt=FMT_LAKHS, bold=True):
        """Write a subtotal/total formula cell with bold formatting."""
        return self.w(row, col, formula,
                      font=font_formula(bold=bold),
                      fill=fill_solid(LIGHT_TEAL),
                      alignment=align_right,
                      border=border_all_thin(TEAL),
                      number_format=fmt,
                      height=ROW_HEIGHT_DATA)

    def write_check(self, row: int, col: int, formula: str):
        """Write a balance check cell (TRUE = green, value = red)."""
        cell = self.ws.cell(row=row, column=col, value=formula)
        cell.alignment = align_center
        # Conditional formatting is set separately via workbook-level rules
        return cell

    def set_col_width(self, col: int, width: float):
        self.ws.column_dimensions[get_column_letter(col)].width = width

    def set_row_height(self, row: int, height: float):
        self.ws.row_dimensions[row].height = height

    def hide_row(self, row: int):
        self.ws.row_dimensions[row].height = ROW_HEIGHT_HIDDEN
        self.ws.row_dimensions[row].hidden = True

    def merge(self, row: int, col_start: int, col_end: int):
        self.ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row,   end_column=col_end
        )

    def freeze(self, row: int, col: int):
        """Freeze panes at (row, col) — rows above and cols left are frozen."""
        from openpyxl.utils.cell import get_column_letter
        self.ws.freeze_panes = f"{get_column_letter(col)}{row}"

    def blank_row(self, row: int):
        self.set_row_height(row, ROW_HEIGHT_DATA)

    def write_year_headers(self, header_row: int, col_start: int,
                            n_years: int, start_date_formula: str = None):
        """Write Year 1, Year 2, … Year N column headers."""
        for yr in range(1, n_years + 1):
            col = col_start + (yr - 1)
            self.write_column_header(header_row, col, f"Year {yr}")
