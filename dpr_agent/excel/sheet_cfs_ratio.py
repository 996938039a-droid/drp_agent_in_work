"""
sheet_cfs.py + sheet_ratio.py
──────────────────────────────
CFS: Indirect method cash flow statement.
  A) Cash from Operations: PAT + non-cash items + WC changes
  B) Cash from Investing:  Capex (negative in year 1)
  C) Cash from Financing:  Term loan drawdown - repayment - interest paid
  Net Change in Cash → Opening + Net Change = Closing Cash

Ratio: Key financial ratios for banker appraisal.
  DSCR (Debt Service Coverage Ratio)
  ROCE (Return on Capital Employed)
  Net Profit Margin
  Break-Even Point
  IRR (approximated using NPV of free cash flows)
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from core.session_store import SessionStore
from core.layout_engine import (
    LayoutEngine, CFSLayout as CFS, RatioLayout as RT,
    PLLayout as PL, BSLayout as BS, TermLoanLayout as TL,
    WCapLayout as WL, DepreciationLayout as DL,
    col_letter, year_col
)
from excel.styles import (
    BaseSheetWriter, FMT_LAKHS, FMT_PCT_2, FMT_NUMBER,
    fill_header, fill_subhdr, fill_alt, fill_white, fill_solid,
    border_all_thin, align_right, align_center,
    ROW_HEIGHT_SUBHDR, ROW_HEIGHT_DATA,
    NAVY, BLUE, TEAL, WHITE,
)

CFS_DATA_COL = CFS.DATA_COL_START   # C = Year 1
PL_DATA_COL  = 6                    # F = Year 1 in PL


class CFSWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years = store.n_years
        self._n_ac   = len(set(a.category for a in store.capital_means.assets))
        super().__init__(wb, "CFS")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = BLUE
        self.ws.column_dimensions[get_column_letter(2)].width = 38
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[
                get_column_letter(CFS_DATA_COL + yr - 1)].width = 15
        self.ws.freeze_panes = f"{get_column_letter(CFS_DATA_COL)}6"

    def _write(self):
        self._write_title()
        self._write_col_headers()
        self._write_operations()
        self._write_investing()
        self._write_financing()
        self._write_net_cash()
        self._register_layout()

    def _write_title(self):
        last = CFS_DATA_COL + self.n_years - 1
        self.write_section_header(1, 2, last,
            "PROJECTED CASH FLOW STATEMENT  (Indirect Method)")
        self.write_section_header(2, 2, last,
            f"{self.store.project_profile.company_name}  |  "
            "(All amounts in INR Lakhs unless otherwise stated)")
        self.ws.row_dimensions[1].height = 18
        self.ws.row_dimensions[2].height = 14

    def _write_col_headers(self):
        self.write_column_header(3, 2, "Particulars", fill=fill_solid(NAVY))
        for yr in range(1, self.n_years + 1):
            self.write_column_header(3, CFS_DATA_COL + yr - 1, f"Year {yr}",
                                      fill=fill_solid(NAVY))
        self.ws.row_dimensions[3].height = ROW_HEIGHT_SUBHDR

    def _write_operations(self):
        last = CFS_DATA_COL + self.n_years - 1
        self.write_sub_header(4, 2, last,
                               "A.  CASH FROM OPERATING ACTIVITIES",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[4].height = 13

        pl = self.layout._map.get("PL", {})
        pat_r      = pl.get("pat_row", (PL.pat_row(), PL_DATA_COL))[0]
        depr_r     = pl.get("depr_opex_row", PL.OE_BASE_ROW + 5)
        int_r      = pl.get("total_interest_row", (PL.total_interest_row(), PL_DATA_COL))[0]

        rows = [
            (CFS.PBT_ROW,          "  Profit Before Tax",              "pbt"),
            (CFS.ADD_DEPR_ROW,     "  Add: Depreciation",              "depr"),
            (CFS.ADD_INT_ROW,      "  Add: Finance Costs (Interest)",   "int"),
            (CFS.CHG_RECEIVABLES_ROW,"  (Increase)/Decrease in Debtors","dbtrs"),
            (CFS.CHG_INVENTORY_ROW,  "  (Increase)/Decrease in Stock",  "stk"),
            (CFS.CHG_CREDITORS_ROW,  "  Increase/(Decrease) in Creditors","cred"),
            (CFS.TAXES_PAID_ROW,   "  Less: Taxes Paid",               "tax"),
            (CFS.NET_OPERATING_ROW,"NET CASH FROM OPERATIONS",         "total"),
        ]

        pbt_src_r  = pl.get("pbt_row", (PL.pbt_row(), PL_DATA_COL))[0]
        tax_src_r  = pl.get("tax_row", (PL.current_tax_row(), PL_DATA_COL))[0]

        for row_const, label, kind in rows:
            bold = kind == "total"
            fill = fill_solid("D5F5E3") if bold else (
                fill_alt if row_const % 2 == 0 else fill_white)
            self.write_label(row_const, 2, label,
                             bold=bold, fill=fill)

            for yr in range(1, self.n_years + 1):
                col   = CFS_DATA_COL + yr - 1
                c     = get_column_letter(col)
                pl_c  = get_column_letter(PL_DATA_COL + yr - 1)
                bs_c  = get_column_letter(3 + yr - 1)  # BS_DATA_COL=3
                bs_pc = get_column_letter(3 + yr - 2) if yr > 1 else None

                if kind == "pbt":
                    formula = f"=PL!{pl_c}{pbt_src_r}"
                elif kind == "depr":
                    formula = f"=PL!{pl_c}{depr_r}"
                elif kind == "int":
                    formula = f"=PL!{pl_c}{int_r}"
                elif kind == "dbtrs":
                    wc_r = WL.DEBTORS_ROW
                    if yr == 1:
                        formula = f"=-'W Cap'!{c}{wc_r}"
                    else:
                        formula = (f"='W Cap'!{bs_pc}{wc_r}"
                                   f"-'W Cap'!{c}{wc_r}")
                elif kind == "stk":
                    wc_r = WL.STOCK_ROW
                    if yr == 1:
                        formula = f"=-'W Cap'!{c}{wc_r}"
                    else:
                        formula = (f"='W Cap'!{bs_pc}{wc_r}"
                                   f"-'W Cap'!{c}{wc_r}")
                elif kind == "cred":
                    wc_r = WL.TOTAL_CL_ROW
                    if yr == 1:
                        formula = f"='W Cap'!{c}{wc_r}"
                    else:
                        formula = (f"='W Cap'!{c}{wc_r}"
                                   f"-'W Cap'!{bs_pc}{wc_r}")
                elif kind == "tax":
                    formula = f"=-PL!{pl_c}{tax_src_r}"
                elif kind == "total":
                    formula = (f"=SUM({c}{CFS.PBT_ROW}:{c}{CFS.TAXES_PAID_ROW})")

                if bold:
                    self._bold_teal(row_const, col, formula)
                else:
                    self.write_formula(row_const, col, formula,
                                       fmt=FMT_LAKHS, fill=fill,
                                       xsheet=(kind != "total"))
            self.ws.row_dimensions[row_const].height = ROW_HEIGHT_DATA

    def _write_investing(self):
        last = CFS_DATA_COL + self.n_years - 1
        self.write_sub_header(CFS.INVEST_HDR_ROW, 2, last,
                               "B.  CASH FROM INVESTING ACTIVITIES",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[CFS.INVEST_HDR_ROW].height = 13

        capex_r   = CFS.CAPEX_ROW
        net_inv_r = CFS.NET_INVESTING_ROW

        total_cost = sum(a.cost_lakhs for a in self.store.capital_means.assets)
        self.write_label(capex_r, 2,
                         "  Capital Expenditure (Fixed Assets)",
                         fill=fill_alt)
        self.write_label(net_inv_r, 2,
                         "NET CASH FROM INVESTING",
                         bold=True, fill=fill_solid("D5F5E3"))

        for yr in range(1, self.n_years + 1):
            col = CFS_DATA_COL + yr - 1
            # Capex only in construction years (year 1 approximation)
            capex_formula = f"={-total_cost}" if yr == 1 else "=0"
            self.write_formula(capex_r, col, capex_formula,
                               fmt=FMT_LAKHS, fill=fill_alt)
            c = get_column_letter(col)
            self._bold_teal(net_inv_r, col, f"={c}{capex_r}")

        for r in [capex_r, net_inv_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    def _write_financing(self):
        last    = CFS_DATA_COL + self.n_years - 1
        fin_hdr = CFS.FIN_HDR_ROW
        self.write_sub_header(fin_hdr, 2, last,
                               "C.  CASH FROM FINANCING ACTIVITIES",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[fin_hdr].height = 13

        tl_r    = CFS.TL_ROW
        int_r   = CFS.INT_PAID_ROW
        net_r   = CFS.NET_FINANCING_ROW

        ann_princ_r = TL.SUMMARY_YEAR_ROW + 3
        ann_int_r   = TL.SUMMARY_YEAR_ROW + 5
        total_loan  = sum(
            s.amount_lakhs for s in self.store.capital_means.finance_sources
            if s.source_type.value == "Term Loan"
        )

        self.write_label(tl_r,  2, "  Term Loan Drawdown / (Repayment)",
                         fill=fill_alt)
        self.write_label(int_r, 2, "  Finance Costs Paid",
                         fill=fill_white)
        self.write_label(net_r, 2, "NET CASH FROM FINANCING",
                         bold=True, fill=fill_solid("D5F5E3"))

        for yr in range(1, self.n_years + 1):
            col  = CFS_DATA_COL + yr - 1
            c    = get_column_letter(col)
            tl_c = get_column_letter(10 + yr - 1)

            # Year 1: full drawdown - repayment; subsequent: just repayment
            if yr == 1:
                tl_formula = f"={total_loan}-'Term Loan'!{tl_c}{ann_princ_r}"
            else:
                tl_formula = f"=-'Term Loan'!{tl_c}{ann_princ_r}"

            self.write_formula(tl_r, col, tl_formula,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            pl = self.layout._map.get("PL", {})
            int_src = pl.get("total_interest_row",
                             (PL.total_interest_row(), PL_DATA_COL))[0]
            pl_c = get_column_letter(PL_DATA_COL + yr - 1)
            self.write_formula(int_r, col,
                f"=-PL!{pl_c}{int_src}",
                fmt=FMT_LAKHS, fill=fill_white, xsheet=True)

            self._bold_teal(net_r, col, f"={c}{tl_r}+{c}{int_r}")

        for r in [tl_r, int_r, net_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    def _write_net_cash(self):
        net_r    = CFS.NET_CHANGE_ROW
        open_r   = CFS.OPENING_CASH_ROW
        close_r  = CFS.CLOSING_CASH_ROW

        self.write_label(net_r,   2, "NET CHANGE IN CASH",
                         bold=True, fill=fill_solid(TEAL))
        self.write_label(open_r,  2, "  Opening Cash Balance",
                         fill=fill_alt)
        self.write_label(close_r, 2, "CLOSING CASH BALANCE",
                         bold=True, fill=fill_solid(TEAL))

        for yr in range(1, self.n_years + 1):
            col = CFS_DATA_COL + yr - 1
            c   = get_column_letter(col)

            # Net change = ops + investing + financing
            self._bold_teal(net_r, col,
                f"={c}{CFS.NET_OPERATING_ROW}"
                f"+{c}{CFS.NET_INVESTING_ROW}"
                f"+{c}{CFS.NET_FINANCING_ROW}")

            # Opening cash (year 1 = 0, subsequent = prior closing)
            if yr == 1:
                self.write_formula(open_r, col, "=0",
                                   fmt=FMT_LAKHS, fill=fill_alt)
            else:
                prev_c = get_column_letter(col - 1)
                self.write_formula(open_r, col, f"={prev_c}{close_r}",
                                   fmt=FMT_LAKHS, fill=fill_alt)

            self._bold_teal(close_r, col,
                f"={c}{open_r}+{c}{net_r}")

        for r in [net_r, open_r, close_r]:
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    def _register_layout(self):
        self.layout._map["CFS"] = {
            "closing_cash_row": (CFS.CLOSING_CASH_ROW, CFS_DATA_COL),
            "net_ops_row":      (CFS.NET_OPERATING_ROW, CFS_DATA_COL),
            "net_inv_row":      (CFS.NET_INVESTING_ROW, CFS_DATA_COL),
            "net_fin_row":      (CFS.NET_FINANCING_ROW, CFS_DATA_COL),
            "data_col_start":   CFS_DATA_COL,
        }

    def _bold_teal(self, row: int, col: int, formula: str):
        cell = self.ws.cell(row=row, column=col, value=formula)
        cell.font          = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill          = fill_solid("1ABC9C")
        cell.number_format = FMT_LAKHS
        cell.alignment     = align_right
        cell.border        = border_all_thin(TEAL)


# ═══════════════════════════════════════════════════════════════════════════════

class RatioWriter(BaseSheetWriter):

    def __init__(self, wb: Workbook, store: SessionStore, layout: LayoutEngine):
        self.store   = store
        self.layout  = layout
        self.n_years = store.n_years
        super().__init__(wb, "Ratio")
        self._write()

    def _setup_sheet(self):
        super()._setup_sheet()
        self.ws.sheet_view.showGridLines = False
        self.ws.sheet_properties.tabColor = TEAL
        self.ws.column_dimensions[get_column_letter(2)].width = 36
        for yr in range(1, self.n_years + 1):
            self.ws.column_dimensions[
                get_column_letter(RT.DATA_COL_START + yr - 1)].width = 14
        self.ws.freeze_panes = f"{get_column_letter(RT.DATA_COL_START)}5"

    def _write(self):
        self._write_title()
        self._write_col_headers()
        self._write_dscr()
        self._write_profitability()
        self._write_roce()
        self._write_bep()

    def _write_title(self):
        last = RT.DATA_COL_START + self.n_years - 1
        self.write_section_header(1, 2, last,
            "KEY FINANCIAL RATIOS & INDICATORS")
        self.write_section_header(2, 2, last,
            f"{self.store.project_profile.company_name}  |  "
            "For Banker Appraisal")
        self.ws.row_dimensions[1].height = 18
        self.ws.row_dimensions[2].height = 14

    def _write_col_headers(self):
        self.write_column_header(3, 2, "Ratio / Indicator",
                                  fill=fill_solid(NAVY))
        for yr in range(1, self.n_years + 1):
            self.write_column_header(3, RT.DATA_COL_START + yr - 1,
                                      f"Year {yr}", fill=fill_solid(NAVY))
        self.ws.row_dimensions[3].height = ROW_HEIGHT_SUBHDR

    def _write_dscr(self):
        last = RT.DATA_COL_START + self.n_years - 1
        self.write_sub_header(RT.DSCR_HDR_ROW, 2, last,
                               "DEBT SERVICE COVERAGE RATIO (DSCR)",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[RT.DSCR_HDR_ROW].height = 13

        pl  = self.layout._map.get("PL", {})
        pat_r   = pl.get("pat_row",       (PL.pat_row(),            6))[0]
        int_r   = pl.get("total_interest_row", (PL.total_interest_row(), 6))[0]
        depr_r  = pl.get("depr_opex_row", PL.OE_BASE_ROW + 5)

        ann_pr  = TL.SUMMARY_YEAR_ROW + 3   # principal repaid
        ann_ir  = TL.SUMMARY_YEAR_ROW + 5   # interest

        labels_rows = [
            (RT.NUMERATOR_ROW,   "  Numerator  (PAT + Depr + Interest)"),
            (RT.DENOMINATOR_ROW, "  Denominator (Principal + Interest)"),
            (RT.DSCR_ROW,        "DSCR"),
            (RT.AVG_DSCR_ROW,    "  Average DSCR (all years)"),
        ]
        for r, lbl in labels_rows:
            bold = "DSCR" in lbl and "Average" not in lbl
            fill = fill_solid("D5F5E3") if bold else fill_alt
            self.write_label(r, 2, lbl, bold=bold, fill=fill)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        dscr_cols = []
        for yr in range(1, self.n_years + 1):
            col   = RT.DATA_COL_START + yr - 1
            c     = get_column_letter(col)
            pl_c  = get_column_letter(6 + yr - 1)
            tl_c  = get_column_letter(10 + yr - 1)

            # Numerator
            num = f"=PL!{pl_c}{pat_r}+PL!{pl_c}{depr_r}+PL!{pl_c}{int_r}"
            self.write_formula(RT.NUMERATOR_ROW, col, num,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            # Denominator
            den = f"='Term Loan'!{tl_c}{ann_pr}+'Term Loan'!{tl_c}{ann_ir}"
            self.write_formula(RT.DENOMINATOR_ROW, col, den,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            # DSCR
            dscr_r_col = RT.DSCR_ROW
            self._bold_teal(dscr_r_col, col,
                f"=IFERROR({c}{RT.NUMERATOR_ROW}/{c}{RT.DENOMINATOR_ROW},0)")
            dscr_cols.append(get_column_letter(col) + str(dscr_r_col))

        # Average DSCR
        avg_range = ",".join(dscr_cols)
        for yr in range(1, self.n_years + 1):
            col = RT.DATA_COL_START + yr - 1
            if yr == 1:
                self.write_formula(RT.AVG_DSCR_ROW, col,
                    f"=AVERAGE({avg_range})",
                    fmt="0.00x", fill=fill_alt)
            else:
                self.ws.cell(RT.AVG_DSCR_ROW, col).value = ""

    def _write_profitability(self):
        last = RT.DATA_COL_START + self.n_years - 1
        self.write_sub_header(RT.PROF_HDR_ROW, 2, last,
                               "PROFITABILITY RATIOS",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[RT.PROF_HDR_ROW].height = 13

        pl  = self.layout._map.get("PL", {})
        rev_r  = pl.get("total_rev_row", (PL.TOTAL_REV_ROW, 6))[0]
        pat_r  = pl.get("pat_row",       (PL.pat_row(),     6))[0]
        ebit_r = pl.get("ebit_row",      (PL.ebit_row(),    6))[0]

        metrics = [
            (RT.PROF_NPM_ROW,        "Net Profit Margin",   "npm"),
            (RT.PROF_OPR_MARGIN_ROW, "Operating Margin",    "opm"),
            (RT.PROF_DE_ROW,         "Debt / Equity",       "de"),
        ]

        for r, lbl, kind in metrics:
            fill = fill_alt if r % 2 == 0 else fill_white
            self.write_label(r, 2, f"  {lbl}", fill=fill)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

            for yr in range(1, self.n_years + 1):
                col  = RT.DATA_COL_START + yr - 1
                c    = get_column_letter(col)
                pl_c = get_column_letter(6 + yr - 1)
                bs_c = get_column_letter(3 + yr - 1)
                tl_c = get_column_letter(10 + yr - 1)

                if kind == "npm":
                    formula = (f"=IFERROR(PL!{pl_c}{pat_r}"
                               f"/PL!{pl_c}{rev_r},0)")
                    self.write_formula(r, col, formula,
                                       fmt=FMT_PCT_2, fill=fill, xsheet=True)
                elif kind == "opm":
                    formula = (f"=IFERROR(PL!{pl_c}{ebit_r}"
                               f"/PL!{pl_c}{rev_r},0)")
                    self.write_formula(r, col, formula,
                                       fmt=FMT_PCT_2, fill=fill, xsheet=True)
                elif kind == "de":
                    ann_out_r = TL.SUMMARY_YEAR_ROW + 4
                    equity_r  = self.layout._map.get("BS", {}).get(
                        "total_liab_row", (BS.TOTAL_LIAB_ROW, 3))[0] - 2
                    formula = (f"=IFERROR('Term Loan'!{tl_c}{ann_out_r}"
                               f"/BS!{bs_c}{equity_r},0)")
                    self.write_formula(r, col, formula,
                                       fmt="0.00x", fill=fill, xsheet=True)

    def _write_roce(self):
        last = RT.DATA_COL_START + self.n_years - 1
        self.write_sub_header(RT.ROCE_HDR_ROW, 2, last,
                               "RETURN ON CAPITAL EMPLOYED (ROCE)",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[RT.ROCE_HDR_ROW].height = 13

        pl   = self.layout._map.get("PL", {})
        ebit_r = pl.get("ebit_row", (PL.ebit_row(), 6))[0]
        tot_assets_r = self.layout._map.get("BS", {}).get(
            "total_assets_row", (BS.TOTAL_ASSETS_ROW, 3))[0]

        labels_rows = [
            (RT.ROCE_ROW,     "ROCE  (EBIT / Total Assets)"),
            (RT.AVG_ROCE_ROW, "  Average ROCE"),
        ]
        for r, lbl in labels_rows:
            bold = "ROCE" in lbl and "Average" not in lbl
            self.write_label(r, 2, lbl, bold=bold,
                             fill=fill_solid("D5F5E3") if bold else fill_alt)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        roce_cells = []
        for yr in range(1, self.n_years + 1):
            col  = RT.DATA_COL_START + yr - 1
            c    = get_column_letter(col)
            pl_c = get_column_letter(6 + yr - 1)
            bs_c = get_column_letter(3 + yr - 1)

            formula = (f"=IFERROR(PL!{pl_c}{ebit_r}"
                       f"/BS!{bs_c}{tot_assets_r},0)")
            self._bold_teal(RT.ROCE_ROW, col, formula)
            self.ws.cell(RT.ROCE_ROW, col).number_format = FMT_PCT_2
            roce_cells.append(f"{c}{RT.ROCE_ROW}")

        for yr in range(1, self.n_years + 1):
            col = RT.DATA_COL_START + yr - 1
            if yr == 1:
                avg_range = ",".join(roce_cells)
                self.write_formula(RT.AVG_ROCE_ROW, col,
                    f"=AVERAGE({avg_range})",
                    fmt=FMT_PCT_2, fill=fill_alt)

    def _write_bep(self):
        last = RT.DATA_COL_START + self.n_years - 1
        self.write_sub_header(RT.BEP_HDR_ROW, 2, last,
                               "BREAK-EVEN POINT ANALYSIS",
                               fill=fill_solid(BLUE))
        self.ws.row_dimensions[RT.BEP_HDR_ROW].height = 13

        pl     = self.layout._map.get("PL", {})
        rev_r  = pl.get("total_rev_row", (PL.TOTAL_REV_ROW, 6))[0]
        cogs_r = pl.get("total_cogs_row", (PL.TOTAL_COGS_ROW, 6))[0]
        depr_r = pl.get("depr_opex_row", PL.OE_BASE_ROW + 5)
        int_r  = pl.get("total_interest_row", (PL.total_interest_row(), 6))[0]
        sga_r  = pl.get("total_opex_row", (PL.total_opex_row(), 6))[0]

        bep_rows = [
            (RT.BEP_SGA_ROW,         "  Fixed Costs (Opex + Interest)"),
            (RT.BEP_CONTRIBUTION_ROW, "  Contribution (Rev - COGS)"),
            (RT.BEP_PCT_ROW,          "BREAK-EVEN POINT (% of Revenue)"),
        ]
        for r, lbl in bep_rows:
            bold = "BREAK-EVEN" in lbl
            fill = fill_solid("D5F5E3") if bold else fill_alt
            self.write_label(r, 2, lbl, bold=bold, fill=fill)
            self.ws.row_dimensions[r].height = ROW_HEIGHT_DATA

        for yr in range(1, self.n_years + 1):
            col  = RT.DATA_COL_START + yr - 1
            c    = get_column_letter(col)
            pl_c = get_column_letter(6 + yr - 1)

            # Fixed costs = total opex - COGS + interest
            fc = (f"=PL!{pl_c}{sga_r}"
                  f"-PL!{pl_c}{cogs_r}"
                  f"+PL!{pl_c}{int_r}"
                  f"+PL!{pl_c}{depr_r}")
            self.write_formula(RT.BEP_SGA_ROW, col, fc,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            # Contribution = Revenue - COGS
            contrib = f"=PL!{pl_c}{rev_r}-PL!{pl_c}{cogs_r}"
            self.write_formula(RT.BEP_CONTRIBUTION_ROW, col, contrib,
                               fmt=FMT_LAKHS, fill=fill_alt, xsheet=True)

            # BEP % = Fixed Cost / Contribution × 100
            bep = (f"=IFERROR({c}{RT.BEP_SGA_ROW}"
                   f"/{c}{RT.BEP_CONTRIBUTION_ROW},0)")
            self._bold_teal(RT.BEP_PCT_ROW, col, bep)
            self.ws.cell(RT.BEP_PCT_ROW, col).number_format = FMT_PCT_2

    def _bold_teal(self, row: int, col: int, formula: str):
        cell = self.ws.cell(row=row, column=col, value=formula)
        cell.font          = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill          = fill_solid("1ABC9C")
        cell.number_format = FMT_LAKHS
        cell.alignment     = align_right
        cell.border        = border_all_thin(TEAL)
