"""
test_phase4.py
──────────────
Tests for Phase 4: Expenses + Term Loan + W Cap + Tax sheets.
"""

import sys, os, tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from openpyxl import load_workbook
from core.session_store import *
from core.layout_engine import (
    LayoutEngine, ExpensesLayout as EL, TermLoanLayout as TL,
    WCapLayout as WL, TaxLayout as TxL, year_col
)
from excel.workbook_builder import WorkbookBuilder
from openpyxl.utils import get_column_letter

GREEN="\033[92m"; RED="\033[91m"; CYAN="\033[96m"; RESET="\033[0m"; BOLD="\033[1m"
passed=0; failed=0; errors=[]

def ok(n):
    global passed; passed+=1; print(f"  {GREEN}✓{RESET}  {n}")
def fail(n,r):
    global failed; failed+=1; errors.append((n,r))
    print(f"  {RED}✗{RESET}  {n}"); print(f"      {RED}{r}{RESET}")
def section(t):
    print(f"\n{BOLD}{CYAN}{'─'*60}{RESET}\n{BOLD}{CYAN}  {t}{RESET}\n{BOLD}{CYAN}{'─'*60}{RESET}")
def assert_eq(n,a,e):
    if a==e: ok(n)
    else: fail(n,f"expected {repr(e)}, got {repr(a)}")
def assert_true(n,c,m=""):
    if c: ok(n)
    else: fail(n,m or "False")
def is_formula(v): return isinstance(v,str) and v.startswith("=")
def xref_in(v,sheet): return is_formula(v) and sheet in str(v)

def make_store():
    s=SessionStore()
    s.project_profile.company_name='Test DPR Co'
    s.project_profile.entity_type=EntityType.COMPANY
    s.project_profile.operation_start_date='2027-04'
    s.project_profile.projection_years=7
    s.project_profile.status=SectionStatus.COMPLETE
    s.capital_means.assets=[
        Asset('Civil Works',AssetCategory.CIVIL_WORKS,202.625),
        Asset('Plant & Machinery',AssetCategory.PLANT_MACHINERY,110.944),
        Asset('Furniture',AssetCategory.FURNITURE,5.0),
    ]
    s.capital_means.finance_sources=[
        FinanceSource(FinanceSourceType.TERM_LOAN,200.0,0.09,84,18,'SBI TL'),
        FinanceSource(FinanceSourceType.OD_LIMIT,70.0,0.09),
        FinanceSource(FinanceSourceType.PROMOTER_EQUITY,117.569),
    ]
    s.capital_means.status=SectionStatus.COMPLETE
    s.revenue_model.products=[
        Product('Mustard Oil','litres',8,0.36,0.5,175.0,0.04),
        Product('Groundnut Oil','litres',8,0.36,0.5,195.0,0.04),
    ]
    s.revenue_model.year1_utilization=0.50
    s.revenue_model.annual_utilization_increment=0.05
    s.revenue_model.max_utilization=0.85
    s.revenue_model.working_days_per_month=28
    s.revenue_model.status=SectionStatus.COMPLETE
    s.cost_structure.raw_materials=[
        RawMaterial('Groundnut Seed','kg',55.0,2.78,0.05),
        RawMaterial('Mustard Seed','kg',44.0,2.78,0.05),
    ]
    s.cost_structure.rm_pct_of_fa=0.02; s.cost_structure.power_pct_revenue=0.06
    s.cost_structure.transport_base_lakhs=7.0; s.cost_structure.misc_base_lakhs=15.0
    s.cost_structure.sga_base_lakhs=6.0; s.cost_structure.status=SectionStatus.COMPLETE
    s.manpower.categories=[
        EmployeeCategory('Manager',1,0.40,True,0.05),
        EmployeeCategory('Operator',5,0.15,True,0.05),
        EmployeeCategory('Admin',3,0.20,True,0.05),
        EmployeeCategory('Guard',1,0.12,True,0.05),
    ]; s.manpower.status=SectionStatus.COMPLETE
    s.finance_wc.debtor_days=10; s.finance_wc.creditor_days_rm=10
    s.finance_wc.creditor_days_admin=30; s.finance_wc.stock_days_rm=7
    s.finance_wc.implementation_months=3; s.finance_wc.status=SectionStatus.COMPLETE
    s.tax_config.entity_type=EntityType.COMPANY; s.tax_config.status=SectionStatus.COMPLETE
    return s

STORE=make_store()
TMP=tempfile.mktemp(suffix="_p4.xlsx")
builder=WorkbookBuilder(STORE); builder.build(TMP)
layout=builder.layout
WB=load_workbook(TMP,data_only=False)
print(f"\n  Generated: {TMP}")
print(f"  Sheets: {WB.sheetnames}\n")

# ─── Section 1: All Phase 4 sheets present ───────────────────────────────────
section("1. All Phase 4 sheets present")
for s in ["Expenses","Term Loan","W Cap","Tax"]:
    assert_true(f"'{s}' sheet exists", s in WB.sheetnames)

# ─── Section 2: Expenses — raw material section ──────────────────────────────
section("2. Expenses — raw material cost formulas")
ws_exp=WB["Expenses"]
n_mats=STORE.n_materials

for i in range(n_mats):
    qty_r  = EL.rm_qty_row(i)
    prc_r  = EL.rm_price_row(i)
    cost_r = EL.rm_cost_row(i)

    qty_y1 = ws_exp.cell(qty_r, year_col(1))
    assert_true(f"RM{i} qty yr1 is formula", is_formula(qty_y1.value),
                f"got {repr(qty_y1.value)}")
    assert_true(f"RM{i} qty references Revenue",
                xref_in(qty_y1.value,"Revenue"),
                f"formula: {qty_y1.value}")

    prc_y1 = ws_exp.cell(prc_r, year_col(1))
    assert_true(f"RM{i} price yr1 references Assumption",
                xref_in(prc_y1.value,"Assumption"),
                f"formula: {prc_y1.value}")

    prc_y2 = ws_exp.cell(prc_r, year_col(2))
    assert_true(f"RM{i} price yr2 escalates",
                is_formula(prc_y2.value) and "Assumption" in str(prc_y2.value),
                f"formula: {prc_y2.value}")

    cost_y1 = ws_exp.cell(cost_r, year_col(1))
    assert_true(f"RM{i} cost is formula",
                is_formula(cost_y1.value),
                f"got {repr(cost_y1.value)}")
    assert_true(f"RM{i} cost divides by 100000",
                "100000" in str(cost_y1.value),
                f"formula: {cost_y1.value}")

# ─── Section 3: Expenses — total COGS and overhead ───────────────────────────
section("3. Expenses — total COGS and overhead formulas")

total_cogs_row = EL.total_cogs_row(n_mats)
tc_y1 = ws_exp.cell(total_cogs_row, year_col(1))
assert_true("Total COGS row is formula", is_formula(tc_y1.value))

# Reference rows
rev_ref_row = EL.revenue_ref_row(n_mats)
nb_ref_row  = EL.net_block_ref_row(n_mats)
rev_y1 = ws_exp.cell(rev_ref_row, year_col(1))
nb_y1  = ws_exp.cell(nb_ref_row,  year_col(1))
assert_true("Revenue ref row references Revenue sheet",
            xref_in(rev_y1.value,"Revenue"), f"formula: {rev_y1.value}")
assert_true("Net block ref row references Depreciation sheet",
            xref_in(nb_y1.value,"Depreciation"), f"formula: {nb_y1.value}")

# Overhead: R&M rate references Assumption
rm_rate_row = EL.rm_rate_row(n_mats)
rm_rate_y1  = ws_exp.cell(rm_rate_row, year_col(1))
assert_true("R&M rate row references Assumption",
            xref_in(rm_rate_y1.value,"Assumption"), f"formula: {rm_rate_y1.value}")

# R&M amount = rate × net block
rm_amt_row = EL.rm_amount_row(n_mats)
rm_amt_y1  = ws_exp.cell(rm_amt_row, year_col(1))
assert_true("R&M amount is formula", is_formula(rm_amt_y1.value))
assert_true("R&M amount references net block ref row",
            str(nb_ref_row) in str(rm_amt_y1.value),
            f"formula: {rm_amt_y1.value}")

# ─── Section 4: Term Loan — monthly schedule ─────────────────────────────────
section("4. Term Loan — monthly schedule structure")
ws_tl=WB["Term Loan"]
loan=STORE.capital_means.term_loans[0]  # 84 months, 18 moratorium

# Moratorium rows should have principal = 0
mora_row = TL.MONTHLY_BASE
mora_cell = ws_tl.cell(mora_row, 4)  # col D = principal
assert_true("Moratorium row 1 principal = 0 formula",
            is_formula(mora_cell.value) and "0" in str(mora_cell.value),
            f"got {repr(mora_cell.value)}")

# First repayment row (month 19)
rep_row = TL.MONTHLY_BASE + 18
rep_princ = ws_tl.cell(rep_row, 4)
assert_true("First repayment month has EMI formula",
            is_formula(rep_princ.value),
            f"got {repr(rep_princ.value)}")

# Interest formula uses rate reference
int_row1 = TL.MONTHLY_BASE
int_cell = ws_tl.cell(int_row1, 5)  # col E = interest
assert_true("Interest col is formula", is_formula(int_cell.value))
assert_true("Interest divides by 12", "/12" in str(int_cell.value))

# Closing = Opening - Principal
closing_r1 = ws_tl.cell(TL.MONTHLY_BASE, 7)
assert_true("Closing is formula", is_formula(closing_r1.value))

# ─── Section 5: Term Loan — annual summary ───────────────────────────────────
section("5. Term Loan — annual summary SUMIF formulas")

base = TL.SUMMARY_YEAR_ROW
princ_r  = base + 3
int_r    = base + 5
outst_r  = base + 4

for yr in range(1, STORE.n_years + 1):
    col = 9 + yr  # AC_START + yr

    princ_cell = ws_tl.cell(princ_r, col)
    assert_true(f"Annual principal yr{yr} uses SUMIF",
                is_formula(princ_cell.value) and "SUMIF" in str(princ_cell.value).upper(),
                f"got {repr(princ_cell.value)}")

    int_cell = ws_tl.cell(int_r, col)
    assert_true(f"Annual interest yr{yr} uses SUMIF",
                is_formula(int_cell.value) and "SUMIF" in str(int_cell.value).upper(),
                f"got {repr(int_cell.value)}")

    outst_cell = ws_tl.cell(outst_r, col)
    assert_true(f"Outstanding yr{yr} is formula",
                is_formula(outst_cell.value), f"got {repr(outst_cell.value)}")

# Annual summary registered in layout
assert_true("annual_interest_row in layout",
            "annual_interest_row" in layout._map.get("Term Loan",{}))
assert_true("annual_outstanding_row in layout",
            "annual_outstanding_row" in layout._map.get("Term Loan",{}))

# ─── Section 6: Term Loan — OD interest ──────────────────────────────────────
section("6. Term Loan — OD interest table")
od_base = TL.od_base_row()

for yr in range(1, STORE.n_years + 1):
    r = od_base + yr
    int_cell = ws_tl.cell(r, 5)  # col E
    assert_true(f"OD interest yr{yr} references Assumption",
                is_formula(int_cell.value) and "Assumption" in str(int_cell.value),
                f"got {repr(int_cell.value)}")

# ─── Section 7: W Cap — current liabilities ──────────────────────────────────
section("7. W Cap — current liabilities")
ws_wc=WB["W Cap"]

cr_rm_cell = ws_wc.cell(WL.CREDITORS_ROW, year_col(1))
assert_true("Creditors RM references Expenses COGS",
            xref_in(cr_rm_cell.value,"Expenses"),
            f"formula: {cr_rm_cell.value}")

cr_adm_cell = ws_wc.cell(WL.ADMIN_CRED_ROW, year_col(1))
assert_true("Creditors admin references Expenses",
            xref_in(cr_adm_cell.value,"Expenses"),
            f"formula: {cr_adm_cell.value}")

total_cl = ws_wc.cell(WL.TOTAL_CL_ROW, year_col(1))
assert_true("Total CL is formula", is_formula(total_cl.value))

# Days reference Assumption
days_cell = ws_wc.cell(WL.CREDITOR_DAYS_ROW, 3)  # COL_BASIS=3
assert_true("Creditor days references Assumption",
            xref_in(days_cell.value,"Assumption"), f"formula: {days_cell.value}")

# ─── Section 8: W Cap — current assets ───────────────────────────────────────
section("8. W Cap — current assets")

stock_cell = ws_wc.cell(WL.STOCK_ROW, year_col(1))
assert_true("Stock references Expenses COGS",
            xref_in(stock_cell.value,"Expenses"),
            f"formula: {stock_cell.value}")

debtor_cell = ws_wc.cell(WL.DEBTORS_ROW, year_col(1))
assert_true("Debtors references Revenue",
            xref_in(debtor_cell.value,"Revenue"),
            f"formula: {debtor_cell.value}")

total_ca = ws_wc.cell(WL.TOTAL_CA_ROW, year_col(1))
assert_true("Total CA is formula", is_formula(total_ca.value))

# ─── Section 9: W Cap — WC requirement ───────────────────────────────────────
section("9. W Cap — net working capital requirement")

wc_req = ws_wc.cell(WL.WC_REQ_ROW, year_col(1))
assert_true("WC requirement = CA - CL formula",
            is_formula(wc_req.value), f"got {repr(wc_req.value)}")
assert_true("WC requirement refs total CA",
            str(WL.TOTAL_CA_ROW) in str(wc_req.value),
            f"formula: {wc_req.value}")
assert_true("WC requirement refs total CL",
            str(WL.TOTAL_CL_ROW) in str(wc_req.value),
            f"formula: {wc_req.value}")

assert_true("wc_req_row in layout",
            "wc_req_row" in layout._map.get("W Cap",{}))
assert_true("debtors_row in layout",
            "debtors_row" in layout._map.get("W Cap",{}))

# ─── Section 10: Tax — computation formulas ──────────────────────────────────
section("10. Tax — tax computation")
ws_tax=WB["Tax"]

# Taxable income references PL
tax_inc_cell = ws_tax.cell(TxL.TAXABLE_INC_ROW, 3)
assert_true("Taxable income references PL",
            xref_in(tax_inc_cell.value,"PL"), f"formula: {tax_inc_cell.value}")
assert_true("Taxable income converts Lakhs to INR (×100000)",
            "100000" in str(tax_inc_cell.value), f"formula: {tax_inc_cell.value}")

# Company tax basic = taxable × rate
co_basic = ws_tax.cell(TxL.CO_BASIC_ROW, 3)
assert_true("Company basic tax is formula", is_formula(co_basic.value))
assert_true("Company basic tax references taxable income row",
            str(TxL.TAXABLE_INC_ROW) in str(co_basic.value),
            f"formula: {co_basic.value}")

# Company total = basic + HEC
co_total = ws_tax.cell(TxL.CO_TOTAL_ROW, 3)
assert_true("Company total tax is formula", is_formula(co_total.value))

# Active tax row uses IF to select entity type
active_r = layout._map.get("Tax",{}).get("active_tax_row",(0,0))[0]
assert_true("active_tax_row registered in layout", active_r > 0)
if active_r:
    active_cell = ws_tax.cell(active_r, 3)
    assert_true("Active tax uses IF formula",
                is_formula(active_cell.value) and "IF(" in str(active_cell.value).upper(),
                f"formula: {active_cell.value}")
    assert_true("Active tax references Assumption entity type",
                "Assumption" in str(active_cell.value),
                f"formula: {active_cell.value}")
    assert_true("Active tax converts to Lakhs (/100000)",
                "100000" in str(active_cell.value),
                f"formula: {active_cell.value}")

# Partnership tax also computed
part_basic = ws_tax.cell(TxL.PART_BASIC_ROW, 3)
assert_true("Partnership basic tax is formula", is_formula(part_basic.value))

# Individual tax computed
ind_basic = ws_tax.cell(TxL.IND_BASIC_ROW, 3)
assert_true("Individual basic tax is formula", is_formula(ind_basic.value))

# ─── Section 11: File integrity ───────────────────────────────────────────────
section("11. File integrity — all 10 sheets, formula counts")

assert_true("All Phase 4 sheets present", all(s in WB.sheetnames for s in ["Expenses","Term Loan","W Cap","Tax"]))
file_size = os.path.getsize(TMP)
assert_true("File size > 20KB", file_size > 20000, f"got {file_size}")

formula_counts = {}
for sn in WB.sheetnames:
    ws=WB[sn]
    formula_counts[sn] = sum(1 for r in ws.iter_rows()
                             for c in r if is_formula(c.value))
    assert_true(f"Sheet '{sn}' has content",
                ws.max_row > 3, f"max_row={ws.max_row}")

# Key sheets have substantial formula counts
assert_true("Term Loan has many formulas (monthly schedule)",
            formula_counts.get("Term Loan",0) > 400,
            f"got {formula_counts.get('Term Loan',0)}")
assert_true("Expenses has formulas",
            formula_counts.get("Expenses",0) > 50,
            f"got {formula_counts.get('Expenses',0)}")
assert_true("Tax has formulas",
            formula_counts.get("Tax",0) > 30,
            f"got {formula_counts.get('Tax',0)}")

# ─── Summary ──────────────────────────────────────────────────────────────────
print(f"\n  {CYAN}Test file: {TMP}{RESET}")
total=passed+failed
print(f"\n{'═'*60}")
print(f"{BOLD}  PHASE 4 TEST RESULTS{RESET}")
print(f"{'═'*60}")
print(f"  {GREEN}Passed: {passed}/{total}{RESET}")
if failed:
    print(f"  {RED}Failed: {failed}/{total}{RESET}")
    for n,r in errors:
        print(f"    {RED}✗ {n}{RESET}"); print(f"      {r}")
else:
    print(f"\n  {GREEN}{BOLD}ALL TESTS PASSED ✓{RESET}")
print(f"{'═'*60}\n")
sys.exit(0 if failed==0 else 1)
