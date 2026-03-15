"""test_phase5.py — PL, BS, CFS, Ratio sheet tests."""

import sys, os, tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from openpyxl import load_workbook
from core.session_store import *
from core.layout_engine import PLLayout as PL, BSLayout as BS, CFSLayout as CFS, RatioLayout as RT, year_col
from excel.workbook_builder import WorkbookBuilder
from openpyxl.utils import get_column_letter

GREEN="\033[92m"; RED="\033[91m"; CYAN="\033[96m"; RESET="\033[0m"; BOLD="\033[1m"
passed=0; failed=0; errors=[]

def ok(n): global passed; passed+=1; print(f"  {GREEN}✓{RESET}  {n}")
def fail(n,r): global failed; failed+=1; errors.append((n,r)); print(f"  {RED}✗{RESET}  {n}"); print(f"      {RED}{r}{RESET}")
def section(t): print(f"\n{BOLD}{CYAN}{'─'*60}{RESET}\n{BOLD}{CYAN}  {t}{RESET}\n{BOLD}{CYAN}{'─'*60}{RESET}")
def assert_true(n,c,m=""): ok(n) if c else fail(n,m or "False")
def is_formula(v): return isinstance(v,str) and v.startswith("=")
def xref(v,sheet): return is_formula(v) and sheet in str(v)

def make_store():
    s=SessionStore()
    s.project_profile.company_name='Test DPR'; s.project_profile.entity_type=EntityType.COMPANY
    s.project_profile.operation_start_date='2027-04'; s.project_profile.projection_years=7
    s.project_profile.status=SectionStatus.COMPLETE
    s.capital_means.assets=[Asset('Civil',AssetCategory.CIVIL_WORKS,200),Asset('PM',AssetCategory.PLANT_MACHINERY,110),Asset('Furn',AssetCategory.FURNITURE,5)]
    s.capital_means.finance_sources=[FinanceSource(FinanceSourceType.TERM_LOAN,200,0.09,84,18),FinanceSource(FinanceSourceType.OD_LIMIT,70,0.09),FinanceSource(FinanceSourceType.PROMOTER_EQUITY,117)]
    s.capital_means.status=SectionStatus.COMPLETE
    s.revenue_model.products=[Product('P1','L',8,0.36,0.5,175,0.04),Product('P2','L',8,0.36,0.5,195,0.04)]
    s.revenue_model.year1_utilization=0.50; s.revenue_model.annual_utilization_increment=0.05
    s.revenue_model.max_utilization=0.85; s.revenue_model.working_days_per_month=28; s.revenue_model.status=SectionStatus.COMPLETE
    s.cost_structure.raw_materials=[RawMaterial('M1','kg',55,2.78,0.05),RawMaterial('M2','kg',44,2.78,0.05)]
    s.cost_structure.rm_pct_of_fa=0.02; s.cost_structure.power_pct_revenue=0.06; s.cost_structure.sga_base_lakhs=6.0
    s.cost_structure.transport_base_lakhs=7.0; s.cost_structure.misc_base_lakhs=15.0; s.cost_structure.status=SectionStatus.COMPLETE
    s.manpower.categories=[EmployeeCategory('M',1,0.40),EmployeeCategory('O',5,0.15),EmployeeCategory('A',3,0.20),EmployeeCategory('G',1,0.12)]
    s.manpower.status=SectionStatus.COMPLETE
    s.finance_wc.debtor_days=10; s.finance_wc.creditor_days_rm=10; s.finance_wc.creditor_days_admin=30
    s.finance_wc.stock_days_rm=7; s.finance_wc.implementation_months=3; s.finance_wc.status=SectionStatus.COMPLETE
    s.tax_config.entity_type=EntityType.COMPANY; s.tax_config.status=SectionStatus.COMPLETE
    return s

STORE=make_store(); TMP=tempfile.mktemp(suffix="_p5.xlsx")
builder=WorkbookBuilder(STORE); builder.build(TMP)
layout=builder.layout; WB=load_workbook(TMP,data_only=False)
print(f"\n  Sheets: {WB.sheetnames}\n")

section("1. All 14 sheets present")
for s in ["Index","Assumption","Cost & Means","Revenue","ManPower","Depreciation","Expenses","Term Loan","W Cap","Tax","PL","BS","CFS","Ratio"]:
    assert_true(f"'{s}' exists", s in WB.sheetnames)

section("2. PL — revenue rows reference Revenue sheet")
ws_pl=WB["PL"]; PL_COL=6
for i in range(STORE.n_products):
    r=PL.SALES_ROW+i
    c=ws_pl.cell(r,PL_COL)
    assert_true(f"PL sales row {i} references Revenue", xref(c.value,"Revenue"), f"{c.value}")

assert_true("PL total revenue references Revenue",
    xref(ws_pl.cell(PL.TOTAL_REV_ROW,PL_COL).value,"Revenue"))

section("3. PL — COGS references Expenses sheet")
assert_true("PL COGS references Expenses",
    xref(ws_pl.cell(PL.COGS_ROW,PL_COL).value,"Expenses"))

section("4. PL — gross profit formula")
gp=ws_pl.cell(PL.GROSS_PROFIT_ROW,PL_COL)
assert_true("Gross profit is formula", is_formula(gp.value))
assert_true("Gross profit references total revenue row",
    str(PL.TOTAL_REV_ROW) in str(gp.value))
assert_true("Gross profit references total COGS row",
    str(PL.TOTAL_COGS_ROW) in str(gp.value))

section("5. PL — opex: depreciation from Depreciation, salary from ManPower")
dep_row=PL.OE_BASE_ROW+5
dep_cell=ws_pl.cell(dep_row,PL_COL)
assert_true("PL depreciation row references Depreciation sheet",
    xref(dep_cell.value,"Depreciation"), f"formula: {dep_cell.value}")

sal_row=PL.OE_BASE_ROW+4
sal_cell=ws_pl.cell(sal_row,PL_COL)
assert_true("PL salary row references ManPower sheet",
    xref(sal_cell.value,"ManPower"), f"formula: {sal_cell.value}")

section("6. PL — interest references Term Loan sheet")
int_cell=ws_pl.cell(PL.interest_tl_row(),PL_COL)
assert_true("PL TL interest references Term Loan",
    xref(int_cell.value,"Term Loan"), f"formula: {int_cell.value}")

section("7. PL — PBT and PAT formulas")
pbt=ws_pl.cell(PL.pbt_row(),PL_COL)
assert_true("PBT is formula", is_formula(pbt.value))
assert_true("PBT references EBIT row", str(PL.ebit_row()) in str(pbt.value))

pat=ws_pl.cell(PL.pat_row(),PL_COL)
assert_true("PAT is formula", is_formula(pat.value))
assert_true("PAT references PBT row", str(PL.pbt_row()) in str(pat.value))

section("8. PL — tax references Tax sheet")
tax_cell=ws_pl.cell(PL.current_tax_row(),PL_COL)
assert_true("PL tax row references Tax sheet",
    xref(tax_cell.value,"Tax"), f"formula: {tax_cell.value}")

section("9. PL — EBITDA adds back depreciation")
ebitda=ws_pl.cell(PL.ebitda_row(),PL_COL)
assert_true("EBITDA is formula", is_formula(ebitda.value))
assert_true("EBITDA references EBIT row", str(PL.ebit_row()) in str(ebitda.value))

section("10. BS — structure and cross-references")
ws_bs=WB["BS"]; BS_COL=3
tl_out=ws_bs.cell(BS.TOTAL_TL_ROW,BS_COL)
assert_true("BS TL outstanding references Term Loan",
    xref(tl_out.value,"Term Loan"), f"formula: {tl_out.value}")

net_block=ws_bs.cell(BS.NET_BLOCK_ROW,BS_COL)
assert_true("BS net block references Depreciation",
    xref(net_block.value,"Depreciation"), f"formula: {net_block.value}")

wc_cl=ws_bs.cell(BS.TOTAL_CL_ROW,BS_COL)
assert_true("BS creditors references W Cap",
    xref(wc_cl.value,"W Cap"), f"formula: {wc_cl.value}")

section("11. BS — balance check row")
check=ws_bs.cell(BS.BALANCE_CHECK_ROW,BS_COL)
assert_true("BS balance check is IF formula",
    is_formula(check.value) and "IF(" in str(check.value).upper())
assert_true("Balance check references TOTAL_ASSETS_ROW",
    str(BS.TOTAL_ASSETS_ROW) in str(check.value))
assert_true("Balance check references TOTAL_LIAB_ROW",
    str(BS.TOTAL_LIAB_ROW) in str(check.value))

section("12. CFS — operating activities from PL")
ws_cfs=WB["CFS"]; CFS_COL=CFS.DATA_COL_START
pbt_cfs=ws_cfs.cell(CFS.PBT_ROW,CFS_COL)
assert_true("CFS PBT references PL", xref(pbt_cfs.value,"PL"), f"formula: {pbt_cfs.value}")

depr_cfs=ws_cfs.cell(CFS.ADD_DEPR_ROW,CFS_COL)
assert_true("CFS add-back depreciation references PL",
    xref(depr_cfs.value,"PL"), f"formula: {depr_cfs.value}")

section("13. CFS — working capital changes reference W Cap")
dbtrs_cfs=ws_cfs.cell(CFS.CHG_RECEIVABLES_ROW,CFS_COL)
assert_true("CFS debtor change references W Cap",
    xref(dbtrs_cfs.value,"W Cap"), f"formula: {dbtrs_cfs.value}")

section("14. CFS — closing cash is cumulative")
close=ws_cfs.cell(CFS.CLOSING_CASH_ROW,CFS_COL)
assert_true("Closing cash is formula", is_formula(close.value))
# Year 2 opening = year 1 closing (that's where the backward reference lives)
open_y2=ws_cfs.cell(CFS.OPENING_CASH_ROW, CFS.DATA_COL_START+1)
assert_true("Year 2 opening references year 1 closing",
    is_formula(open_y2.value) and str(CFS.CLOSING_CASH_ROW) in str(open_y2.value),
    f"got: {open_y2.value}")

section("15. Ratio — DSCR uses PAT+Depr+Interest / Principal+Interest")
ws_rat=WB["Ratio"]; R_COL=RT.DATA_COL_START
num=ws_rat.cell(RT.NUMERATOR_ROW,R_COL)
assert_true("DSCR numerator references PL", xref(num.value,"PL"), f"formula: {num.value}")
den=ws_rat.cell(RT.DENOMINATOR_ROW,R_COL)
assert_true("DSCR denominator references Term Loan",
    xref(den.value,"Term Loan"), f"formula: {den.value}")
dscr=ws_rat.cell(RT.DSCR_ROW,R_COL)
assert_true("DSCR row is IFERROR formula",
    is_formula(dscr.value) and "IFERROR" in str(dscr.value).upper())

section("16. Ratio — profitability margins")
npm=ws_rat.cell(RT.PROF_NPM_ROW,R_COL)
assert_true("NPM references PL", xref(npm.value,"PL"), f"formula: {npm.value}")

section("17. Ratio — ROCE references PL and BS")
roce=ws_rat.cell(RT.ROCE_ROW,R_COL)
assert_true("ROCE references PL (EBIT)",  xref(roce.value,"PL"),  f"formula: {roce.value}")
assert_true("ROCE references BS (assets)", xref(roce.value,"BS"), f"formula: {roce.value}")

section("18. Ratio — BEP uses fixed costs / contribution")
bep=ws_rat.cell(RT.BEP_PCT_ROW,R_COL)
assert_true("BEP is IFERROR formula",
    is_formula(bep.value) and "IFERROR" in str(bep.value).upper())

section("19. File integrity — 14 sheets, 1000+ formulas")
assert_true("14 sheets built", len(WB.sheetnames)==14, f"got {len(WB.sheetnames)}")
total=sum(sum(1 for r in WB[s].iter_rows() for c in r if isinstance(c.value,str) and c.value.startswith('=')) for s in WB.sheetnames)
assert_true("Total formulas > 1000", total>1000, f"got {total}")
assert_true("File size > 30KB", os.path.getsize(TMP)>30000, f"{os.path.getsize(TMP)}")
print(f"  Formula count: {total}")

section("20. Layout maps registered for all key sheets")
for sheet in ["PL","BS","CFS"]:
    assert_true(f"{sheet} layout map registered",
        sheet in layout._map and len(layout._map[sheet])>2)

total_t=passed+failed
print(f"\n{'═'*60}\n{BOLD}  PHASE 5 TEST RESULTS{RESET}\n{'═'*60}")
print(f"  {GREEN}Passed: {passed}/{total_t}{RESET}")
if failed:
    print(f"  {RED}Failed: {failed}/{total_t}{RESET}")
    for n,r in errors: print(f"    {RED}✗ {n}{RESET}\n      {r}")
else: print(f"\n  {GREEN}{BOLD}ALL TESTS PASSED ✓{RESET}")
print(f"{'═'*60}\n")
sys.exit(0 if failed==0 else 1)
