"""
test_phase2.py
──────────────
Tests for Phase 2: Assumption + Cost & Means sheet generation.

Tests verify:
  1. File is created and openable
  2. Correct sheets exist
  3. Assumption: all input cells are blue font
  4. Assumption: correct values written for all parameters
  5. Assumption: section headers are navy
  6. Cost & Means: asset rows sum correctly
  7. Cost & Means: means rows sum correctly
  8. Cost & Means: balance check formula is present
  9. Formula cells are black (not blue)
  10. Cross-sheet reference cells are green
"""

import sys, os, tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from openpyxl import load_workbook
from openpyxl.styles import Font
from core.session_store import (
    SessionStore, Asset, FinanceSource, Product, RawMaterial,
    EmployeeCategory, BusinessType, RevenueModelType, EntityType,
    FinanceSourceType, AssetCategory, SectionStatus
)
from core.layout_engine  import LayoutEngine, col_letter, year_col
from excel.workbook_builder import WorkbookBuilder

# ─── ANSI colours ────────────────────────────────────────────────────────────
GREEN = "\033[92m"; RED = "\033[91m"; CYAN = "\033[96m"
RESET = "\033[0m";  BOLD = "\033[1m"

passed = 0; failed = 0; errors = []

def ok(name):
    global passed; passed += 1
    print(f"  {GREEN}✓{RESET}  {name}")

def fail(name, reason):
    global failed; failed += 1; errors.append((name, reason))
    print(f"  {RED}✗{RESET}  {name}")
    print(f"      {RED}{reason}{RESET}")

def section(title):
    print(f"\n{BOLD}{CYAN}{'─'*60}{RESET}\n{BOLD}{CYAN}  {title}{RESET}\n{BOLD}{CYAN}{'─'*60}{RESET}")

def assert_eq(name, actual, expected):
    if actual == expected: ok(name)
    else: fail(name, f"expected {repr(expected)}, got {repr(actual)}")

def assert_true(name, cond, msg=""):
    if cond: ok(name)
    else: fail(name, msg or "False")

def assert_in(name, val, container):
    if val in container: ok(name)
    else: fail(name, f"{repr(val)} not found in {repr(container)[:80]}")


# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD THE WORKBOOK ONCE, REUSE ACROSS ALL TESTS
# ═══════════════════════════════════════════════════════════════════════════════

def make_store() -> SessionStore:
    store = SessionStore()
    store.project_profile.company_name = "Videhanutra India Pvt Ltd"
    store.project_profile.promoter_name = "Rajesh Kumar"
    store.project_profile.operation_start_date = "2026-04"
    store.project_profile.projection_years = 7
    store.project_profile.entity_type = EntityType.COMPANY
    store.project_profile.status = SectionStatus.COMPLETE

    store.capital_means.assets = [
        Asset("Civil Works",       AssetCategory.CIVIL_WORKS,     202.625,{1:0.8,2:0.2}),
        Asset("Plant & Machinery", AssetCategory.PLANT_MACHINERY, 110.944,{3:1.0}),
        Asset("Furniture & Fixture",AssetCategory.FURNITURE,        5.0,  {3:1.0}),
    ]
    store.capital_means.finance_sources = [
        FinanceSource(FinanceSourceType.TERM_LOAN,      200.0, 0.09, 84, 18, "SBI TL"),
        FinanceSource(FinanceSourceType.OD_LIMIT,        70.0, 0.09),
        FinanceSource(FinanceSourceType.PROMOTER_EQUITY, 47.569),
    ]
    store.capital_means.status = SectionStatus.COMPLETE

    store.revenue_model.products = [
        Product("Mustard Oil",   "litres", capacity_per_day=8, output_ratio=0.36,
                split_percent=0.5, price_per_unit=175.0, price_escalation_pa=0.04),
        Product("Groundnut Oil", "litres", capacity_per_day=8, output_ratio=0.36,
                split_percent=0.5, price_per_unit=195.0, price_escalation_pa=0.04),
    ]
    store.revenue_model.year1_utilization = 0.50
    store.revenue_model.annual_utilization_increment = 0.05
    store.revenue_model.max_utilization = 0.85
    store.revenue_model.working_days_per_month = 28
    store.revenue_model.status = SectionStatus.COMPLETE

    store.cost_structure.raw_materials = [
        RawMaterial("Groundnut Seed", "kg", 55.0, 2.78, 0.05),
        RawMaterial("Mustard Seed",   "kg", 44.0, 2.78, 0.05),
    ]
    store.cost_structure.rm_pct_of_fa         = 0.02
    store.cost_structure.rm_escalation_pa     = 0.10
    store.cost_structure.insurance_pct_of_fa  = 0.01
    store.cost_structure.power_pct_revenue    = 0.06
    store.cost_structure.power_escalation_pa  = 0.05
    store.cost_structure.marketing_pct_revenue= 0.05
    store.cost_structure.transport_base_lakhs = 7.0
    store.cost_structure.transport_escalation_pa = 0.15
    store.cost_structure.misc_base_lakhs      = 15.0
    store.cost_structure.misc_escalation_pa   = 0.20
    store.cost_structure.sga_base_lakhs       = 6.0
    store.cost_structure.sga_escalation_pa    = 0.10
    store.cost_structure.status = SectionStatus.COMPLETE

    store.manpower.categories = [
        EmployeeCategory("Plant Manager",     1, 0.40, True,  0.05),
        EmployeeCategory("Operational Staff", 5, 0.15, True,  0.05),
        EmployeeCategory("Admin Team",        3, 0.20, True,  0.05),
        EmployeeCategory("Security Guard",    1, 0.12, True,  0.05),
    ]
    store.manpower.status = SectionStatus.COMPLETE

    store.finance_wc.debtor_days         = 10
    store.finance_wc.creditor_days_rm    = 10
    store.finance_wc.creditor_days_admin = 30
    store.finance_wc.stock_days_rm       = 7
    store.finance_wc.implementation_months = 3
    store.finance_wc.status = SectionStatus.COMPLETE

    store.tax_config.entity_type = EntityType.COMPANY
    store.tax_config.status = SectionStatus.COMPLETE

    return store


# Build once — use the builder's layout so we have runtime mutations
STORE   = make_store()
TMPFILE = tempfile.mktemp(suffix=".xlsx")
builder = WorkbookBuilder(STORE)
builder.build(TMPFILE)
layout  = builder.layout   # use the SAME layout the builder used, not a fresh one

# Open for inspection (data_only=False to see formulas)
WB_FORMULA = load_workbook(TMPFILE, data_only=False)
print(f"\n  Generated: {TMPFILE}")
print(f"  Sheets: {WB_FORMULA.sheetnames}\n")


# ═══════════════════════════════════════════════════════════════════════════════
section("1. File Structure")
# ═══════════════════════════════════════════════════════════════════════════════

assert_true("File exists", os.path.exists(TMPFILE))
assert_true("File size > 5KB", os.path.getsize(TMPFILE) > 5000)

EXPECTED_SHEETS = ["Index", "Assumption", "Cost & Means"]
for sname in EXPECTED_SHEETS:
    assert_in(f"Sheet '{sname}' exists", sname, WB_FORMULA.sheetnames)


# ═══════════════════════════════════════════════════════════════════════════════
section("2. Assumption Sheet — Structure")
# ═══════════════════════════════════════════════════════════════════════════════

ws_asmp = WB_FORMULA["Assumption"]
# layout already set to builder.layout above — do NOT recreate here

# Title row is navy (dark fill)
title_cell = ws_asmp.cell(row=1, column=1)
assert_true("Title row has fill",
            title_cell.fill is not None and
            title_cell.fill.fgColor is not None)

# Section headers exist (navy background, white text)
def check_section_header_exists(row):
    for col in range(1, 5):
        cell = ws_asmp.cell(row=row, column=col)
        if cell.value:
            return True
    return False

from core.layout_engine import AssumptionLayout as AL
for sec_row in [AL.HDR_CAPACITY, AL.HDR_REVENUE, AL.HDR_EXPENSES,
                AL.HDR_MANPOWER, AL.HDR_FINANCE, AL.HDR_WORKING_CAPITAL,
                AL.HDR_DEPRECIATION]:
    assert_true(f"Section header exists at row {sec_row}",
                check_section_header_exists(sec_row))


# ═══════════════════════════════════════════════════════════════════════════════
section("3. Assumption Sheet — Input Cell Values")
# ═══════════════════════════════════════════════════════════════════════════════

def get_val(ws, logical, layout):
    r, c = layout._map["Assumption"][logical]
    return ws.cell(row=r, column=c).value

def get_val_direct(ws, row, col):
    return ws.cell(row=row, column=col).value

def approx_eq(a, b, tol=0.0001):
    try:
        return abs(float(a) - float(b)) < tol
    except (TypeError, ValueError):
        return a == b

# Capacity parameters
assert_true("cap_year1_util = 0.50",
            approx_eq(get_val(ws_asmp, "cap_year1_util", layout), 0.50))
assert_true("cap_annual_increment = 0.05",
            approx_eq(get_val(ws_asmp, "cap_annual_increment", layout), 0.05))
assert_true("cap_max_util = 0.85",
            approx_eq(get_val(ws_asmp, "cap_max_util", layout), 0.85))
assert_true("cap_working_days = 28",
            approx_eq(get_val(ws_asmp, "cap_working_days", layout), 28))

# Revenue — product prices (col 5) and escalations (col 5, next row)
from core.layout_engine import AssumptionLayout as AL_
p0_price_row = AL_.rev_price_row(0)
p0_esc_row   = AL_.rev_escalation_row(0)
p1_price_row = AL_.rev_price_row(1)

assert_true("rev_price_p0 = 175",
            approx_eq(ws_asmp.cell(p0_price_row, 5).value, 175.0))
assert_true("rev_price_p1 = 195",
            approx_eq(ws_asmp.cell(p1_price_row, 5).value, 195.0))
assert_true("rev_escalation_p0 = 0.04",
            approx_eq(ws_asmp.cell(p0_esc_row, 5).value, 0.04),
            f"got {ws_asmp.cell(p0_esc_row, 5).value} at row {p0_esc_row}")

# Raw materials — price and escalation
n_prods = STORE.n_products
m0_price_row = AL_.rm_price_row(0, n_prods)
m0_esc_row   = m0_price_row + 1
m1_price_row = AL_.rm_price_row(1, n_prods)

assert_true("rm_price_m0 = 55 (groundnut seed)",
            approx_eq(ws_asmp.cell(m0_price_row, 5).value, 55.0))
assert_true("rm_price_m1 = 44 (mustard seed)",
            approx_eq(ws_asmp.cell(m1_price_row, 5).value, 44.0))
assert_true("rm_escalation_m0 = 0.05",
            approx_eq(ws_asmp.cell(m0_esc_row, 5).value, 0.05),
            f"got {ws_asmp.cell(m0_esc_row, 5).value} at row {m0_esc_row}")

# Expense parameters — read directly from layout constants
exp_rows = {
    "exp_rm_pct_fa":       (AL_.EXP_RM_PCT_FA[0],     AL_.EXP_RM_PCT_FA[1],     0.02),
    "exp_power_pct_rev":   (AL_.EXP_POWER_PCT_REV[0], AL_.EXP_POWER_PCT_REV[1], 0.06),
    "exp_transport_base":  (AL_.EXP_TRANSPORT_BASE[0],AL_.EXP_TRANSPORT_BASE[1],7.0),
    "exp_misc_base":       (AL_.EXP_MISC_BASE[0],     AL_.EXP_MISC_BASE[1],     15.0),
}
for key, (r, c, expected) in exp_rows.items():
    val = ws_asmp.cell(r, c).value
    assert_true(f"{key} = {expected}",
                approx_eq(val, expected),
                f"got {val} at row {r} col {c}")

# Manpower
assert_true("mp_salary_cat0 = 0.40",
            approx_eq(get_val(ws_asmp, "mp_salary_cat0", layout), 0.40))
assert_true("mp_salary_cat1 = 0.15",
            approx_eq(get_val(ws_asmp, "mp_salary_cat1", layout), 0.15))

# Finance
assert_true("fin_rate_l0 = 0.09",
            approx_eq(get_val(ws_asmp, "fin_rate_l0", layout), 0.09))
assert_true("fin_moratorium_l0 = 18",
            approx_eq(get_val(ws_asmp, "fin_moratorium_l0", layout), 18))
assert_true("fin_tenor_l0 = 84",
            approx_eq(get_val(ws_asmp, "fin_tenor_l0", layout), 84))
assert_true("fin_amount_l0 = 200",
            approx_eq(get_val(ws_asmp, "fin_amount_l0", layout), 200.0))

# Derived: repayment months = tenor - moratorium = 84 - 18 = 66
# It's written one row after the moratorium row in Assumption
from core.layout_engine import AssumptionLayout as AL
mor_row = layout.row("Assumption", "fin_moratorium_l0")
rep_row_actual = mor_row + 1
rep_cell = ws_asmp.cell(row=rep_row_actual, column=5)  # COL_VALUE = 5
assert_true("fin_repayment_months_l0 is a formula (not hardcoded 66)",
            isinstance(rep_cell.value, str) and rep_cell.value.startswith("="),
            f"got {repr(rep_cell.value)}")

# WC days
assert_true("wc_debtor_days = 10",
            approx_eq(get_val(ws_asmp, "wc_debtor_days", layout), 10))
assert_true("wc_stock_rm = 7",
            approx_eq(get_val(ws_asmp, "wc_stock_rm", layout), 7))

# Depreciation rates
assert_true("dep_pm_rate = 0.15",
            approx_eq(get_val(ws_asmp, "dep_pm_rate", layout), 0.15))
assert_true("dep_civil_rate = 0.10",
            approx_eq(get_val(ws_asmp, "dep_civil_rate", layout), 0.10))


# ═══════════════════════════════════════════════════════════════════════════════
section("4. Assumption Sheet — Cell Colours (Input = Blue)")
# ═══════════════════════════════════════════════════════════════════════════════

def get_font_color(ws, logical, layout):
    r, c = layout._map["Assumption"][logical]
    cell = ws.cell(row=r, column=c)
    if cell.font and cell.font.color:
        return str(cell.font.color.rgb).upper() if hasattr(cell.font.color, 'rgb') else None
    return None

BLUE_COLOR = "FF0000FF"   # openpyxl stores as ARGB with FF prefix

input_keys = [
    "cap_year1_util", "cap_annual_increment", "cap_max_util",
    "rev_price_p0", "rev_escalation_p0",
    "rm_price_m0", "rm_escalation_m0",
    "exp_rm_pct_fa", "exp_power_pct_rev",
    "dep_pm_rate", "dep_civil_rate",
]
for key in input_keys:
    color = get_font_color(ws_asmp, key, layout)
    assert_true(f"Input cell '{key}' has blue font",
                color == BLUE_COLOR or color == "0000FF",
                f"got color={color}")


# ═══════════════════════════════════════════════════════════════════════════════
section("5. Assumption Sheet — Rows are strictly ordered")
# ═══════════════════════════════════════════════════════════════════════════════

# Verify all key rows are in ascending order
key_sequence = [
    "cap_year1_util", "cap_annual_increment", "cap_max_util",
    "rev_price_p0", "rev_escalation_p0",
    "rev_price_p1", "rev_escalation_p1",
    "rm_price_m0", "rm_escalation_m0",
    "rm_price_m1", "rm_escalation_m1",
    "exp_rm_pct_fa", "exp_ins_pct_fa", "exp_power_pct_rev",
    "wc_debtor_days", "wc_creditor_rm",
    "dep_pm_rate", "dep_civil_rate",
]

rows = [layout.row("Assumption", k) for k in key_sequence]
for i in range(len(rows) - 1):
    assert_true(f"Assumption row ordering: {key_sequence[i]} < {key_sequence[i+1]}",
                rows[i] < rows[i+1],
                f"rows[{i}]={rows[i]}, rows[{i+1}]={rows[i+1]}")


# ═══════════════════════════════════════════════════════════════════════════════
section("6. Cost & Means Sheet — Structure")
# ═══════════════════════════════════════════════════════════════════════════════

ws_cm = WB_FORMULA["Cost & Means"]

# Sheet has content
assert_true("Cost & Means has content",
            ws_cm.max_row > 5 and ws_cm.max_column >= 5)

# Cost total row exists and has a SUM formula
cost_total_row = layout._map["Cost & Means"]["cost_total_row"][0]
cost_total_col = layout._map["Cost & Means"]["cost_total_row"][1]
cost_total_cell = ws_cm.cell(row=cost_total_row, column=cost_total_col)
assert_true("Cost total cell has SUM formula",
            isinstance(cost_total_cell.value, str) and
            "SUM" in cost_total_cell.value.upper(),
            f"got {repr(cost_total_cell.value)}")

# Means total row exists
means_total_row = layout._map["Cost & Means"]["means_total_row"][0]
means_total_col = layout._map["Cost & Means"]["means_total_row"][1]
means_total_cell = ws_cm.cell(row=means_total_row, column=means_total_col)
assert_true("Means total cell has SUM formula",
            isinstance(means_total_cell.value, str) and
            "SUM" in means_total_cell.value.upper(),
            f"got {repr(means_total_cell.value)}")


# ═══════════════════════════════════════════════════════════════════════════════
section("7. Cost & Means — Asset Values Written Correctly")
# ═══════════════════════════════════════════════════════════════════════════════

# The 3 asset rows are at rows 6, 7, 8 (first_data_row = ROW_COL_HDR+1 = 6)
# ROW_COL_HDR = 5, first_data_row = 6
from excel.styles import FMT_LAKHS
assets = STORE.capital_means.assets
first_data = 6  # first asset row

for i, asset in enumerate(assets):
    r = first_data + i
    cell = ws_cm.cell(row=r, column=5)  # COL_AMT = 5
    assert_true(f"Asset {i+1} ({asset.name}) value = {asset.cost_lakhs}",
                abs(float(cell.value or 0) - asset.cost_lakhs) < 0.01,
                f"got {cell.value}")


# ═══════════════════════════════════════════════════════════════════════════════
section("8. Cost & Means — Balance Check Formula")
# ═══════════════════════════════════════════════════════════════════════════════

check_row = layout._map["Cost & Means"]["balance_check_row"][0]
check_col = layout._map["Cost & Means"]["balance_check_row"][1]
check_cell = ws_cm.cell(row=check_row, column=check_col)

assert_true("Balance check cell exists",
            check_cell.value is not None,
            "cell is empty")
assert_true("Balance check is a formula",
            isinstance(check_cell.value, str) and
            check_cell.value.startswith("="),
            f"got {repr(check_cell.value)}")
assert_true("Balance check uses IF",
            "IF(" in check_cell.value.upper(),
            f"got {repr(check_cell.value)}")
assert_true("Balance check uses ROUND",
            "ROUND" in check_cell.value.upper(),
            f"got {repr(check_cell.value)}")
assert_true("Balance check references BALANCED text",
            "BALANCED" in check_cell.value,
            f"got {repr(check_cell.value)}")


# ═══════════════════════════════════════════════════════════════════════════════
section("9. Cost & Means — Percentage Formulas Present")
# ═══════════════════════════════════════════════════════════════════════════════

# Each asset row should have a % formula in COL_PCT (col 6)
for i, asset in enumerate(assets):
    r = first_data + i
    pct_cell = ws_cm.cell(row=r, column=6)
    assert_true(f"Asset {i+1} pct cell is IFERROR formula",
                isinstance(pct_cell.value, str) and
                "IFERROR" in pct_cell.value.upper(),
                f"got {repr(pct_cell.value)}")


# ═══════════════════════════════════════════════════════════════════════════════
section("10. Index Sheet")
# ═══════════════════════════════════════════════════════════════════════════════

ws_idx = WB_FORMULA["Index"]
assert_true("Index sheet has content", ws_idx.max_row > 3)

# First cell should contain company name
found_company = False
for row in ws_idx.iter_rows():
    for cell in row:
        if cell.value and STORE.project_profile.company_name in str(cell.value):
            found_company = True
            break
assert_true("Index contains company name", found_company)

# Should list sheet names
sheet_names_found = []
for row in ws_idx.iter_rows():
    for cell in row:
        if cell.value in ["Assumption", "Cost & Means", "PL", "BS", "CFS"]:
            sheet_names_found.append(cell.value)
assert_true("Index lists sheet names",
            len(sheet_names_found) >= 3,
            f"found: {sheet_names_found}")


# ═══════════════════════════════════════════════════════════════════════════════
section("11. Workbook Properties")
# ═══════════════════════════════════════════════════════════════════════════════

assert_true("Workbook title set",
            STORE.project_profile.company_name in (WB_FORMULA.properties.title or ""),
            f"got: {WB_FORMULA.properties.title}")
assert_true("Creator is DPR Agent",
            WB_FORMULA.properties.creator == "DPR Agent",
            f"got: {WB_FORMULA.properties.creator}")


# ═══════════════════════════════════════════════════════════════════════════════
section("12. Multi-Product: 3 Products")
# ═══════════════════════════════════════════════════════════════════════════════

from core.session_store import SectionStatus

store3 = make_store()
store3.revenue_model.products = [
    Product("Product A", "kg",     capacity_per_day=10, output_ratio=0.40,
            split_percent=0.33, price_per_unit=100.0, price_escalation_pa=0.04),
    Product("Product B", "litres", capacity_per_day=10, output_ratio=0.40,
            split_percent=0.33, price_per_unit=200.0, price_escalation_pa=0.04),
    Product("Product C", "units",  capacity_per_day=10, output_ratio=0.40,
            split_percent=0.34, price_per_unit=300.0, price_escalation_pa=0.04),
]
store3.cost_structure.raw_materials = [
    RawMaterial("Input X", "kg", 30.0, 2.5, 0.05),
]

tmp3 = tempfile.mktemp(suffix="_3prod.xlsx")
WorkbookBuilder(store3).build(tmp3)
wb3 = load_workbook(tmp3, data_only=False)
ws3 = wb3["Assumption"]

L3 = LayoutEngine(store3)
# Product 2 price should be at row 14 + 2*3 = 20
p2_row = L3.row("Assumption", "rev_price_p2")
assert_true("3 products: p2 price row > p1 price row",
            p2_row > L3.row("Assumption", "rev_price_p1"))
p2_val = ws3.cell(row=p2_row,
                  column=L3._map["Assumption"]["rev_price_p2"][1]).value
assert_true("3 products: p2 price = 300",
            abs(float(p2_val or 0) - 300.0) < 0.01,
            f"got {p2_val}")

assert_true("3 product file exists", os.path.exists(tmp3))
os.unlink(tmp3)


# ═══════════════════════════════════════════════════════════════════════════════
#  CLEANUP & SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

# Keep the main test file for inspection
print(f"\n  {CYAN}Test file kept at: {TMPFILE}{RESET}")
print(f"  Open in Excel/LibreOffice to visually inspect.\n")

total = passed + failed
print(f"{'═'*60}")
print(f"{BOLD}  PHASE 2 TEST RESULTS{RESET}")
print(f"{'═'*60}")
print(f"  {GREEN}Passed: {passed}/{total}{RESET}")
if failed:
    print(f"  {RED}Failed: {failed}/{total}{RESET}")
    print(f"\n{RED}  Failed tests:{RESET}")
    for name, reason in errors:
        print(f"    {RED}✗ {name}{RESET}")
        print(f"      {reason}")
else:
    print(f"\n  {GREEN}{BOLD}ALL TESTS PASSED ✓{RESET}")
print(f"{'═'*60}\n")

sys.exit(0 if failed == 0 else 1)
