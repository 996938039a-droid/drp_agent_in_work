"""
test_phase3.py
──────────────
Tests for Phase 3: Revenue + ManPower + Depreciation sheets.

Tests verify:
  1. All three sheets exist in the workbook
  2. Revenue: correct formula structure for 1, 2, 5 products
  3. Revenue: utilisation ramp uses Assumption references
  4. Revenue: total revenue row sums all product revenue rows
  5. Revenue: price escalation references Assumption
  6. ManPower: salary projections compound correctly
  7. ManPower: P&L transfer row exists
  8. Depreciation: opening WDV matches asset costs
  9. Depreciation: depreciation charge uses IF half-year rule
  10. Depreciation: closing WDV = opening + additions - charge
  11. Depreciation: net block = gross block - cumulative depreciation
  12. Depreciation: balance check row is a formula
  13. All formulas reference Assumption (no raw hardcoded constants for rates)
"""

import sys, os, tempfile
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from openpyxl import load_workbook
from openpyxl.styles import Font
from core.session_store import (
    SessionStore, Asset, FinanceSource, Product, RawMaterial,
    EmployeeCategory, EntityType, FinanceSourceType, AssetCategory,
    SectionStatus
)
from core.layout_engine import (
    LayoutEngine, RevenueLayout as RL, ManPowerLayout as ML,
    DepreciationLayout as DL, col_letter, year_col
)
from excel.workbook_builder import WorkbookBuilder

GREEN = "\033[92m"; RED = "\033[91m"; CYAN = "\033[96m"
RESET = "\033[0m";  BOLD = "\033[1m"

passed = 0; failed = 0; errors = []

def ok(name):
    global passed; passed += 1
    print(f"  {GREEN}✓{RESET}  {name}")

def fail(name, reason):
    global failed; failed += 1; errors.append((name, reason))
    print(f"  {RED}✗{RESET}  {name}")
    print(f"      {RED}{reason}{RESET}")

def section(title):
    print(f"\n{BOLD}{CYAN}{'─'*60}{RESET}\n{BOLD}{CYAN}  {title}{RESET}\n{BOLD}{CYAN}{'─'*60}{RESET}")

def assert_eq(name, actual, expected):
    if actual == expected: ok(name)
    else: fail(name, f"expected {repr(expected)}, got {repr(actual)}")

def assert_true(name, cond, msg=""):
    if cond: ok(name)
    else: fail(name, msg or "False")

def assert_contains(name, text, substr):
    if substr in str(text): ok(name)
    else: fail(name, f"'{substr}' not in '{str(text)[:80]}'")

def is_formula(val):
    return isinstance(val, str) and val.startswith("=")

def approx(a, b, tol=0.01):
    try: return abs(float(a) - float(b)) < tol
    except: return False


# ═══════════════════════════════════════════════════════════════════════════════
#  FIXTURE
# ═══════════════════════════════════════════════════════════════════════════════

def make_store(n_products=2, n_employees=4, n_asset_classes=3) -> SessionStore:
    store = SessionStore()
    store.project_profile.company_name = "Videhanutra India Pvt Ltd"
    store.project_profile.promoter_name = "Rajesh Kumar"
    store.project_profile.operation_start_date = "2026-04"
    store.project_profile.projection_years = 7
    store.project_profile.entity_type = EntityType.COMPANY
    store.project_profile.status = SectionStatus.COMPLETE

    store.capital_means.assets = [
        Asset("Civil Works",        AssetCategory.CIVIL_WORKS,     202.625, {1:0.8,2:0.2}),
        Asset("Plant & Machinery",  AssetCategory.PLANT_MACHINERY, 110.944, {3:1.0}),
        Asset("Furniture & Fixture",AssetCategory.FURNITURE,         5.0,   {3:1.0}),
    ]
    store.capital_means.finance_sources = [
        FinanceSource(FinanceSourceType.TERM_LOAN,      200.0, 0.09, 84, 18, "SBI TL"),
        FinanceSource(FinanceSourceType.OD_LIMIT,        70.0, 0.09),
        FinanceSource(FinanceSourceType.PROMOTER_EQUITY, 47.569),
    ]
    store.capital_means.status = SectionStatus.COMPLETE

    products = [
        Product("Mustard Oil",   "litres", 8, 0.36, 0.5, 175.0, 0.04),
        Product("Groundnut Oil", "litres", 8, 0.36, 0.5, 195.0, 0.04),
        Product("Product C",     "kg",     5, 0.50, 1.0, 120.0, 0.03),
        Product("Product D",     "units",  100, 1.0, 1.0, 50.0, 0.05),
        Product("Product E",     "litres", 3, 0.40, 1.0, 300.0, 0.04),
    ]
    store.revenue_model.products = products[:n_products]
    store.revenue_model.year1_utilization = 0.50
    store.revenue_model.annual_utilization_increment = 0.05
    store.revenue_model.max_utilization = 0.85
    store.revenue_model.working_days_per_month = 28
    store.revenue_model.status = SectionStatus.COMPLETE

    store.cost_structure.raw_materials = [
        RawMaterial("Groundnut Seed", "kg", 55.0, 2.78, 0.05),
        RawMaterial("Mustard Seed",   "kg", 44.0, 2.78, 0.05),
    ]
    store.cost_structure.rm_pct_of_fa = 0.02
    store.cost_structure.power_pct_revenue = 0.06
    store.cost_structure.transport_base_lakhs = 7.0
    store.cost_structure.misc_base_lakhs = 15.0
    store.cost_structure.status = SectionStatus.COMPLETE

    desigs = ["Plant Manager","Operator","Admin","Guard","Supervisor",
              "Technician","Driver","Helper"]
    cats = []
    for i in range(n_employees):
        cats.append(EmployeeCategory(
            designation=desigs[i % len(desigs)],
            count=1, monthly_salary_lakhs=0.40 - i*0.03,
            is_fixed=True, annual_increment_pa=0.05
        ))
    store.manpower.categories = cats
    store.manpower.status = SectionStatus.COMPLETE

    store.finance_wc.debtor_days = 10
    store.finance_wc.creditor_days_rm = 10
    store.finance_wc.creditor_days_admin = 30
    store.finance_wc.stock_days_rm = 7
    store.finance_wc.implementation_months = 3
    store.finance_wc.status = SectionStatus.COMPLETE

    store.tax_config.entity_type = EntityType.COMPANY
    store.tax_config.status = SectionStatus.COMPLETE
    return store


# Build the main test workbook (2 products, 4 employees, 3 asset classes)
STORE   = make_store(n_products=2, n_employees=4)
TMPFILE = tempfile.mktemp(suffix="_phase3.xlsx")
builder = WorkbookBuilder(STORE)
builder.build(TMPFILE)
layout  = builder.layout
WB      = load_workbook(TMPFILE, data_only=False)
print(f"\n  Generated: {TMPFILE}")
print(f"  Sheets: {WB.sheetnames}\n")


# ═══════════════════════════════════════════════════════════════════════════════
section("1. All sheets present")
# ═══════════════════════════════════════════════════════════════════════════════

for sheet in ["Index", "Assumption", "Cost & Means",
              "Revenue", "ManPower", "Depreciation"]:
    assert_true(f"Sheet '{sheet}' exists", sheet in WB.sheetnames)


# ═══════════════════════════════════════════════════════════════════════════════
section("2. Revenue — sheet structure")
# ═══════════════════════════════════════════════════════════════════════════════

ws_rev = WB["Revenue"]

# Has content
assert_true("Revenue has content", ws_rev.max_row > 10)

# Row 4 has date formulas in year columns
for yr in range(1, STORE.n_years + 1):
    col = year_col(yr)
    cell = ws_rev.cell(row=4, column=col)
    assert_true(f"Revenue row 4 year {yr} has date formula",
                cell.value is not None,
                f"cell is empty at col {col}")

# Row 5 — months = 12
for yr in range(1, STORE.n_years + 1):
    cell = ws_rev.cell(row=5, column=year_col(yr))
    assert_true(f"Months row yr{yr} formula",
                is_formula(cell.value),
                f"got {repr(cell.value)}")

# Row 10 — utilisation
yr1_util = ws_rev.cell(row=10, column=year_col(1))
yr2_util = ws_rev.cell(row=10, column=year_col(2))
assert_true("Yr1 util references Assumption",
            is_formula(yr1_util.value) and "Assumption" in str(yr1_util.value),
            f"got {repr(yr1_util.value)}")
assert_true("Yr2 util has IF formula for ramp",
            is_formula(yr2_util.value) and "IF(" in str(yr2_util.value).upper(),
            f"got {repr(yr2_util.value)}")


# ═══════════════════════════════════════════════════════════════════════════════
section("3. Revenue — product blocks (2 products)")
# ═══════════════════════════════════════════════════════════════════════════════

n_prods = STORE.n_products  # 2

for i in range(n_prods):
    vol_row = RL.prod_volume_row(i)
    lit_row = RL.prod_liters_row(i)
    spl_row = RL.prod_split_row(i)
    prc_row = RL.prod_price_row(i)
    rev_row = RL.prod_revenue_row(i)

    prod = STORE.revenue_model.products[i]

    # Volume row — formula multiplying rows 5, 6, capacity, yield, split, util
    vol_y1 = ws_rev.cell(row=vol_row, column=year_col(1))
    assert_true(f"Product {i} volume row is formula",
                is_formula(vol_y1.value),
                f"got {repr(vol_y1.value)}")
    # Must reference row 5 (months), row 6 (days), row 10 (utilisation)
    formula_str = str(vol_y1.value)
    assert_true(f"Product {i} volume references months (row5)",
                "5" in formula_str,
                f"formula: {formula_str}")
    assert_true(f"Product {i} volume references util (row10)",
                "10" in formula_str,
                f"formula: {formula_str}")

    # Price row Year 1 — references Assumption
    prc_y1 = ws_rev.cell(row=prc_row, column=year_col(1))
    assert_true(f"Product {i} price yr1 references Assumption",
                is_formula(prc_y1.value) and "Assumption" in str(prc_y1.value),
                f"got {repr(prc_y1.value)}")

    # Price row Year 2 — escalation formula with prior year
    prc_y2 = ws_rev.cell(row=prc_row, column=year_col(2))
    assert_true(f"Product {i} price yr2 escalates from yr1",
                is_formula(prc_y2.value) and "Assumption" in str(prc_y2.value),
                f"got {repr(prc_y2.value)}")

    # Revenue row — (split_row × price_row) / 100000
    rev_y1 = ws_rev.cell(row=rev_row, column=year_col(1))
    assert_true(f"Product {i} revenue is formula",
                is_formula(rev_y1.value),
                f"got {repr(rev_y1.value)}")
    assert_true(f"Product {i} revenue divides by 100000",
                "100000" in str(rev_y1.value),
                f"formula: {rev_y1.value}")

# Total revenue row
total_row = RL.total_revenue_row(n_prods)
total_y1  = ws_rev.cell(row=total_row, column=year_col(1))
assert_true("Total revenue row is formula",
            is_formula(total_y1.value),
            f"got {repr(total_y1.value)}")
# Must reference all product revenue rows
for i in range(n_prods):
    rev_row_i = RL.prod_revenue_row(i)
    assert_true(f"Total revenue references product {i} rev row",
                str(rev_row_i) in str(total_y1.value),
                f"formula: {total_y1.value}")


# ═══════════════════════════════════════════════════════════════════════════════
section("4. Revenue — utilisation ramp correctness")
# ═══════════════════════════════════════════════════════════════════════════════

# The IF formula structure: =IF(prev<max, prev+increment, prev)
for yr in range(2, STORE.n_years + 1):
    cell = ws_rev.cell(row=10, column=year_col(yr))
    formula = str(cell.value)
    assert_true(f"Util yr{yr} has IF ramp formula",
                "IF(" in formula.upper(),
                f"got {formula[:60]}")
    assert_true(f"Util yr{yr} references Assumption max",
                "Assumption" in formula,
                f"got {formula[:60]}")


# ═══════════════════════════════════════════════════════════════════════════════
section("5. Revenue — 5 product dynamic row structure")
# ═══════════════════════════════════════════════════════════════════════════════

store5 = make_store(n_products=5)
tmp5   = tempfile.mktemp(suffix="_5prod.xlsx")
WorkbookBuilder(store5).build(tmp5)
wb5    = load_workbook(tmp5, data_only=False)
ws_r5  = wb5["Revenue"]

# Product 4 (5th) volume row
vol_p4 = RL.prod_volume_row(4)
rev_p4 = RL.prod_revenue_row(4)
total_5= RL.total_revenue_row(5)

assert_true("5-prod: product 4 volume row has formula",
            is_formula(ws_r5.cell(vol_p4, year_col(1)).value),
            f"got {ws_r5.cell(vol_p4, year_col(1)).value}")
assert_true("5-prod: product 4 revenue row has formula",
            is_formula(ws_r5.cell(rev_p4, year_col(1)).value),
            f"got {ws_r5.cell(rev_p4, year_col(1)).value}")
assert_true("5-prod: total revenue row has formula",
            is_formula(ws_r5.cell(total_5, year_col(1)).value),
            f"got {ws_r5.cell(total_5, year_col(1)).value}")
assert_true("5-prod: total references product 4 rev row",
            str(rev_p4) in str(ws_r5.cell(total_5, year_col(1)).value))

os.unlink(tmp5)


# ═══════════════════════════════════════════════════════════════════════════════
section("6. ManPower — sheet structure")
# ═══════════════════════════════════════════════════════════════════════════════

ws_mp = WB["ManPower"]

assert_true("ManPower has content", ws_mp.max_row > 8)

# Employee rows exist
for i in range(STORE.n_employee_categories):
    r = ML.emp_row(i)
    label_cell = ws_mp.cell(row=r, column=2)
    assert_true(f"Employee {i} row has designation label",
                label_cell.value is not None,
                f"empty at row {r}")

# Count references Assumption
count_cell = ws_mp.cell(row=ML.emp_row(0), column=3)
assert_true("Employee count references Assumption",
            is_formula(count_cell.value) and "Assumption" in str(count_cell.value),
            f"got {repr(count_cell.value)}")

# Monthly salary references Assumption
sal_cell = ws_mp.cell(row=ML.emp_row(0), column=4)
assert_true("Monthly salary references Assumption",
            is_formula(sal_cell.value) and "Assumption" in str(sal_cell.value),
            f"got {repr(sal_cell.value)}")


# ═══════════════════════════════════════════════════════════════════════════════
section("7. ManPower — annual salary projections")
# ═══════════════════════════════════════════════════════════════════════════════

# The salary projection row
sal_proj_row = builder.wb["ManPower"]  # get from built workbook

# Find the projection row by scanning for formula with increment reference
found_proj = False
for row in ws_mp.iter_rows():
    for cell in row:
        if (cell.value and is_formula(cell.value) and
                "Assumption" in str(cell.value) and
                "(1+" in str(cell.value)):
            found_proj = True
            break
    if found_proj:
        break

assert_true("ManPower has salary escalation formula referencing Assumption",
            found_proj)

# Year 2+ formulas compound from year 1
# Find the salary projection row (scans for first row with compound formula)
proj_row = None
for r in range(ML.BASE_ROW + STORE.n_employee_categories + 2,
               ML.BASE_ROW + STORE.n_employee_categories + 20):
    cell_y2 = ws_mp.cell(row=r, column=year_col(2))
    if (cell_y2.value and is_formula(cell_y2.value) and
            "(1+" in str(cell_y2.value)):
        proj_row = r
        break

assert_true("ManPower projection row found", proj_row is not None,
            "No row with compounding formula found")

if proj_row:
    # Year 1 must reference base year total
    y1_cell = ws_mp.cell(row=proj_row, column=year_col(1))
    assert_true("Year 1 salary references base total",
                is_formula(y1_cell.value),
                f"got {repr(y1_cell.value)}")

    # Year 7 must be formula
    y7_cell = ws_mp.cell(row=proj_row, column=year_col(7))
    assert_true("Year 7 salary is formula",
                is_formula(y7_cell.value),
                f"got {repr(y7_cell.value)}")


# ═══════════════════════════════════════════════════════════════════════════════
section("8. ManPower — P&L transfer row")
# ═══════════════════════════════════════════════════════════════════════════════

# The P&L transfer row should be a formula row near the bottom of the sheet
pl_transfer_present = False
for row in ws_mp.iter_rows(
        min_row=ML.BASE_ROW + STORE.n_employee_categories + 5,
        max_row=ws_mp.max_row):
    for cell in row:
        if (cell.value and is_formula(cell.value) and
                "12" in str(cell.value) and
                cell.column == year_col(1)):
            pl_transfer_present = True
            break
    if pl_transfer_present:
        break

assert_true("ManPower P&L transfer row exists with formula", pl_transfer_present)

# Layout map has pl_salary_row registered
assert_true("ManPower pl_salary_row registered in layout",
            "pl_salary_row" in layout._map.get("ManPower", {}))


# ═══════════════════════════════════════════════════════════════════════════════
section("9. Depreciation — sheet structure and asset blocks")
# ═══════════════════════════════════════════════════════════════════════════════

ws_dep = WB["Depreciation"]

assert_true("Depreciation has content", ws_dep.max_row > 10)

# Months row
months_cell = ws_dep.cell(row=DL.MONTHS_ROW, column=year_col(1))
assert_true("Depreciation months row has formula",
            is_formula(months_cell.value),
            f"got {repr(months_cell.value)}")

# Check all 3 asset classes have opening/charge/closing formulas
from excel.sheet_depreciation import DEPR_RATE_MAP
n_classes = len(set(a.category for a in STORE.capital_means.assets))

for i in range(n_classes):
    opening_row = DL.opening_row(i)
    charge_row  = DL.charge_row(i)
    closing_row = DL.closing_row(i)

    # Opening Year 1 — must be a number or formula
    open_y1 = ws_dep.cell(row=opening_row, column=year_col(1))
    assert_true(f"Asset class {i} opening yr1 has value",
                open_y1.value is not None,
                f"empty at row {opening_row}")

    # Opening Year 2 — must reference closing of Year 1
    open_y2 = ws_dep.cell(row=opening_row, column=year_col(2))
    assert_true(f"Asset class {i} opening yr2 references yr1 closing",
                is_formula(open_y2.value) and str(closing_row) in str(open_y2.value),
                f"got {repr(open_y2.value)}")

    # Charge row — must have IF formula (half-year rule)
    charge_y1 = ws_dep.cell(row=charge_row, column=year_col(1))
    assert_true(f"Asset class {i} charge has IF formula",
                is_formula(charge_y1.value) and "IF(" in str(charge_y1.value).upper(),
                f"got {repr(charge_y1.value)}")

    # Charge references Assumption depreciation rate
    assert_true(f"Asset class {i} charge references Assumption rate",
                "Assumption" in str(charge_y1.value),
                f"formula: {charge_y1.value}")

    # Closing = opening + additions - charge
    close_y1 = ws_dep.cell(row=closing_row, column=year_col(1))
    assert_true(f"Asset class {i} closing is formula",
                is_formula(close_y1.value),
                f"got {repr(close_y1.value)}")
    formula_str = str(close_y1.value)
    assert_true(f"Asset class {i} closing references opening row",
                str(opening_row) in formula_str,
                f"formula: {formula_str}")
    assert_true(f"Asset class {i} closing references charge row",
                str(charge_row) in formula_str,
                f"formula: {formula_str}")


# ═══════════════════════════════════════════════════════════════════════════════
section("10. Depreciation — summary block")
# ═══════════════════════════════════════════════════════════════════════════════

gb_row = DL.gross_block_row(n_classes)
cd_row = DL.cumul_depr_row(n_classes)
nb_row = DL.net_block_row(n_classes)

# Gross block Year 1
gb_y1 = ws_dep.cell(row=gb_row, column=year_col(1))
assert_true("Gross block yr1 is formula",
            is_formula(gb_y1.value),
            f"got {repr(gb_y1.value)}")

# Cumulative depr Year 1
cd_y1 = ws_dep.cell(row=cd_row, column=year_col(1))
assert_true("Cumulative depr yr1 is formula",
            is_formula(cd_y1.value),
            f"got {repr(cd_y1.value)}")

# Cumulative depr Year 2 references Year 1 cumulative
cd_y2 = ws_dep.cell(row=cd_row, column=year_col(2))
assert_true("Cumulative depr yr2 references yr1",
            is_formula(cd_y2.value) and str(cd_row) in str(cd_y2.value),
            f"got {repr(cd_y2.value)}")

# Net block = gross - cumulative
nb_y1 = ws_dep.cell(row=nb_row, column=year_col(1))
assert_true("Net block yr1 is formula",
            is_formula(nb_y1.value),
            f"got {repr(nb_y1.value)}")
assert_true("Net block references gross block row",
            str(gb_row) in str(nb_y1.value),
            f"formula: {nb_y1.value}")
assert_true("Net block references cumulative depr row",
            str(cd_row) in str(nb_y1.value),
            f"formula: {nb_y1.value}")

# Net block registered in layout
assert_true("net_block registered in Depreciation layout",
            "net_block" in layout._map.get("Depreciation", {}))


# ═══════════════════════════════════════════════════════════════════════════════
section("11. Depreciation — balance check")
# ═══════════════════════════════════════════════════════════════════════════════

check_row = DL.check_row(n_classes)
check_y1  = ws_dep.cell(row=check_row, column=year_col(1))
assert_true("Balance check row has formula",
            is_formula(check_y1.value),
            f"got {repr(check_y1.value)}")
assert_true("Balance check uses IF for OK/fail",
            "IF(" in str(check_y1.value).upper(),
            f"formula: {check_y1.value}")
assert_true("Balance check references net block row",
            str(nb_row) in str(check_y1.value),
            f"formula: {check_y1.value}")


# ═══════════════════════════════════════════════════════════════════════════════
section("12. Depreciation — WDV declines each year (sanity check)")
# ═══════════════════════════════════════════════════════════════════════════════

# We can't compute values without recalc, but we CAN verify the formula
# structure ensures decline: closing < opening (depreciation > 0)
# Verify by checking all charge formulas include the rate × opening product
for i in range(n_classes):
    charge_y1 = ws_dep.cell(row=DL.charge_row(i), column=year_col(1))
    # Charge must reference opening row (not just a fixed number)
    assert_true(f"Asset class {i} charge references opening row",
                str(DL.opening_row(i)) in str(charge_y1.value),
                f"formula: {charge_y1.value}")


# ═══════════════════════════════════════════════════════════════════════════════
section("13. Cross-sheet — no hardcoded rate constants in Revenue formulas")
# ═══════════════════════════════════════════════════════════════════════════════

# All utilisation-related formulas in Revenue must reference Assumption
# (not contain raw rate like 0.05 hardcoded)
# Price escalation must reference Assumption for year 2+
for i in range(STORE.n_products):
    prc_row = RL.prod_price_row(i)
    for yr in range(1, STORE.n_years + 1):
        cell = ws_rev.cell(row=prc_row, column=year_col(yr))
        if yr >= 1:
            assert_true(f"Product {i} price yr{yr} references Assumption",
                        "Assumption" in str(cell.value),
                        f"formula: {cell.value}")


# ═══════════════════════════════════════════════════════════════════════════════
section("14. File integrity — workbook opens, all sheets have data")
# ═══════════════════════════════════════════════════════════════════════════════

file_size = os.path.getsize(TMPFILE)
assert_true("File size > 15KB (all 6 sheets written)",
            file_size > 15000,
            f"got {file_size} bytes")

for sheet_name in ["Revenue", "ManPower", "Depreciation"]:
    ws = WB[sheet_name]
    assert_true(f"{sheet_name} max_row > 5",
                ws.max_row > 5,
                f"max_row = {ws.max_row}")
    assert_true(f"{sheet_name} max_column >= 4",
                ws.max_column >= 4,
                f"max_column = {ws.max_column}")


# ═══════════════════════════════════════════════════════════════════════════════
#  SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

print(f"\n  {CYAN}Test file: {TMPFILE}{RESET}")
total = passed + failed
print(f"\n{'═'*60}")
print(f"{BOLD}  PHASE 3 TEST RESULTS{RESET}")
print(f"{'═'*60}")
print(f"  {GREEN}Passed: {passed}/{total}{RESET}")
if failed:
    print(f"  {RED}Failed: {failed}/{total}{RESET}")
    print(f"\n{RED}  Failed tests:{RESET}")
    for name, reason in errors:
        print(f"    {RED}✗ {name}{RESET}")
        print(f"      {reason}")
else:
    print(f"\n  {GREEN}{BOLD}ALL TESTS PASSED ✓{RESET}")
print(f"{'═'*60}\n")

sys.exit(0 if failed == 0 else 1)
